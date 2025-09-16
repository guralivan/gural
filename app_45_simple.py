# -*- coding: utf-8 -*-
import pandas as pd
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime, timedelta
import json
import os
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Настройка страницы
st.set_page_config(page_title="Анализ 45.xlsx", layout="wide")

st.title("📊 Анализ данных из файла 45.xlsx")

# Функции для работы с кешем данных
def save_data_cache(df, cache_file="data_cache.csv"):
    """Сохраняет данные в кеш"""
    try:
        df.to_csv(cache_file, index=False, encoding='utf-8-sig')
        return True
    except Exception as e:
        st.error(f"Ошибка сохранения кеша: {e}")
        return False

def load_data_cache(cache_file="data_cache.csv"):
    """Загружает данные из кеша"""
    try:
        if os.path.exists(cache_file):
            df = pd.read_csv(cache_file)
            df['Дата'] = pd.to_datetime(df['Дата'])
            return df
        return None
    except Exception as e:
        st.error(f"Ошибка загрузки кеша: {e}")
        return None

def merge_new_data(existing_df, new_df):
    """Объединяет существующие данные с новыми, добавляя только недостающие записи"""
    if existing_df is None or existing_df.empty:
        return new_df
    
    # Объединяем данные
    combined_df = pd.concat([existing_df, new_df], ignore_index=True)
    
    # Удаляем дубликаты по ключевым полям (Дата, Артикул WB, Артикул продавца)
    key_columns = ['Дата', 'Артикул WB', 'Артикул продавца']
    if all(col in combined_df.columns for col in key_columns):
        combined_df = combined_df.drop_duplicates(subset=key_columns, keep='last')
    
    # Сортируем по дате
    combined_df = combined_df.sort_values('Дата')
    
    return combined_df

def get_cache_info(cache_file="data_cache.csv"):
    """Получает информацию о кеше"""
    if os.path.exists(cache_file):
        try:
            df = pd.read_csv(cache_file)
            df['Дата'] = pd.to_datetime(df['Дата'])
            return {
                'exists': True,
                'records': len(df),
                'start_date': df['Дата'].min().strftime('%d.%m.%Y'),
                'end_date': df['Дата'].max().strftime('%d.%m.%Y'),
                'years': sorted(df['Дата'].dt.year.unique())
            }
        except:
            return {'exists': False}
    return {'exists': False}

def process_uploaded_excel_file(file):
    """Обрабатывает загруженный Excel файл"""
    try:
        # Пытаемся прочитать файл
        df = pd.read_excel(file, sheet_name='Товары', header=1)
        
        # Проверяем структуру файла
        required_columns = ['Дата', 'Артикул WB', 'Артикул продавца', 'Заказали, шт', 'Выкупили, шт']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            return {
                'success': False,
                'error': f"Неверная структура файла. Отсутствуют колонки: {', '.join(missing_columns)}"
            }
        
        # Преобразуем данные
        df['Дата'] = pd.to_datetime(df['Дата'])
        
        # Преобразуем числовые столбцы
        numeric_cols = ['Заказали, шт', 'Выкупили, шт', 'Выкупили на сумму, ₽', 
                       'Переходы в карточку', 'Положили в корзину', 'Процент выкупа',
                       'Заказали на сумму, ₽']
        for col in numeric_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
        
        # Добавляем год для сравнения
        df['Год'] = df['Дата'].dt.year
        df['Месяц'] = df['Дата'].dt.month
        df['Неделя'] = df['Дата'].dt.isocalendar().week
        df['Месяц_название'] = df['Дата'].dt.strftime('%B')
        df['День_недели'] = df['Дата'].dt.strftime('%A')
        df['День_месяца'] = df['Дата'].dt.day
        df['День_года'] = df['Дата'].dt.strftime('%m-%d')
        
        return {
            'success': True,
            'data': df,
            'records': len(df),
            'period': f"{df['Дата'].min().strftime('%d.%m.%Y')} - {df['Дата'].max().strftime('%d.%m.%Y')}",
            'years': sorted(df['Дата'].dt.year.unique())
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': f"Ошибка обработки файла: {str(e)}"
        }

def merge_uploaded_files_to_cache(uploaded_files):
    """Объединяет загруженные файлы с кешем"""
    if not uploaded_files:
        return None
    
    # Загружаем существующий кеш
    cached_df = load_data_cache()
    
    # Обрабатываем каждый файл
    processed_files = []
    total_new_records = 0
    
    for file in uploaded_files:
        result = process_uploaded_excel_file(file)
        if result['success']:
            processed_files.append({
                'name': file.name,
                'data': result['data'],
                'records': result['records'],
                'period': result['period'],
                'years': result['years']
            })
        else:
            st.error(f"❌ Ошибка в файле {file.name}: {result['error']}")
    
    if not processed_files:
        st.error("❌ Не удалось обработать ни одного файла")
        return None
    
    # Объединяем все данные
    all_data = []
    if cached_df is not None:
        all_data.append(cached_df)
    
    for file_info in processed_files:
        all_data.append(file_info['data'])
    
    # Объединяем и удаляем дубликаты
    combined_df = pd.concat(all_data, ignore_index=True)
    
    # Удаляем дубликаты по ключевым полям
    key_columns = ['Дата', 'Артикул WB', 'Артикул продавца']
    if all(col in combined_df.columns for col in key_columns):
        before_dedup = len(combined_df)
        combined_df = combined_df.drop_duplicates(subset=key_columns, keep='last')
        after_dedup = len(combined_df)
        total_new_records = before_dedup - after_dedup
    
    # Сортируем по дате
    combined_df = combined_df.sort_values('Дата')
    
    # Сохраняем обновленный кеш
    if save_data_cache(combined_df):
        return {
            'success': True,
            'combined_data': combined_df,
            'processed_files': processed_files,
            'total_records': len(combined_df),
            'new_records': total_new_records
        }
    else:
        return None

def create_full_report(df):
    """Создает полный отчет с аналитикой"""
    try:
        # Основные метрики
        total_orders = df['Заказали, шт'].sum()
        total_sales = df['Выкупили, шт'].sum()
        total_revenue = df['Выкупили на сумму, ₽'].sum()
        total_orders_amount = df['Заказали на сумму, ₽'].sum()
        
        # Конверсия
        conversion_rate = (total_sales / total_orders * 100) if total_orders > 0 else 0
        
        # Период данных
        start_date = df['Дата'].min()
        end_date = df['Дата'].max()
        
        # Анализ по товарам
        product_analysis = df.groupby('Артикул продавца').agg({
            'Заказали, шт': 'sum',
            'Выкупили, шт': 'sum',
            'Выкупили на сумму, ₽': 'sum',
            'Заказали на сумму, ₽': 'sum',
            'Переходы в карточку': 'sum',
            'Положили в корзину': 'sum'
        }).reset_index()
        
        # Добавляем конверсию по товарам
        product_analysis['Конверсия, %'] = (product_analysis['Выкупили, шт'] / product_analysis['Заказали, шт'] * 100).round(2)
        product_analysis = product_analysis.fillna(0)
        
        # Анализ по месяцам
        monthly_analysis = df.groupby(['Год', 'Месяц']).agg({
            'Заказали, шт': 'sum',
            'Выкупили, шт': 'sum',
            'Выкупили на сумму, ₽': 'sum',
            'Заказали на сумму, ₽': 'sum'
        }).reset_index()
        
        # Анализ по неделям
        weekly_analysis = df.groupby(['Год', 'Неделя']).agg({
            'Заказали, шт': 'sum',
            'Выкупили, шт': 'sum',
            'Выкупили на сумму, ₽': 'sum',
            'Заказали на сумму, ₽': 'sum'
        }).reset_index()
        
        # Топ товары
        top_products = product_analysis.nlargest(10, 'Выкупили на сумму, ₽')
        
        # Статистика по дням недели
        daily_stats = df.groupby('День_недели').agg({
            'Заказали, шт': 'mean',
            'Выкупили, шт': 'mean',
            'Выкупили на сумму, ₽': 'mean'
        }).reset_index()
        
        return {
            'summary': {
                'total_orders': total_orders,
                'total_sales': total_sales,
                'total_revenue': total_revenue,
                'total_orders_amount': total_orders_amount,
                'conversion_rate': conversion_rate,
                'start_date': start_date,
                'end_date': end_date,
                'total_records': len(df),
                'unique_products': df['Артикул продавца'].nunique()
            },
            'product_analysis': product_analysis,
            'monthly_analysis': monthly_analysis,
            'weekly_analysis': weekly_analysis,
            'top_products': top_products,
            'daily_stats': daily_stats,
            'raw_data': df
        }
    except Exception as e:
        st.error(f"Ошибка создания отчета: {e}")
        return None

def create_excel_report(report_data):
    """Создает Excel отчет с несколькими листами"""
    try:
        wb = Workbook()
        
        # Удаляем стандартный лист
        wb.remove(wb.active)
        
        # Стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 1. Сводка
        ws_summary = wb.create_sheet("📊 Сводка")
        summary_data = [
            ["Показатель", "Значение"],
            ["Общее количество заказов", f"{report_data['summary']['total_orders']:,}"],
            ["Общее количество выкупов", f"{report_data['summary']['total_sales']:,}"],
            ["Общая выручка", f"{report_data['summary']['total_revenue']:,.2f} ₽"],
            ["Общая сумма заказов", f"{report_data['summary']['total_orders_amount']:,.2f} ₽"],
            ["Конверсия", f"{report_data['summary']['conversion_rate']:.2f}%"],
            ["Период данных", f"{report_data['summary']['start_date'].strftime('%d.%m.%Y')} - {report_data['summary']['end_date'].strftime('%d.%m.%Y')}"],
            ["Всего записей", f"{report_data['summary']['total_records']:,}"],
            ["Уникальных товаров", f"{report_data['summary']['unique_products']:,}"]
        ]
        
        for row in summary_data:
            ws_summary.append(row)
        
        # Применяем стили к заголовкам
        for cell in ws_summary[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # 2. Анализ по товарам
        ws_products = wb.create_sheet("📦 Товары")
        for r in dataframe_to_rows(report_data['product_analysis'], index=False, header=True):
            ws_products.append(r)
        
        # Применяем стили
        for cell in ws_products[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # 3. Топ товары
        ws_top = wb.create_sheet("🏆 Топ товары")
        for r in dataframe_to_rows(report_data['top_products'], index=False, header=True):
            ws_top.append(r)
        
        # Применяем стили
        for cell in ws_top[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # 4. Анализ по месяцам
        ws_monthly = wb.create_sheet("📅 По месяцам")
        for r in dataframe_to_rows(report_data['monthly_analysis'], index=False, header=True):
            ws_monthly.append(r)
        
        # Применяем стили
        for cell in ws_monthly[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # 5. Анализ по неделям
        ws_weekly = wb.create_sheet("📆 По неделям")
        for r in dataframe_to_rows(report_data['weekly_analysis'], index=False, header=True):
            ws_weekly.append(r)
        
        # Применяем стили
        for cell in ws_weekly[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # 6. Статистика по дням недели
        ws_daily = wb.create_sheet("📊 По дням недели")
        for r in dataframe_to_rows(report_data['daily_stats'], index=False, header=True):
            ws_daily.append(r)
        
        # Применяем стили
        for cell in ws_daily[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # 7. Исходные данные
        ws_raw = wb.create_sheet("📋 Исходные данные")
        # Ограничиваем количество строк для производительности
        raw_data_sample = report_data['raw_data'].head(10000)  # Первые 10k записей
        for r in dataframe_to_rows(raw_data_sample, index=False, header=True):
            ws_raw.append(r)
        
        # Применяем стили
        for cell in ws_raw[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Автоподбор ширины колонок
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем в буфер
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
        
    except Exception as e:
        st.error(f"Ошибка создания Excel отчета: {e}")
        return None

# Загрузка данных
@st.cache_data
def load_data():
    try:
        df = pd.read_excel('45.xlsx', sheet_name='Товары', header=1)
        df['Дата'] = pd.to_datetime(df['Дата'])
        
        # Преобразуем числовые столбцы
        numeric_cols = ['Заказали, шт', 'Выкупили, шт', 'Выкупили на сумму, ₽', 
                       'Переходы в карточку', 'Положили в корзину', 'Процент выкупа',
                       'Заказали на сумму, ₽']
        for col in numeric_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
        
        # Добавляем год для сравнения
        df['Год'] = df['Дата'].dt.year
        df['Месяц'] = df['Дата'].dt.month
        df['Неделя'] = df['Дата'].dt.isocalendar().week
        df['Месяц_название'] = df['Дата'].dt.strftime('%B')
        df['День_недели'] = df['Дата'].dt.strftime('%A')
        df['День_месяца'] = df['Дата'].dt.day
        df['День_года'] = df['Дата'].dt.strftime('%m-%d')  # Формат MM-DD для сравнения дней
        
        return df
    except Exception as e:
        st.error(f"Ошибка загрузки: {e}")
        return None

def load_data_with_cache():
    """Загружает данные с учетом кеша"""
    # Проверяем информацию о кеше
    cache_info = get_cache_info()
    
    # Загружаем новые данные из файла
    new_df = load_data()
    if new_df is None:
        return None
    
    # Если кеш существует, объединяем данные
    if cache_info['exists']:
        cached_df = load_data_cache()
        if cached_df is not None:
            # Объединяем данные
            combined_df = merge_new_data(cached_df, new_df)
            
            # Показываем информацию о слиянии
            new_records = len(combined_df) - len(cached_df)
            if new_records > 0:
                st.success(f"✅ Добавлено {new_records} новых записей в кеш")
            else:
                st.info("ℹ️ Новых данных не найдено")
            
            # Сохраняем обновленный кеш
            save_data_cache(combined_df)
            return combined_df
    
    # Если кеша нет, сохраняем новые данные
    save_data_cache(new_df)
    st.success(f"✅ Создан новый кеш с {len(new_df)} записями")
    return new_df

# Загрузка данных
df = load_data_with_cache()

if df is None:
    st.error("Не удалось загрузить данные из файла 45.xlsx")
    st.stop()

# Обработка загруженных файлов
if 'process_uploaded_files' in st.session_state and st.session_state.process_uploaded_files:
    # Получаем загруженные файлы из session state
    if 'uploaded_files' in st.session_state:
        st.info("🔄 Обрабатываю загруженные файлы...")
        
        # Объединяем файлы с кешем
        result = merge_uploaded_files_to_cache(st.session_state.uploaded_files)
        
        if result and result['success']:
            st.success(f"✅ Файлы успешно обработаны!")
            st.write(f"📊 **Всего записей в кеше:** {result['total_records']:,}")
            
            if result['new_records'] > 0:
                st.write(f"🆕 **Новых записей добавлено:** {result['new_records']:,}")
            
            st.write("📁 **Обработанные файлы:**")
            for file_info in result['processed_files']:
                st.write(f"  - {file_info['name']}: {file_info['records']:,} записей, период: {file_info['period']}")
            
            # Обновляем основные данные
            df = result['combined_data']
            
            # Очищаем флаг обработки
            st.session_state.process_uploaded_files = False
            del st.session_state.uploaded_files
            
            st.rerun()
        else:
            st.error("❌ Ошибка при обработке файлов")
            st.session_state.process_uploaded_files = False

# Информация о данных
st.sidebar.header("📋 Информация о данных")
st.sidebar.write(f"Записей: {len(df):,}")
st.sidebar.write(f"Товаров: {df['Артикул продавца'].nunique()}")
st.sidebar.write(f"Артикулов WB: {df['Артикул WB'].nunique()}")
st.sidebar.write(f"Период: {df['Дата'].min().strftime('%d.%m.%Y')} - {df['Дата'].max().strftime('%d.%m.%Y')}")

# Информация о текущей неделе
st.sidebar.header("📅 Текущая неделя")
current_date = datetime.now()
current_week = current_date.isocalendar().week
current_year = current_date.year
current_month = current_date.strftime('%B')
current_day = current_date.strftime('%A')

st.sidebar.info(f"**Сегодня:** {current_date.strftime('%d.%m.%Y')}")
st.sidebar.info(f"**День недели:** {current_day}")
st.sidebar.info(f"**Неделя:** {current_week}")
st.sidebar.info(f"**Месяц:** {current_month}")
st.sidebar.info(f"**Год:** {current_year}")

# Подсветка текущей недели в данных
if current_week in df['Неделя'].values:
    current_week_data = df[df['Неделя'] == current_week]
    current_week_orders = current_week_data['Заказали, шт'].sum()
    current_week_sales = current_week_data['Выкупили, шт'].sum()
    
    st.sidebar.success(f"**📊 Данные за текущую неделю {current_week}:**")
    st.sidebar.write(f"Заказы: {current_week_orders:,}")
    st.sidebar.write(f"Выкупы: {current_week_sales:,}")
else:
    st.sidebar.warning(f"**⚠️ Нет данных за текущую неделю {current_week}**")

# Информация о кеше
cache_info = get_cache_info()
if cache_info['exists']:
    st.sidebar.header("💾 Информация о кеше")
    st.sidebar.write(f"Записей в кеше: {cache_info['records']:,}")
    st.sidebar.write(f"Период кеша: {cache_info['start_date']} - {cache_info['end_date']}")
    st.sidebar.write(f"Годы в кеше: {', '.join(map(str, cache_info['years']))}")
    
    # Кнопка очистки кеша
    if st.sidebar.button("🗑️ Очистить кеш", type="secondary"):
        try:
            os.remove("data_cache.csv")
            st.sidebar.success("✅ Кеш очищен")
            st.rerun()
        except:
            st.sidebar.error("❌ Ошибка очистки кеша")
    
    # Экспорт объединенных данных
    if st.sidebar.button("📥 Экспорт данных", type="secondary"):
        try:
            cached_df = load_data_cache()
            if cached_df is not None:
                csv = cached_df.to_csv(index=False, encoding='utf-8-sig')
                st.sidebar.download_button(
                    label="💾 Скачать CSV",
                    data=csv,
                    file_name=f"объединенные_данные_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv"
                )
        except Exception as e:
            st.sidebar.error(f"❌ Ошибка экспорта: {e}")
    
    # Полный отчет с аналитикой
    if st.sidebar.button("📊 Полный отчет", type="primary"):
        try:
            # Загружаем все данные (основные + кеш)
            full_df = load_data_with_cache()
            if full_df is not None:
                # Создаем полный отчет с аналитикой
                report_data = create_full_report(full_df)
                
                # Экспорт в Excel с несколькими листами
                excel_buffer = create_excel_report(report_data)
                
                st.sidebar.download_button(
                    label="📊 Скачать полный отчет (Excel)",
                    data=excel_buffer,
                    file_name=f"полный_отчет_wb_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                st.sidebar.success("✅ Полный отчет готов к скачиванию!")
        except Exception as e:
            st.sidebar.error(f"❌ Ошибка создания отчета: {e}")

# Загрузка дополнительных отчетов
st.sidebar.header("📁 Загрузка отчетов")
st.sidebar.markdown("**Загрузите дополнительные Excel файлы для объединения данных:**")

# Информация о структуре файлов
with st.sidebar.expander("ℹ️ Требования к файлам", expanded=False):
    st.write("**📋 Структура файла должна содержать:**")
    st.write("• Лист 'Товары'")
    st.write("• Заголовки со 2-й строки")
    st.write("• Обязательные колонки:")
    st.write("  - Дата")
    st.write("  - Артикул WB")
    st.write("  - Артикул продавца")
    st.write("  - Заказали, шт")
    st.write("  - Выкупили, шт")
    st.write("• Поддерживаемые форматы: .xlsx, .xls")

# Виджет загрузки файлов
uploaded_files = st.sidebar.file_uploader(
    "Выберите Excel файлы",
    type=['xlsx', 'xls'],
    accept_multiple_files=True,
    help="Можно загрузить несколько файлов одновременно"
)

# Обработка загруженных файлов
if uploaded_files:
    st.sidebar.success(f"✅ Загружено файлов: {len(uploaded_files)}")
    
    # Показываем информацию о загруженных файлах
    for i, file in enumerate(uploaded_files):
        file_info = f"📄 {file.name} ({file.size / 1024:.1f} KB)"
        st.sidebar.write(file_info)
    
    # Сохраняем файлы в session state для обработки
    st.session_state.uploaded_files = uploaded_files
    
    # Кнопка обработки файлов
    if st.sidebar.button("🔄 Обработать загруженные файлы", type="primary"):
        st.session_state.process_uploaded_files = True
        st.rerun()

# Выбор режима анализа
analysis_mode = st.sidebar.selectbox(
    "🎯 Режим анализа",
    ["Общий анализ", "Анализ по товару"],
    help="Выберите режим анализа: общий или по конкретному товару"
)

# Фильтры
st.sidebar.header("🔧 Фильтры")

# Период агрегации
period = st.sidebar.selectbox("Период", ['D', 'W', 'M'], 
                             format_func=lambda x: {'D': 'Дни', 'W': 'Недели', 'M': 'Месяцы'}[x])

# Диапазон дат
date_range = st.sidebar.date_input(
    "Диапазон дат",
    value=(df['Дата'].min().date(), df['Дата'].max().date())
)

# Фильтр по текущей неделе
st.sidebar.markdown("---")
st.sidebar.subheader("🎯 Быстрые фильтры")

# Кнопка для фильтрации по текущей неделе
if st.sidebar.button("📅 Текущая неделя", type="secondary"):
    current_date = datetime.now()
    current_week = current_date.isocalendar().week
    current_year = current_date.year
    
    # Находим даты текущей недели
    week_start = current_date - timedelta(days=current_date.weekday())
    week_end = week_start + timedelta(days=6)
    
    # Обновляем фильтр дат
    st.session_state.quick_filter_week = True
    st.session_state.week_start = week_start.date()
    st.session_state.week_end = week_end.date()
    st.rerun()

# Кнопка для сброса фильтров
if st.sidebar.button("🔄 Сбросить фильтры", type="secondary"):
    st.session_state.quick_filter_week = False
    st.rerun()

# Применяем быстрый фильтр по неделе
if 'quick_filter_week' in st.session_state and st.session_state.quick_filter_week:
    st.sidebar.success(f"✅ Применен фильтр: текущая неделя {current_week}")
    # Обновляем диапазон дат
    if 'week_start' in st.session_state and 'week_end' in st.session_state:
        date_range = (st.session_state.week_start, st.session_state.week_end)

# Фильтрация данных
if len(date_range) == 2:
    filtered_df = df[
        (df['Дата'].dt.date >= date_range[0]) &
        (df['Дата'].dt.date <= date_range[1])
    ]
else:
    filtered_df = df

# Выбор товара для детального анализа
if analysis_mode == "Анализ по товару":
    st.sidebar.header("🏷️ Выбор товара")
    
    # Получаем список товаров с названиями
    products_info = filtered_df[['Артикул WB', 'Артикул продавца', 'Название']].drop_duplicates()
    products_info['display_name'] = products_info['Артикул WB'].astype(str) + ' - ' + products_info['Название'].fillna('Без названия')
    
    selected_product_display = st.sidebar.selectbox(
        "Выберите товар",
        options=products_info['display_name'].tolist(),
        help="Выберите товар для детального анализа"
    )
    
    # Получаем выбранный артикул WB
    selected_wb_article = products_info[products_info['display_name'] == selected_product_display]['Артикул WB'].iloc[0]
    
    # Фильтруем данные по выбранному товару
    filtered_df = filtered_df[filtered_df['Артикул WB'] == selected_wb_article]
    
    # Показываем информацию о выбранном товаре
    product_info = filtered_df[['Артикул WB', 'Артикул продавца', 'Название', 'Предмет', 'Бренд']].iloc[0]
    
    st.header(f"🏷️ Анализ товара: {selected_product_display}")
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.info(f"**Артикул WB:** {product_info['Артикул WB']}")
    with col2:
        st.info(f"**Артикул продавца:** {product_info['Артикул продавца']}")
    with col3:
        st.info(f"**Предмет:** {product_info['Предмет']}")
    with col4:
        st.info(f"**Бренд:** {product_info['Бренд']}")
else:
    st.header("📊 Общий анализ данных")

# Перемещаем вкладки вверх, сразу после заголовка
tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs(["📊 Тренды", "🔄 Конверсия", "💰 Выручка", "📅 По периодам", "📅 Сравнение по дням", "📦 Калькулятор баланса"])

# Информация о загруженных файлах
if 'uploaded_files' in st.session_state and st.session_state.uploaded_files:
    st.info(f"📁 **Загружено файлов:** {len(st.session_state.uploaded_files)}")
    
    # Показываем краткую информацию о файлах
    col1, col2, col3 = st.columns(3)
    with col1:
        st.write("**📄 Файлы:**")
        for file in st.session_state.uploaded_files:
            st.write(f"• {file.name}")
    
    with col2:
        st.write("**📊 Размеры:**")
        for file in st.session_state.uploaded_files:
            st.write(f"• {file.size / 1024:.1f} KB")
    
    with col3:
        st.write("**⚡ Действие:**")
        st.write("Нажмите 'Обработать загруженные файлы' в сайдбаре для объединения данных")
    
    st.markdown("---")

# Информация об анализируемом периоде
st.header("📅 Анализируемый период")

# Показываем выбранный период
if len(date_range) == 2:
    start_date = date_range[0].strftime('%d %B %Y')
    end_date = date_range[1].strftime('%d %B %Y')
    period_days = (date_range[1] - date_range[0]).days + 1
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.info(f"**Начало периода:** {start_date}")
    with col2:
        st.info(f"**Конец периода:** {end_date}")
    with col3:
        st.info(f"**Дней в периоде:** {period_days}")

# Информация о текущей неделе
st.subheader("📅 Текущая неделя")
current_date = datetime.now()
current_week = current_date.isocalendar().week
current_year = current_date.year
current_month = current_date.strftime('%B')
current_day = current_date.strftime('%A')

col1, col2, col3, col4, col5 = st.columns(5)
with col1:
    st.metric("📅 Сегодня", current_date.strftime('%d.%m.%Y'))
with col2:
    st.metric("📆 День недели", current_day)
with col3:
    st.metric("📊 Неделя", current_week)
with col4:
    st.metric("🗓️ Месяц", current_month)
with col5:
    st.metric("📅 Год", current_year)

# Данные за текущую неделю
if current_week in df['Неделя'].values:
    current_week_data = df[df['Неделя'] == current_week]
    current_week_orders = current_week_data['Заказали, шт'].sum()
    current_week_sales = current_week_data['Выкупили, шт'].sum()
    current_week_revenue = current_week_data['Выкупили на сумму, ₽'].sum()
    
    st.success(f"📊 **Данные за текущую неделю {current_week}:**")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("📦 Заказы", f"{current_week_orders:,}")
    with col2:
        st.metric("💰 Выкупы", f"{current_week_sales:,}")
    with col3:
        st.metric("💵 Выручка", f"{current_week_revenue:,.0f} ₽")
else:
    st.warning(f"⚠️ **Нет данных за текущую неделю {current_week}**")
    st.info("Данные за текущую неделю появятся после загрузки соответствующих отчетов")

# Расширенные KPI
st.header("📊 Ключевые показатели (KPI)")

# Основные KPI - первая строка
col1, col2, col3, col4 = st.columns(4)

with col1:
    total_orders = filtered_df['Заказали, шт'].sum()
    total_orders_sum = filtered_df['Заказали на сумму, ₽'].sum()
    st.metric("📦 Заказы", f"{total_orders:,.0f}", f"на {total_orders_sum:,.0f} ₽")

with col2:
    total_sales = filtered_df['Выкупили, шт'].sum()
    total_revenue = filtered_df['Выкупили на сумму, ₽'].sum()
    st.metric("💰 Выкупы", f"{total_sales:,.0f}", f"на {total_revenue:,.0f} ₽")

with col3:
    conversion = (total_sales / total_orders * 100) if total_orders > 0 else 0
    st.metric("📈 Конверсия", f"{conversion:.1f}%")

with col4:
    avg_percent = filtered_df['Процент выкупа'].mean()
    st.metric("🎯 Ср. % выкупа", f"{avg_percent:.1f}%")

# Дополнительные KPI - вторая строка
col1, col2, col3, col4 = st.columns(4)

with col1:
    total_views = filtered_df['Переходы в карточку'].sum()
    st.metric("👁️ Переходы", f"{total_views:,.0f}")

with col2:
    total_cart = filtered_df['Положили в корзину'].sum()
    cart_conversion = (total_cart / total_views * 100) if total_views > 0 else 0
    st.metric("🛒 В корзину", f"{total_cart:,.0f}", f"{cart_conversion:.1f}%")

with col3:
    order_conversion = (total_orders / total_cart * 100) if total_cart > 0 else 0
    st.metric("📋 Конв. заказ", f"{order_conversion:.1f}%")

with col4:
    avg_revenue = total_revenue / total_sales if total_sales > 0 else 0
    st.metric("💎 Ср. чек", f"{avg_revenue:,.0f} ₽")

# Расширенные KPI - третья строка
col1, col2, col3, col4 = st.columns(4)

with col1:
    # Заказ на сумму
    st.metric("📦 Заказ на сумму", f"{total_orders_sum:,.0f} ₽")

with col2:
    # Выкуплено на сумму
    st.metric("💰 Выкуплено на сумму", f"{total_revenue:,.0f} ₽")

with col3:
    # Количество уникальных товаров
    unique_products = filtered_df['Артикул WB'].nunique()
    st.metric("🏷️ Уникальных товаров", f"{unique_products:,}")

with col4:
    # Количество дней с продажами
    days_with_sales = filtered_df[filtered_df['Выкупили, шт'] > 0]['Дата'].nunique()
    st.metric("📅 Дней с продажами", f"{days_with_sales}")

# Дополнительные метрики - четвертая строка
col1, col2, col3, col4 = st.columns(4)

with col1:
    # Средние продажи в день
    avg_sales_per_day = total_sales / period_days if period_days > 0 else 0
    st.metric("📊 Ср. продажи/день", f"{avg_sales_per_day:.1f}")

with col2:
    # Средняя выручка в день
    avg_revenue_per_day = total_revenue / period_days if period_days > 0 else 0
    st.metric("💰 Ср. выручка/день", f"{avg_revenue_per_day:,.0f} ₽")

with col3:
    # Средний чек заказа
    avg_order_value = total_orders_sum / total_orders if total_orders > 0 else 0
    st.metric("📈 Ср. чек заказа", f"{avg_order_value:,.0f} ₽")

with col4:
    # Эффективность (выручка на товар)
    revenue_per_product = total_revenue / unique_products if unique_products > 0 else 0
    st.metric("🎯 Выручка/товар", f"{revenue_per_product:,.0f} ₽")

# Агрегация данных
agg_data = filtered_df.groupby(pd.Grouper(key='Дата', freq=period)).agg({
    'Заказали, шт': 'sum',
    'Выкупили, шт': 'sum',
    'Заказали на сумму, ₽': 'sum',
    'Выкупили на сумму, ₽': 'sum',
    'Переходы в карточку': 'sum',
    'Положили в корзину': 'sum'
}).reset_index()

# Графики
st.header("📈 Графики")

with tab1:
    # Совмещенный график по неделям с подписями месяцев (вверху)
    st.subheader("📊 Совмещенный график: Заказы и выкупы 2024 vs 2025 по неделям")
    
    # Создаем данные по неделям с подписями месяцев
    def create_weekly_data_with_months():
        weekly_data = []
        
        # Данные 2024 года
        data_2024 = filtered_df[filtered_df['Год'] == 2024]
        if not data_2024.empty:
            weekly_2024 = data_2024.groupby(['Неделя', 'Месяц_название']).agg({
                'Заказали, шт': 'sum',
                'Выкупили, шт': 'sum'
            }).reset_index()
            
            for _, row in weekly_2024.iterrows():
                week_label = f"Неделя {row['Неделя']} ({row['Месяц_название']})"
                weekly_data.append({
                    'week_label': week_label,
                    'week_num': row['Неделя'],
                    'month': row['Месяц_название'],
                    'orders_2024': row['Заказали, шт'],
                    'sales_2024': row['Выкупили, шт'],
                    'orders_2025': 0,
                    'sales_2025': 0
                })
        
        # Данные 2025 года
        data_2025 = filtered_df[filtered_df['Год'] == 2025]
        if not data_2025.empty:
            weekly_2025 = data_2025.groupby(['Неделя', 'Месяц_название']).agg({
                'Заказали, шт': 'sum',
                'Выкупили, шт': 'sum'
            }).reset_index()
            
            for _, row in weekly_2025.iterrows():
                week_label = f"Неделя {row['Неделя']} ({row['Месяц_название']})"
                
                # Ищем существующую неделю или создаем новую
                existing_week = next((w for w in weekly_data if w['week_label'] == week_label), None)
                if existing_week:
                    existing_week['orders_2025'] = row['Заказали, шт']
                    existing_week['sales_2025'] = row['Выкупили, шт']
                else:
                    weekly_data.append({
                        'week_label': week_label,
                        'week_num': row['Неделя'],
                        'month': row['Месяц_название'],
                        'orders_2024': 0,
                        'sales_2024': 0,
                        'orders_2025': row['Заказали, шт'],
                        'sales_2025': row['Выкупили, шт']
                    })
        
        # Сортируем по номеру недели
        weekly_data.sort(key=lambda x: x['week_num'])
        return weekly_data
    
    weekly_data = create_weekly_data_with_months()
    
    if weekly_data:
        # Создаем график
        fig_weekly = go.Figure()
        
        # Подготавливаем данные для графика
        week_labels = [w['week_label'] for w in weekly_data]
        orders_2024 = [w['orders_2024'] for w in weekly_data]
        sales_2024 = [w['sales_2024'] for w in weekly_data]
        orders_2025 = [w['orders_2025'] for w in weekly_data]
        sales_2025 = [w['sales_2025'] for w in weekly_data]
        
        # Заказы 2024
        fig_weekly.add_trace(go.Bar(
            x=week_labels,
            y=orders_2024,
            name='Заказы 2024',
            marker_color='blue',
            opacity=0.8
        ))
        
        # Выкупы 2024
        fig_weekly.add_trace(go.Bar(
            x=week_labels,
            y=sales_2024,
            name='Выкупы 2024',
            marker_color='lightblue',
            opacity=0.8
        ))
        
        # Заказы 2025
        fig_weekly.add_trace(go.Bar(
            x=week_labels,
            y=orders_2025,
            name='Заказы 2025',
            marker_color='red',
            opacity=0.8
        ))
        
        # Выкупы 2025
        fig_weekly.add_trace(go.Bar(
            x=week_labels,
            y=sales_2025,
            name='Выкупы 2025',
            marker_color='orange',
            opacity=0.8
        ))
        
        # Настройка графика
        fig_weekly.update_layout(
            title='Заказы и выкупы по неделям: 2024 vs 2025',
            xaxis_title='Неделя (месяц)',
            yaxis_title='Количество',
            barmode='group',
            height=500,
            xaxis=dict(
                tickangle=45,
                tickmode='array',
                ticktext=week_labels,
                tickvals=list(range(len(week_labels)))
            )
        )
        
        st.plotly_chart(fig_weekly, width='stretch')
        
        # Статистика по годам
        st.subheader("📊 Статистика по годам")
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            total_orders_2024 = sum(w['orders_2024'] for w in weekly_data)
            st.metric("📦 Заказы 2024", f"{total_orders_2024:,}")
        
        with col2:
            total_orders_2025 = sum(w['orders_2025'] for w in weekly_data)
            st.metric("📦 Заказы 2025", f"{total_orders_2025:,}")
        
        with col3:
            total_sales_2024 = sum(w['sales_2024'] for w in weekly_data)
            st.metric("💰 Выкупы 2024", f"{total_sales_2024:,}")
        
        with col4:
            total_sales_2025 = sum(w['sales_2025'] for w in weekly_data)
            st.metric("💰 Выкупы 2025", f"{total_sales_2025:,}")
        
        # Сравнение
        if total_orders_2024 > 0 and total_orders_2025 > 0:
            st.subheader("🔄 Сравнение")
            col1, col2 = st.columns(2)
            
            with col1:
                orders_growth = ((total_orders_2025 - total_orders_2024) / total_orders_2024 * 100)
                st.metric(
                    "📈 Рост заказов", 
                    f"{orders_growth:+.1f}%",
                    delta_color="normal"
                )
            
            with col2:
                sales_growth = ((total_sales_2025 - total_sales_2024) / total_sales_2024 * 100) if total_sales_2024 > 0 else 0
                st.metric(
                    "📈 Рост выкупов", 
                    f"{sales_growth:+.1f}%",
                    delta_color="normal"
                )
    else:
        st.info("Нет данных для отображения графика по неделям")
    
    st.markdown("---")
    st.subheader("📊 Настройки графика")
    
    # Выбор типа отображения
    chart_type = st.radio(
        "Выберите тип отображения:",
        ["Общий тренд", "Сравнение по годам", "Совмещенный график"],
        horizontal=True,
        help="Общий тренд - все данные, Сравнение по годам - отдельные графики, Совмещенный график - 2024 и 2025 на одном графике"
    )
    
    if chart_type == "Общий тренд":
        # Оригинальный график
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=agg_data['Дата'], y=agg_data['Заказали, шт'], 
                                name='Заказы', line=dict(color='blue')))
        fig.add_trace(go.Scatter(x=agg_data['Дата'], y=agg_data['Выкупили, шт'], 
                                name='Выкупы', line=dict(color='green')))
        fig.update_layout(title='Тренд заказов и выкупов', xaxis_title='Дата', yaxis_title='Количество')
        st.plotly_chart(fig, width='stretch')
        
    elif chart_type == "Сравнение по годам":
        # Отдельные графики для каждого года
        col1, col2 = st.columns(2)
        
        with col1:
            # График 2024 года
            data_2024 = filtered_df[filtered_df['Год'] == 2024]
            if not data_2024.empty:
                agg_2024 = data_2024.groupby(pd.Grouper(key='Дата', freq=period)).agg({
                    'Заказали, шт': 'sum',
                    'Выкупили, шт': 'sum'
                }).reset_index()
                
                fig_2024 = go.Figure()
                fig_2024.add_trace(go.Scatter(x=agg_2024['Дата'], y=agg_2024['Заказали, шт'], 
                                            name='Заказы', line=dict(color='blue')))
                fig_2024.add_trace(go.Scatter(x=agg_2024['Дата'], y=agg_2024['Выкупили, шт'], 
                                            name='Выкупы', line=dict(color='green')))
                fig_2024.update_layout(title='2024 год - Заказы и выкупы', xaxis_title='Дата', yaxis_title='Количество')
                st.plotly_chart(fig_2024, use_container_width=True)
            else:
                st.info("Нет данных за 2024 год")
        
        with col2:
            # График 2025 года
            data_2025 = filtered_df[filtered_df['Год'] == 2025]
            if not data_2025.empty:
                agg_2025 = data_2025.groupby(pd.Grouper(key='Дата', freq=period)).agg({
                    'Заказали, шт': 'sum',
                    'Выкупили, шт': 'sum'
                }).reset_index()
                
                fig_2025 = go.Figure()
                fig_2025.add_trace(go.Scatter(x=agg_2025['Дата'], y=agg_2025['Заказали, шт'], 
                                            name='Заказы', line=dict(color='blue')))
                fig_2025.add_trace(go.Scatter(x=agg_2025['Дата'], y=agg_2025['Выкупили, шт'], 
                                            name='Выкупы', line=dict(color='green')))
                fig_2025.update_layout(title='2025 год - Заказы и выкупы', xaxis_title='Дата', yaxis_title='Количество')
                st.plotly_chart(fig_2025, use_container_width=True)
            else:
                st.info("Нет данных за 2025 год")
    
    elif chart_type == "Совмещенный график":
        # Совмещенный график 2024 и 2025 года
        fig = go.Figure()
        
        # Данные 2024 года
        data_2024 = filtered_df[filtered_df['Год'] == 2024]
        if not data_2024.empty:
            agg_2024 = data_2024.groupby(pd.Grouper(key='Дата', freq=period)).agg({
                'Заказали, шт': 'sum',
                'Выкупили, шт': 'sum'
            }).reset_index()
            
            # Заказы 2024
            fig.add_trace(go.Scatter(
                x=agg_2024['Дата'], 
                y=agg_2024['Заказали, шт'], 
                name='Заказы 2024', 
                line=dict(color='blue', width=2),
                mode='lines+markers'
            ))
            
            # Выкупы 2024
            fig.add_trace(go.Scatter(
                x=agg_2024['Дата'], 
                y=agg_2024['Выкупили, шт'], 
                name='Выкупы 2024', 
                line=dict(color='lightblue', width=2),
                mode='lines+markers'
            ))
        
        # Данные 2025 года
        data_2025 = filtered_df[filtered_df['Год'] == 2025]
        if not data_2025.empty:
            agg_2025 = data_2025.groupby(pd.Grouper(key='Дата', freq=period)).agg({
                'Заказали, шт': 'sum',
                'Выкупили, шт': 'sum'
            }).reset_index()
            
            # Заказы 2025
            fig.add_trace(go.Scatter(
                x=agg_2025['Дата'], 
                y=agg_2025['Заказали, шт'], 
                name='Заказы 2025', 
                line=dict(color='red', width=2),
                mode='lines+markers'
            ))
            
            # Выкупы 2025
            fig.add_trace(go.Scatter(
                x=agg_2025['Дата'], 
                y=agg_2025['Выкупили, шт'], 
                name='Выкупы 2025', 
                line=dict(color='orange', width=2),
                mode='lines+markers'
            ))
        
        # Настройка графика
        fig.update_layout(
            title='Совмещенный график: Заказы и выкупы 2024 vs 2025',
            xaxis_title='Дата',
            yaxis_title='Количество',
            hovermode='x unified',
            legend=dict(
                orientation="h",
                yanchor="bottom",
                y=1.02,
                xanchor="right",
                x=1
            )
        )
        
        st.plotly_chart(fig, width='stretch')
        
        # Статистика по годам
        if not data_2024.empty or not data_2025.empty:
            st.subheader("📊 Статистика по годам")
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                if not data_2024.empty:
                    orders_2024 = data_2024['Заказали, шт'].sum()
                    st.metric("📦 Заказы 2024", f"{orders_2024:,}")
                else:
                    st.metric("📦 Заказы 2024", "Нет данных")
            
            with col2:
                if not data_2025.empty:
                    orders_2025 = data_2025['Заказали, шт'].sum()
                    st.metric("📦 Заказы 2025", f"{orders_2025:,}")
                else:
                    st.metric("📦 Заказы 2025", "Нет данных")
            
            with col3:
                if not data_2024.empty:
                    sales_2024 = data_2024['Выкупили, шт'].sum()
                    st.metric("💰 Выкупы 2024", f"{sales_2024:,}")
                else:
                    st.metric("💰 Выкупы 2024", "Нет данных")
            
            with col4:
                if not data_2025.empty:
                    sales_2025 = data_2025['Выкупили, шт'].sum()
                    st.metric("💰 Выкупы 2025", f"{sales_2025:,}")
                else:
                    st.metric("💰 Выкупы 2025", "Нет данных")
            
            # Сравнение
            if not data_2024.empty and not data_2025.empty:
                st.subheader("🔄 Сравнение")
                col1, col2 = st.columns(2)
                
                with col1:
                    orders_growth = ((orders_2025 - orders_2024) / orders_2024 * 100) if orders_2024 > 0 else 0
                    st.metric(
                        "📈 Рост заказов", 
                        f"{orders_growth:+.1f}%",
                        delta_color="normal"
                    )
                
                with col2:
                    sales_growth = ((sales_2025 - sales_2024) / sales_2024 * 100) if sales_2024 > 0 else 0
                    st.metric(
                        "📈 Рост выкупов", 
                        f"{sales_growth:+.1f}%",
                        delta_color="normal"
                    )
        
        # KPI анализ: текущий период vs аналогичный период прошлого года
        st.subheader("🎯 KPI: Сравнение с 35 недели (2024 vs 2025)")
        
        # Получаем текущую дату и неделю
        from datetime import datetime, timedelta
        current_date = datetime.now()
        current_year = current_date.year
        current_week = current_date.isocalendar()[1]  # Номер недели в году
        
        # Определяем период для анализа (с 35 недели 2025 года до текущей недели)
        # Находим дату начала 35 недели 2025 года
        week_35_2025 = datetime.strptime(f"{current_year}-W35-1", "%Y-W%W-%w")
        period_start = week_35_2025
        period_end = current_date
        
        # Аналогичный период прошлого года (с 35 недели 2024 года)
        prev_year = current_year - 1
        week_35_2024 = datetime.strptime(f"{prev_year}-W35-1", "%Y-W%W-%w")
        prev_period_start = week_35_2024
        
        # Рассчитываем количество недель для сравнения
        weeks_to_compare = current_week - 35 + 1  # +1 чтобы включить 35 неделю
        prev_period_end = week_35_2024 + timedelta(weeks=weeks_to_compare-1, days=6)  # Конец последней недели
        
        # Фильтруем данные для текущего периода
        current_period_data = filtered_df[
            (filtered_df['Дата'] >= period_start) & 
            (filtered_df['Дата'] <= period_end)
        ]
        
        # Фильтруем данные для аналогичного периода прошлого года
        prev_period_data = filtered_df[
            (filtered_df['Дата'] >= prev_period_start) & 
            (filtered_df['Дата'] <= prev_period_end)
        ]
        
        # Проверяем наличие данных и адаптируемся к доступным данным
        has_current_data = not current_period_data.empty
        has_prev_data = not prev_period_data.empty
        
        if has_current_data:
            # Рассчитываем KPI для текущего периода
            current_orders = current_period_data['Заказали, шт'].sum()
            current_sales = current_period_data['Выкупили, шт'].sum()
            current_revenue = current_period_data['Выкупили на сумму, ₽'].sum()
            current_conversion = (current_sales / current_orders * 100) if current_orders > 0 else 0
            
            # Рассчитываем KPI для аналогичного периода прошлого года (если есть данные)
            if has_prev_data:
                prev_orders = prev_period_data['Заказали, шт'].sum()
                prev_sales = prev_period_data['Выкупили, шт'].sum()
                prev_revenue = prev_period_data['Выкупили на сумму, ₽'].sum()
                prev_conversion = (prev_sales / prev_orders * 100) if prev_orders > 0 else 0
                
                # Рассчитываем изменения
                orders_change = ((current_orders - prev_orders) / prev_orders * 100) if prev_orders > 0 else 0
                sales_change = ((current_sales - prev_sales) / prev_sales * 100) if prev_sales > 0 else 0
                revenue_change = ((current_revenue - prev_revenue) / prev_revenue * 100) if prev_revenue > 0 else 0
                conversion_change = current_conversion - prev_conversion
            else:
                # Если нет данных за прошлый год, показываем только текущие показатели
                prev_orders = prev_sales = prev_revenue = prev_conversion = 0
                orders_change = sales_change = revenue_change = conversion_change = 0
            
            # Отображаем KPI в виде метрик
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                if has_prev_data:
                    st.metric(
                        "📦 Заказы",
                        f"{current_orders:,}",
                        f"{orders_change:+.1f}%",
                        delta_color="normal"
                    )
                    st.caption(f"vs {prev_orders:,} в {prev_year}")
                else:
                    st.metric("📦 Заказы", f"{current_orders:,}")
                    st.caption("⚠️ Нет данных за прошлый год")
            
            with col2:
                if has_prev_data:
                    st.metric(
                        "💰 Выкупы",
                        f"{current_sales:,}",
                        f"{sales_change:+.1f}%",
                        delta_color="normal"
                    )
                    st.caption(f"vs {prev_sales:,} в {prev_year}")
                else:
                    st.metric("💰 Выкупы", f"{current_sales:,}")
                    st.caption("⚠️ Нет данных за прошлый год")
            
            with col3:
                if has_prev_data:
                    st.metric(
                        "💵 Выручка",
                        f"{current_revenue:,.0f} ₽",
                        f"{revenue_change:+.1f}%",
                        delta_color="normal"
                    )
                    st.caption(f"vs {prev_revenue:,.0f} ₽ в {prev_year}")
                else:
                    st.metric("💵 Выручка", f"{current_revenue:,.0f} ₽")
                    st.caption("⚠️ Нет данных за прошлый год")
            
            with col4:
                if has_prev_data:
                    st.metric(
                        "📊 Конверсия",
                        f"{current_conversion:.1f}%",
                        f"{conversion_change:+.1f}п.п.",
                        delta_color="normal"
                    )
                    st.caption(f"vs {prev_conversion:.1f}% в {prev_year}")
                else:
                    st.metric("📊 Конверсия", f"{current_conversion:.1f}%")
                    st.caption("⚠️ Нет данных за прошлый год")
            
            # Дополнительная аналитика
            st.subheader("📈 Детальный анализ")
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.write("**Период анализа:**")
                st.write(f"• Текущий: {period_start.strftime('%d.%m.%Y')} - {period_end.strftime('%d.%m.%Y')}")
                st.write(f"• Недели: 35-{current_week} ({weeks_to_compare} недель)")
                
                if has_prev_data:
                    st.write(f"• Прошлый год: {prev_period_start.strftime('%d.%m.%Y')} - {prev_period_end.strftime('%d.%m.%Y')}")
                    st.write(f"• Недели: 35-{current_week} ({weeks_to_compare} недель)")
                    st.write(f"• Сравниваются **{weeks_to_compare} недель** с 35 недели")
                else:
                    st.write(f"• Прошлый год: {prev_period_start.strftime('%d.%m.%Y')} - {prev_period_end.strftime('%d.%m.%Y')}")
                    st.write(f"• Недели: 35-{current_week} ({weeks_to_compare} недель)")
                    st.write("⚠️ **Данные за прошлый год отсутствуют**")
                
                # Средние показатели
                current_days = (period_end - period_start).days + 1
                st.write("**Средние показатели в день:**")
                st.write(f"• Заказы: {current_orders/current_days:.1f}")
                st.write(f"• Выкупы: {current_sales/current_days:.1f}")
                st.write(f"• Выручка: {current_revenue/current_days:,.0f} ₽")
                
                if has_prev_data:
                    prev_days = (prev_period_end - prev_period_start).days + 1
                    st.write("**Сравнение с прошлым годом:**")
                    st.write(f"• Заказы: {current_orders/current_days:.1f} vs {prev_orders/prev_days:.1f}")
                    st.write(f"• Выкупы: {current_sales/current_days:.1f} vs {prev_sales/prev_days:.1f}")
                    st.write(f"• Выручка: {current_revenue/current_days:,.0f} ₽ vs {prev_revenue/prev_days:,.0f} ₽")
            
            with col2:
                # Индикаторы эффективности
                if has_prev_data:
                    st.write("**Индикаторы эффективности:**")
                    
                    # Цветовая индикация
                    def get_color_indicator(change):
                        if change > 10:
                            return "🟢"
                        elif change > 0:
                            return "🟡"
                        elif change > -10:
                            return "🟠"
                        else:
                            return "🔴"
                    
                    st.write(f"{get_color_indicator(orders_change)} Заказы: {'Рост' if orders_change > 0 else 'Спад'} {abs(orders_change):.1f}%")
                    st.write(f"{get_color_indicator(sales_change)} Выкупы: {'Рост' if sales_change > 0 else 'Спад'} {abs(sales_change):.1f}%")
                    st.write(f"{get_color_indicator(revenue_change)} Выручка: {'Рост' if revenue_change > 0 else 'Спад'} {abs(revenue_change):.1f}%")
                    st.write(f"{get_color_indicator(conversion_change)} Конверсия: {'Рост' if conversion_change > 0 else 'Спад'} {abs(conversion_change):.1f}п.п.")
                    
                    # Общая оценка
                    total_score = (orders_change + sales_change + revenue_change + conversion_change) / 4
                    if total_score > 10:
                        st.success(f"🎉 Отличные результаты! Общий рост: {total_score:.1f}%")
                    elif total_score > 0:
                        st.info(f"📈 Положительная динамика: {total_score:.1f}%")
                    elif total_score > -10:
                        st.warning(f"⚠️ Небольшой спад: {total_score:.1f}%")
                    else:
                        st.error(f"📉 Значительный спад: {total_score:.1f}%")
                else:
                    st.write("**Текущие показатели:**")
                    st.write(f"📦 Заказы: {current_orders:,}")
                    st.write(f"💰 Выкупы: {current_sales:,}")
                    st.write(f"💵 Выручка: {current_revenue:,.0f} ₽")
                    st.write(f"📊 Конверсия: {current_conversion:.1f}%")
                    
                    st.info("📊 Показаны только текущие данные. Для сравнения добавьте данные за прошлый год.")
        
        else:
            st.warning("⚠️ Недостаточно данных для сравнения периодов")

with tab2:
    # Воронка конверсии с процентами
    total_views = agg_data['Переходы в карточку'].sum()
    total_cart = agg_data['Положили в корзину'].sum()
    total_orders = agg_data['Заказали, шт'].sum()
    total_sales = agg_data['Выкупили, шт'].sum()
    
    # KPI для воронки конверсии
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("👁️ Переходы", f"{total_views:,.0f}")
    
    with col2:
        cart_conv_rate = (total_cart / total_views * 100) if total_views > 0 else 0
        st.metric("🛒 В корзину", f"{total_cart:,.0f}", f"{cart_conv_rate:.1f}%")
    
    with col3:
        order_conv_rate = (total_orders / total_cart * 100) if total_cart > 0 else 0
        st.metric("📋 Заказы", f"{total_orders:,.0f}", f"{order_conv_rate:.1f}%")
    
    with col4:
        sales_conv_rate = (total_sales / total_orders * 100) if total_orders > 0 else 0
        st.metric("💰 Выкупы", f"{total_sales:,.0f}", f"{sales_conv_rate:.1f}%")
    
    # Воронка конверсии
    fig = go.Figure(go.Funnel(
        y=['Переходы', 'В корзину', 'Заказы', 'Выкупы'],
        x=[total_views, total_cart, total_orders, total_sales],
        textinfo="value+percent initial",
        textposition="inside",
        marker=dict(color=['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728'])
    ))
    
    fig.update_layout(
        title='Воронка конверсии',
        annotations=[
            dict(x=0.5, y=0.9, xref="paper", yref="paper", 
                 text=f"Конверсия в корзину: {cart_conv_rate:.1f}%", 
                 showarrow=False, font=dict(size=14)),
            dict(x=0.5, y=0.8, xref="paper", yref="paper", 
                 text=f"Конверсия в заказ: {order_conv_rate:.1f}%", 
                 showarrow=False, font=dict(size=14)),
            dict(x=0.5, y=0.7, xref="paper", yref="paper", 
                 text=f"Процент выкупа: {sales_conv_rate:.1f}%", 
                 showarrow=False, font=dict(size=14))
        ]
    )
    st.plotly_chart(fig, width='stretch')

with tab3:
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=agg_data['Дата'], y=agg_data['Выкупили на сумму, ₽'], 
                            name='Выручка', fill='tonexty'))
    fig.update_layout(title='Тренд выручки', xaxis_title='Дата', yaxis_title='Выручка, ₽')
    st.plotly_chart(fig, width='stretch')

with tab4:
    st.subheader("📅 Настройки графиков по периодам")
    
    # Выбор типа отображения
    period_chart_type = st.radio(
        "Выберите тип отображения:",
        ["Общие данные", "Сравнение по годам", "Совмещенные графики"],
        horizontal=True,
        help="Общие данные - все данные, Сравнение по годам - отдельные графики, Совмещенные графики - 2024 и 2025 на одном графике"
    )
    
    if period_chart_type == "Общие данные":
        # Оригинальные графики
        fig = make_subplots(
            rows=2, cols=1,
            subplot_titles=('Заказы по месяцам', 'Заказы по неделям'),
            vertical_spacing=0.1
        )
        
        # По месяцам
        monthly_data = filtered_df.groupby('Месяц').agg({
            'Заказали, шт': 'sum',
            'Выкупили, шт': 'sum'
        }).reset_index()
        
        fig.add_trace(
            go.Bar(x=monthly_data['Месяц'], y=monthly_data['Заказали, шт'], 
                   name='Заказы', marker_color='blue'),
            row=1, col=1
        )
        fig.add_trace(
            go.Bar(x=monthly_data['Месяц'], y=monthly_data['Выкупили, шт'], 
                   name='Выкупы', marker_color='green'),
            row=1, col=1
        )
        
        # По неделям
        weekly_data = filtered_df.groupby('Неделя').agg({
            'Заказали, шт': 'sum',
            'Выкупили, шт': 'sum'
        }).reset_index()
        
        fig.add_trace(
            go.Bar(x=weekly_data['Неделя'], y=weekly_data['Заказали, шт'], 
                   name='Заказы', marker_color='blue', showlegend=False),
            row=2, col=1
        )
        fig.add_trace(
            go.Bar(x=weekly_data['Неделя'], y=weekly_data['Выкупили, шт'], 
                   name='Выкупы', marker_color='green', showlegend=False),
            row=2, col=1
        )
        
        fig.update_layout(height=600, title_text="Заказы по месяцам и неделям")
        st.plotly_chart(fig, width='stretch')
        
    elif period_chart_type == "Сравнение по годам":
        # Отдельные графики для каждого года
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("📊 2024 год")
            data_2024 = filtered_df[filtered_df['Год'] == 2024]
            if not data_2024.empty:
                # По месяцам 2024
                monthly_2024 = data_2024.groupby('Месяц').agg({
                    'Заказали, шт': 'sum',
                    'Выкупили, шт': 'sum'
                }).reset_index()
                
                fig_monthly_2024 = go.Figure()
                fig_monthly_2024.add_trace(go.Bar(x=monthly_2024['Месяц'], y=monthly_2024['Заказали, шт'], 
                                               name='Заказы', marker_color='blue'))
                fig_monthly_2024.add_trace(go.Bar(x=monthly_2024['Месяц'], y=monthly_2024['Выкупили, шт'], 
                                               name='Выкупы', marker_color='green'))
                fig_monthly_2024.update_layout(title='2024 - По месяцам', height=300)
                st.plotly_chart(fig_monthly_2024, use_container_width=True)
                
                # По неделям 2024
                weekly_2024 = data_2024.groupby('Неделя').agg({
                    'Заказали, шт': 'sum',
                    'Выкупили, шт': 'sum'
                }).reset_index()
                
                fig_weekly_2024 = go.Figure()
                fig_weekly_2024.add_trace(go.Bar(x=weekly_2024['Неделя'], y=weekly_2024['Заказали, шт'], 
                                              name='Заказы', marker_color='blue'))
                fig_weekly_2024.add_trace(go.Bar(x=weekly_2024['Неделя'], y=weekly_2024['Выкупили, шт'], 
                                              name='Выкупы', marker_color='green'))
                fig_weekly_2024.update_layout(title='2024 - По неделям', height=300)
                st.plotly_chart(fig_weekly_2024, use_container_width=True)
            else:
                st.info("Нет данных за 2024 год")
        
        with col2:
            st.subheader("📊 2025 год")
            data_2025 = filtered_df[filtered_df['Год'] == 2025]
            if not data_2025.empty:
                # По месяцам 2025
                monthly_2025 = data_2025.groupby('Месяц').agg({
                    'Заказали, шт': 'sum',
                    'Выкупили, шт': 'sum'
                }).reset_index()
                
                fig_monthly_2025 = go.Figure()
                fig_monthly_2025.add_trace(go.Bar(x=monthly_2025['Месяц'], y=monthly_2025['Заказали, шт'], 
                                               name='Заказы', marker_color='red'))
                fig_monthly_2025.add_trace(go.Bar(x=monthly_2025['Месяц'], y=monthly_2025['Выкупили, шт'], 
                                               name='Выкупы', marker_color='orange'))
                fig_monthly_2025.update_layout(title='2025 - По месяцам', height=300)
                st.plotly_chart(fig_monthly_2025, use_container_width=True)
                
                # По неделям 2025
                weekly_2025 = data_2025.groupby('Неделя').agg({
                    'Заказали, шт': 'sum',
                    'Выкупили, шт': 'sum'
                }).reset_index()
                
                fig_weekly_2025 = go.Figure()
                fig_weekly_2025.add_trace(go.Bar(x=weekly_2025['Неделя'], y=weekly_2025['Заказали, шт'], 
                                              name='Заказы', marker_color='red'))
                fig_weekly_2025.add_trace(go.Bar(x=weekly_2025['Неделя'], y=weekly_2025['Выкупили, шт'], 
                                              name='Выкупы', marker_color='orange'))
                fig_weekly_2025.update_layout(title='2025 - По неделям', height=300)
                st.plotly_chart(fig_weekly_2025, use_container_width=True)
            else:
                st.info("Нет данных за 2025 год")
    
    elif period_chart_type == "Совмещенные графики":
        # Совмещенные графики 2024 и 2025 года
        st.subheader("📊 Совмещенные графики: 2024 vs 2025")
        
        # График по месяцам
        fig_monthly = go.Figure()
        
        # Данные 2024 года по месяцам
        data_2024 = filtered_df[filtered_df['Год'] == 2024]
        if not data_2024.empty:
            monthly_2024 = data_2024.groupby('Месяц').agg({
                'Заказали, шт': 'sum',
                'Выкупили, шт': 'sum'
            }).reset_index()
            
            fig_monthly.add_trace(go.Bar(
                x=monthly_2024['Месяц'], 
                y=monthly_2024['Заказали, шт'], 
                name='Заказы 2024', 
                marker_color='blue',
                opacity=0.8
            ))
            
            fig_monthly.add_trace(go.Bar(
                x=monthly_2024['Месяц'], 
                y=monthly_2024['Выкупили, шт'], 
                name='Выкупы 2024', 
                marker_color='lightblue',
                opacity=0.8
            ))
        
        # Данные 2025 года по месяцам
        data_2025 = filtered_df[filtered_df['Год'] == 2025]
        if not data_2025.empty:
            monthly_2025 = data_2025.groupby('Месяц').agg({
                'Заказали, шт': 'sum',
                'Выкупили, шт': 'sum'
            }).reset_index()
            
            fig_monthly.add_trace(go.Bar(
                x=monthly_2025['Месяц'], 
                y=monthly_2025['Заказали, шт'], 
                name='Заказы 2025', 
                marker_color='red',
                opacity=0.8
            ))
            
            fig_monthly.add_trace(go.Bar(
                x=monthly_2025['Месяц'], 
                y=monthly_2025['Выкупили, шт'], 
                name='Выкупы 2025', 
                marker_color='orange',
                opacity=0.8
            ))
        
        fig_monthly.update_layout(
            title='Заказы и выкупы по месяцам: 2024 vs 2025',
            xaxis_title='Месяц',
            yaxis_title='Количество',
            barmode='group',
            height=400
        )
        st.plotly_chart(fig_monthly, width='stretch')
        
        # График по неделям
        fig_weekly = go.Figure()
        
        # Данные 2024 года по неделям
        if not data_2024.empty:
            weekly_2024 = data_2024.groupby('Неделя').agg({
                'Заказали, шт': 'sum',
                'Выкупили, шт': 'sum'
            }).reset_index()
            
            fig_weekly.add_trace(go.Bar(
                x=weekly_2024['Неделя'], 
                y=weekly_2024['Заказали, шт'], 
                name='Заказы 2024', 
                marker_color='blue',
                opacity=0.8
            ))
            
            fig_weekly.add_trace(go.Bar(
                x=weekly_2024['Неделя'], 
                y=weekly_2024['Выкупили, шт'], 
                name='Выкупы 2024', 
                marker_color='lightblue',
                opacity=0.8
            ))
        
        # Данные 2025 года по неделям
        if not data_2025.empty:
            weekly_2025 = data_2025.groupby('Неделя').agg({
                'Заказали, шт': 'sum',
                'Выкупили, шт': 'sum'
            }).reset_index()
            
            fig_weekly.add_trace(go.Bar(
                x=weekly_2025['Неделя'], 
                y=weekly_2025['Заказали, шт'], 
                name='Заказы 2025', 
                marker_color='red',
                opacity=0.8
            ))
            
            fig_weekly.add_trace(go.Bar(
                x=weekly_2025['Неделя'], 
                y=weekly_2025['Выкупили, шт'], 
                name='Выкупы 2025', 
                marker_color='orange',
                opacity=0.8
            ))
        
        fig_weekly.update_layout(
            title='Заказы и выкупы по неделям: 2024 vs 2025',
            xaxis_title='Неделя',
            yaxis_title='Количество',
            barmode='group',
            height=400
        )
        st.plotly_chart(fig_weekly, width='stretch')

with tab5:
    st.header("📅 Сравнение заказов по дням: 2024 vs 2025")
    
    # Выбор месяца для сравнения
    available_months = sorted(df['Дата'].dt.month.unique())
    month_names = {1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель', 5: 'Май', 6: 'Июнь',
                  7: 'Июль', 8: 'Август', 9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'}
    
    # Определяем текущий месяц
    current_month = datetime.now().month
    
    # Если текущий месяц есть в данных, используем его, иначе последний доступный
    if current_month in available_months:
        default_index = available_months.index(current_month)
    else:
        default_index = len(available_months) - 1
    
    selected_month = st.selectbox(
        "Выберите месяц для сравнения",
        options=available_months,
        format_func=lambda x: month_names[x],
        index=default_index  # По умолчанию текущий месяц
    )
    
    selected_month_name = month_names[selected_month]
    
    # Фильтруем данные только за выбранный месяц
    selected_month_data = df[df['Дата'].dt.month == selected_month]
    
    daily_comparison = selected_month_data.groupby(['День_месяца', 'Год']).agg({
        'Заказали, шт': 'sum',
        'Выкупили, шт': 'sum',
        'Выкупили на сумму, ₽': 'sum'
    }).reset_index()
    
    # Переименовываем столбец для удобства
    daily_comparison = daily_comparison.rename(columns={'День_месяца': 'День'})
    
    # Сортируем по дню месяца
    daily_comparison = daily_comparison.sort_values('День')
    
    # Создаем отдельные графики для 2024 и 2025 годов
    col1, col2 = st.columns(2)
    
    with col1:
        # График для 2024 года
        fig_2024 = go.Figure()
        
        data_2024 = daily_comparison[daily_comparison['Год'] == 2024]
        if not data_2024.empty:
            # Заказы 2024
            fig_2024.add_trace(go.Scatter(
                x=data_2024['День'], 
                y=data_2024['Заказали, шт'],
                name='Заказы',
                line=dict(color='blue', width=2),
                mode='lines+markers'
            ))
            # Выкупы 2024
            fig_2024.add_trace(go.Scatter(
                x=data_2024['День'], 
                y=data_2024['Выкупили, шт'],
                name='Выкупы',
                line=dict(color='lightblue', width=2),
                mode='lines+markers'
            ))
        
        fig_2024.update_layout(
            title=f'2024 год - Заказы и выкупы по дням ({selected_month_name})',
            xaxis_title='День месяца',
            yaxis_title='Количество',
            height=400,
            hovermode='x unified'
        )
        
        # Настройка оси X для лучшей читаемости
        fig_2024.update_xaxes(tickangle=45)
        
        st.plotly_chart(fig_2024, width='stretch')
    
    with col2:
        # График для 2025 года
        fig_2025 = go.Figure()
        
        data_2025 = daily_comparison[daily_comparison['Год'] == 2025]
        if not data_2025.empty:
            # Заказы 2025
            fig_2025.add_trace(go.Scatter(
                x=data_2025['День'], 
                y=data_2025['Заказали, шт'],
                name='Заказы',
                line=dict(color='red', width=2),
                mode='lines+markers'
            ))
            # Выкупы 2025
            fig_2025.add_trace(go.Scatter(
                x=data_2025['День'], 
                y=data_2025['Выкупили, шт'],
                name='Выкупы',
                line=dict(color='orange', width=2),
                mode='lines+markers'
            ))
        
        fig_2025.update_layout(
            title=f'2025 год - Заказы и выкупы по дням ({selected_month_name})',
            xaxis_title='День месяца',
            yaxis_title='Количество',
            height=400,
            hovermode='x unified'
        )
        
        # Настройка оси X для лучшей читаемости
        fig_2025.update_xaxes(tickangle=45)
        
        st.plotly_chart(fig_2025, width='stretch')
    
    # Статистика сравнения
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("📋 Статистика по годам")
        if not data_2024.empty:
            st.write(f"**2024 год:**")
            st.write(f"- Всего заказов: {data_2024['Заказали, шт'].sum():,.0f}")
            st.write(f"- Среднее заказов в день: {data_2024['Заказали, шт'].mean():,.1f}")
            st.write(f"- Максимум заказов: {data_2024['Заказали, шт'].max():,.0f}")
        
        if not data_2025.empty:
            st.write(f"**2025 год:**")
            st.write(f"- Всего заказов: {data_2025['Заказали, шт'].sum():,.0f}")
            st.write(f"- Среднее заказов в день: {data_2025['Заказали, шт'].mean():,.1f}")
            st.write(f"- Максимум заказов: {data_2025['Заказали, шт'].max():,.0f}")
    
    with col2:
        st.subheader("📈 Сравнение")
        if not data_2024.empty and not data_2025.empty:
            total_2024 = data_2024['Заказали, шт'].sum()
            total_2025 = data_2025['Заказали, шт'].sum()
            avg_2024 = data_2024['Заказали, шт'].mean()
            avg_2025 = data_2025['Заказали, шт'].mean()
            
            growth_total = ((total_2025 - total_2024) / total_2024 * 100) if total_2024 > 0 else 0
            growth_avg = ((avg_2025 - avg_2024) / avg_2024 * 100) if avg_2024 > 0 else 0
            
            st.write(f"**Рост общего количества заказов:** {growth_total:+.1f}%")
            st.write(f"**Рост среднего количества заказов:** {growth_avg:+.1f}%")
            
            if growth_total > 0:
                st.success(f"📈 Продажи в 2025 году выросли на {growth_total:.1f}%")
            else:
                st.error(f"📉 Продажи в 2025 году снизились на {abs(growth_total):.1f}%")
    
    # Таблица всех дней с заказами и выкупами (за выбранный месяц)
    st.subheader(f"📋 Все дни {selected_month_name}: заказы и выкупы 2024 vs 2025")
    
    # Создаем полную таблицу всех дней
    all_days_data = []
    
    # Получаем все уникальные дни месяца
    all_days = sorted(daily_comparison['День'].unique())
    
    for day in all_days:
        # Данные для 2024 года
        day_2024_orders = data_2024[data_2024['День'] == day]['Заказали, шт'].iloc[0] if not data_2024[data_2024['День'] == day].empty else 0
        day_2024_sales = data_2024[data_2024['День'] == day]['Выкупили, шт'].iloc[0] if not data_2024[data_2024['День'] == day].empty else 0
        day_2024_revenue = data_2024[data_2024['День'] == day]['Выкупили на сумму, ₽'].iloc[0] if not data_2024[data_2024['День'] == day].empty else 0
        
        # Данные для 2025 года
        day_2025_orders = data_2025[data_2025['День'] == day]['Заказали, шт'].iloc[0] if not data_2025[data_2025['День'] == day].empty else 0
        day_2025_sales = data_2025[data_2025['День'] == day]['Выкупили, шт'].iloc[0] if not data_2025[data_2025['День'] == day].empty else 0
        day_2025_revenue = data_2025[data_2025['День'] == day]['Выкупили на сумму, ₽'].iloc[0] if not data_2025[data_2025['День'] == day].empty else 0
        
        all_days_data.append({
            'День месяца': day,
            'Заказы 2024': day_2024_orders,
            'Заказы 2025': day_2025_orders,
            'Выкупы 2024': day_2024_sales,
            'Выкупы 2025': day_2025_sales,
            'Выручка 2024': day_2024_revenue,
            'Выручка 2025': day_2025_revenue,
            'Рост заказов %': ((day_2025_orders - day_2024_orders) / day_2024_orders * 100) if day_2024_orders > 0 else 0,
            'Рост выкупов %': ((day_2025_sales - day_2024_sales) / day_2024_sales * 100) if day_2024_sales > 0 else 0
        })
    
    all_days_df = pd.DataFrame(all_days_data)
    
    # Форматируем проценты
    all_days_df['Рост заказов %'] = all_days_df['Рост заказов %'].round(1)
    all_days_df['Рост выкупов %'] = all_days_df['Рост выкупов %'].round(1)
    
    # Сортируем по дню месяца
    all_days_df = all_days_df.sort_values('День месяца')
    
    st.dataframe(all_days_df, width='stretch')

# Сравнение по периодам (месяцы, конкретные даты, недели)
st.header("📅 Сравнение по периодам")

# Выбор типа сравнения
comparison_type = st.selectbox(
    "Выберите тип сравнения",
    ["По месяцам", "По конкретным датам", "По неделям"],
    help="Выберите тип сравнения для анализа"
)

if comparison_type == "По месяцам":
    # Получаем доступные месяцы
    available_months = filtered_df['Месяц_название'].unique()
    available_years = filtered_df['Год'].unique()
    
    col1, col2 = st.columns(2)
    
    with col1:
        selected_month = st.selectbox("Выберите месяц для сравнения", available_months)
        
    with col2:
        st.write("**Сравнение по годам:**")
        for year in sorted(available_years):
            month_data = filtered_df[
                (filtered_df['Месяц_название'] == selected_month) & 
                (filtered_df['Год'] == year)
            ]
            if not month_data.empty:
                orders = month_data['Заказали, шт'].sum()
                sales = month_data['Выкупили, шт'].sum()
                revenue = month_data['Выкупили на сумму, ₽'].sum()
                st.write(f"**{year}:** Заказы: {orders:,.0f}, Выкупы: {sales:,.0f}, Выручка: {revenue:,.0f} ₽")
    
    # График сравнения по месяцам
    if st.button("Показать график сравнения по месяцам"):
        month_comparison = filtered_df[filtered_df['Месяц_название'] == selected_month].groupby('Год').agg({
            'Заказали, шт': 'sum',
            'Выкупили, шт': 'sum',
            'Выкупили на сумму, ₽': 'sum'
        }).reset_index()
        
        fig = make_subplots(
            rows=1, cols=3,
            subplot_titles=(f'Заказы в {selected_month}', f'Выкупы в {selected_month}', f'Выручка в {selected_month}'),
            specs=[[{"type": "bar"}, {"type": "bar"}, {"type": "bar"}]]
        )
        
        fig.add_trace(
            go.Bar(x=month_comparison['Год'], y=month_comparison['Заказали, шт'], 
                   name='Заказы', marker_color=['blue', 'red']),
            row=1, col=1
        )
        
        fig.add_trace(
            go.Bar(x=month_comparison['Год'], y=month_comparison['Выкупили, шт'], 
                   name='Выкупы', marker_color=['green', 'orange']),
            row=1, col=2
        )
        
        fig.add_trace(
            go.Bar(x=month_comparison['Год'], y=month_comparison['Выкупили на сумму, ₽'], 
                   name='Выручка', marker_color=['purple', 'brown']),
            row=1, col=3
        )
        
        fig.update_layout(height=400, title_text=f"Сравнение {selected_month} по годам")
        st.plotly_chart(fig, width='stretch')

elif comparison_type == "По конкретным датам":
    # Получаем доступные месяцы и дни
    available_months = filtered_df['Месяц_название'].unique()
    available_days = filtered_df['День_месяца'].unique()
    available_years = filtered_df['Год'].unique()
    
    col1, col2 = st.columns(2)
    
    with col1:
        selected_month = st.selectbox("Выберите месяц", available_months)
        selected_day = st.selectbox("Выберите день месяца", sorted(available_days))
        
    with col2:
        st.write("**Сравнение по годам:**")
        for year in sorted(available_years):
            date_data = filtered_df[
                (filtered_df['Месяц_название'] == selected_month) & 
                (filtered_df['День_месяца'] == selected_day) &
                (filtered_df['Год'] == year)
            ]
            if not date_data.empty:
                orders = date_data['Заказали, шт'].sum()
                sales = date_data['Выкупили, шт'].sum()
                revenue = date_data['Выкупили на сумму, ₽'].sum()
                st.write(f"**{year} ({selected_day} {selected_month}):** Заказы: {orders:,.0f}, Выкупы: {sales:,.0f}, Выручка: {revenue:,.0f} ₽")
    
    # График сравнения по конкретным датам
    if st.button("Показать график сравнения по датам"):
        date_comparison = filtered_df[
            (filtered_df['Месяц_название'] == selected_month) & 
            (filtered_df['День_месяца'] == selected_day)
        ].groupby('Год').agg({
            'Заказали, шт': 'sum',
            'Выкупили, шт': 'sum',
            'Выкупили на сумму, ₽': 'sum'
        }).reset_index()
        
        fig = make_subplots(
            rows=1, cols=3,
            subplot_titles=(f'Заказы {selected_day} {selected_month}', f'Выкупы {selected_day} {selected_month}', f'Выручка {selected_day} {selected_month}'),
            specs=[[{"type": "bar"}, {"type": "bar"}, {"type": "bar"}]]
        )
        
        fig.add_trace(
            go.Bar(x=date_comparison['Год'], y=date_comparison['Заказали, шт'], 
                   name='Заказы', marker_color=['blue', 'red']),
            row=1, col=1
        )
        
        fig.add_trace(
            go.Bar(x=date_comparison['Год'], y=date_comparison['Выкупили, шт'], 
                   name='Выкупы', marker_color=['green', 'orange']),
            row=1, col=2
        )
        
        fig.add_trace(
            go.Bar(x=date_comparison['Год'], y=date_comparison['Выкупили на сумму, ₽'], 
                   name='Выручка', marker_color=['purple', 'brown']),
            row=1, col=3
        )
        
        fig.update_layout(height=400, title_text=f"Сравнение {selected_day} {selected_month} по годам")
        st.plotly_chart(fig, width='stretch')

elif comparison_type == "По неделям":
    # Получаем доступные недели
    available_weeks = filtered_df['Неделя'].unique()
    available_years = filtered_df['Год'].unique()
    
    col1, col2 = st.columns(2)
    
    with col1:
        selected_week = st.selectbox("Выберите неделю для сравнения", sorted(available_weeks))
        
    with col2:
        st.write("**Сравнение по годам:**")
        for year in sorted(available_years):
            week_data = filtered_df[
                (filtered_df['Неделя'] == selected_week) & 
                (filtered_df['Год'] == year)
            ]
            if not week_data.empty:
                orders = week_data['Заказали, шт'].sum()
                sales = week_data['Выкупили, шт'].sum()
                revenue = week_data['Выкупили на сумму, ₽'].sum()
                st.write(f"**{year} (Неделя {selected_week}):** Заказы: {orders:,.0f}, Выкупы: {sales:,.0f}, Выручка: {revenue:,.0f} ₽")
    
    # График сравнения по неделям
    if st.button("Показать график сравнения по неделям"):
        week_comparison = filtered_df[filtered_df['Неделя'] == selected_week].groupby('Год').agg({
            'Заказали, шт': 'sum',
            'Выкупили, шт': 'sum',
            'Выкупили на сумму, ₽': 'sum'
        }).reset_index()
        
        fig = make_subplots(
            rows=1, cols=3,
            subplot_titles=(f'Заказы в неделю {selected_week}', f'Выкупы в неделю {selected_week}', f'Выручка в неделю {selected_week}'),
            specs=[[{"type": "bar"}, {"type": "bar"}, {"type": "bar"}]]
        )
        
        fig.add_trace(
            go.Bar(x=week_comparison['Год'], y=week_comparison['Заказали, шт'], 
                   name='Заказы', marker_color=['blue', 'red']),
            row=1, col=1
        )
        
        fig.add_trace(
            go.Bar(x=week_comparison['Год'], y=week_comparison['Выкупили, шт'], 
                   name='Выкупы', marker_color=['green', 'orange']),
            row=1, col=2
        )
        
        fig.add_trace(
            go.Bar(x=week_comparison['Год'], y=week_comparison['Выкупили на сумму, ₽'], 
                   name='Выручка', marker_color=['purple', 'brown']),
            row=1, col=3
        )
        
        fig.update_layout(height=400, title_text=f"Сравнение недели {selected_week} по годам")
        st.plotly_chart(fig, width='stretch')

# Детальная таблица
st.header("📋 Детальные данные")

if st.checkbox("Показать таблицу"):
    st.dataframe(filtered_df, width='stretch')
    
    # Экспорт
    col1, col2 = st.columns(2)
    
    with col1:
        csv = filtered_df.to_csv(index=False, encoding='utf-8-sig')
        st.download_button(
            label="📥 Скачать CSV",
            data=csv,
            file_name=f"wb_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv"
        )
    
    with col2:
        if st.button("📊 Полный отчет (Excel)", type="primary"):
            try:
                # Загружаем все данные (основные + кеш)
                full_df = load_data_with_cache()
                if full_df is not None:
                    # Создаем полный отчет с аналитикой
                    report_data = create_full_report(full_df)
                    
                    # Экспорт в Excel с несколькими листами
                    excel_buffer = create_excel_report(report_data)
                    
                    st.download_button(
                        label="📊 Скачать полный отчет (Excel)",
                        data=excel_buffer,
                        file_name=f"полный_отчет_wb_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                    st.success("✅ Полный отчет готов к скачиванию!")
            except Exception as e:
                st.error(f"❌ Ошибка создания отчета: {e}")

with tab6:
    st.header("📦 Калькулятор заказов и баланса товаров")
    st.markdown("---")
    
    # Определение месяцев
    months = [
        "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
        "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
    ]
    
    # Инициализация сессии для калькулятора
    if 'orders_data' not in st.session_state:
        st.session_state.orders_data = {}
    if 'return_percentage' not in st.session_state:
        st.session_state.return_percentage = 20.0  # По умолчанию 20%
    if 'initial_balance' not in st.session_state:
        st.session_state.initial_balance = 500  # По умолчанию 500
    
    # Получение данных из основного приложения
    def get_orders_from_data():
        """Получает данные заказов из основного приложения"""
        orders_data = {}
        
        # Группируем данные по месяцам и неделям
        weekly_orders = filtered_df.groupby(['Месяц', 'Неделя']).agg({
            'Заказали, шт': 'sum'
        }).reset_index()
        
        # Преобразуем в формат для калькулятора
        for _, row in weekly_orders.iterrows():
            month_num = int(row['Месяц'])
            week_num = int(row['Неделя'])
            orders = int(row['Заказали, шт'])
            
            if orders > 0:  # Только ненулевые заказы
                month_name = months[month_num - 1]  # Индексация с 0
                key = f"{month_name}_{week_num}"
                orders_data[key] = orders
        
        return orders_data
    
    # Автоматически загружаем данные из основного приложения
    if 'auto_load_data' not in st.session_state:
        st.session_state.auto_load_data = True
    
    # Инициализация переменных для сохранения выбранного диапазона
    if 'selected_start_period' not in st.session_state:
        st.session_state.selected_start_period = None
    if 'selected_end_period' not in st.session_state:
        st.session_state.selected_end_period = None
    
    if st.session_state.auto_load_data:
        auto_orders = get_orders_from_data()
        if auto_orders:
            st.session_state.orders_data = auto_orders
            st.session_state.auto_load_data = False
            st.success("✅ Данные заказов автоматически загружены из основного приложения!")
    
    # Настройки и анализ
    st.subheader("⚙️ Настройки и анализ")
    
    # Создаем колонки для настроек
    col_settings1, col_settings2, col_settings3 = st.columns(3)
    
    with col_settings1:
        # Процент выкупа
        return_percentage = st.number_input(
            "Процент выкупа (%)",
            min_value=0.0,
            max_value=100.0,
            value=st.session_state.return_percentage,
            step=0.1,
            help="Процент товаров, которые будут выкуплены"
        )
        st.session_state.return_percentage = return_percentage
    
    with col_settings2:
        # Начальный баланс
        initial_balance = st.number_input(
            "Начальный баланс товаров",
            min_value=0,
            value=st.session_state.initial_balance,
            step=1,
            help="Количество товаров на складе в начале периода"
        )
        st.session_state.initial_balance = initial_balance
    
    with col_settings3:
        st.write("")
        st.write("")
        # Кнопка для обновления данных
        if st.button("🔄 Обновить данные", type="primary"):
            # Сохраняем текущий выбранный диапазон
            if 'start_period' in locals() and 'end_period' in locals():
                st.session_state.selected_start_period = start_period
                st.session_state.selected_end_period = end_period
            
            auto_orders = get_orders_from_data()
            if auto_orders:
                st.session_state.orders_data = auto_orders
                st.success("✅ Данные обновлены из основного приложения!")
                st.rerun()
            else:
                st.warning("⚠️ Нет данных для загрузки")
        
        # Кнопка для очистки данных
        if st.button("🗑️ Очистить все данные", type="secondary"):
            st.session_state.orders_data = {}
            st.rerun()
    
    # Расчет минимального необходимого баланса
    def calculate_min_balance():
        if not st.session_state.orders_data:
            return {
                'simple': 0,
                'optimal': 0,
                'with_returns': 0
            }
        
        # Простой расчет - общее количество заказов
        total_orders = sum(st.session_state.orders_data.values())
        
        # Продвинутый расчет с учетом возвратов
        def calculate_optimal_balance():
            if not st.session_state.orders_data:
                return 0
            
            # Начинаем с нулевого баланса для поиска минимального необходимого
            test_balance = 0
            max_deficit = 0
            returned_from_previous = 0
            total_orders_in_period = 0
            
            # Сортируем данные по месяцам и неделям для правильного порядка
            sorted_orders = []
            for key, value in st.session_state.orders_data.items():
                month, week = key.split('_')
                month_index = months.index(month)
                sorted_orders.append((month_index, int(week), value))
            
            # Сортируем по индексу месяца и номеру недели
            sorted_orders.sort(key=lambda x: (x[0], x[1]))
            
            for month_index, week, orders in sorted_orders:
                if orders > 0:
                    total_orders_in_period += orders
                    # Списываем заказы с тестового баланса
                    test_balance = test_balance - orders + returned_from_previous
                    
                    # Если баланс отрицательный, это дефицит
                    if test_balance < 0:
                        deficit = abs(test_balance)
                        if deficit > max_deficit:
                            max_deficit = deficit
                    
                    # Возвращенные товары придут на следующую неделю
                    returned = orders * (1 - return_percentage / 100)
                    returned_from_previous = returned
            
            # Если нет дефицита, возвращаем максимальный заказ как минимальный баланс
            if max_deficit == 0 and total_orders_in_period > 0:
                max_order = max(st.session_state.orders_data.values())
                return max_order
            
            return max_deficit
        
        optimal_balance = calculate_optimal_balance()
        
        return {
            'simple': round(total_orders),
            'optimal': round(optimal_balance),
            'with_returns': round(total_orders - (total_orders * (return_percentage / 100)))
        }
    
    # Функция расчета баланса на конец периода
    def calculate_final_balance():
        if not st.session_state.orders_data:
            return 0
        
        current_balance = st.session_state.initial_balance
        returned_from_previous_week = 0
        
        # Сортируем данные по месяцам и неделям для правильного порядка
        sorted_orders = []
        for key, value in st.session_state.orders_data.items():
            month, week = key.split('_')
            month_index = months.index(month)
            sorted_orders.append((month_index, int(week), value))
        
        # Сортируем по индексу месяца и номеру недели
        sorted_orders.sort(key=lambda x: (x[0], x[1]))
        
        for month_index, week, orders in sorted_orders:
            if orders > 0:
                # Списываем заказы, добавляем возвращенные с предыдущей недели
                current_balance = current_balance - orders + returned_from_previous_week
                
                # Возвращенные товары придут на следующую неделю
                returned = orders * (1 - return_percentage / 100)
                returned_from_previous_week = returned
        
        return round(current_balance)
    
    min_required_balance = calculate_min_balance()
    final_balance = calculate_final_balance()
    
    # Функция выбора диапазона для анализа
    st.markdown("---")
    st.subheader("📅 Выбор диапазона для анализа")
    
    # Получаем доступные месяцы и недели с правильной сортировкой
    available_periods = []
    period_data = []
    
    for key in st.session_state.orders_data.keys():
        month, week = key.split('_')
        month_index = months.index(month)
        period_data.append({
            'key': key,
            'month': month,
            'month_index': month_index,
            'week': int(week),
            'display': f"{month} - неделя {week}",
            'orders': st.session_state.orders_data[key]
        })
    
    # Сортируем по индексу месяца и номеру недели
    period_data.sort(key=lambda x: (x['month_index'], x['week']))
    
    # Создаем отсортированный список для отображения
    available_periods = [p['display'] for p in period_data]
    
    if available_periods:
        # Красивый дизайн выбора периода
        st.markdown("**🎯 Выберите период для анализа:**")
        
        # Создаем красивый интерфейс с информацией о периодах
        col_info1, col_info2 = st.columns(2)
        
        with col_info1:
            st.info(f"📊 **Доступно периодов:** {len(available_periods)}")
            st.info(f"📦 **Общее количество заказов:** {sum(st.session_state.orders_data.values()):,}")
        
        with col_info2:
            st.info(f"📅 **Первый период:** {available_periods[0]}")
            st.info(f"📅 **Последний период:** {available_periods[-1]}")
        
        # Выбор начального и конечного периода с улучшенным дизайном
        st.markdown("**🔽 Выберите диапазон:**")
        
        col_range1, col_range2 = st.columns(2)
        
        with col_range1:
            # Определяем индекс для начального периода
            start_index = 0
            if st.session_state.selected_start_period and st.session_state.selected_start_period in available_periods:
                start_index = available_periods.index(st.session_state.selected_start_period)
            
            start_period = st.selectbox(
                "📍 Начальный период",
                options=available_periods,
                index=start_index,
                help="Выберите начальный период для анализа"
            )
        
        with col_range2:
            # Определяем индекс для конечного периода
            end_index = len(available_periods)-1
            if st.session_state.selected_end_period and st.session_state.selected_end_period in available_periods:
                end_index = available_periods.index(st.session_state.selected_end_period)
            
            end_period = st.selectbox(
                "📍 Конечный период",
                options=available_periods,
                index=end_index,
                help="Выберите конечный период для анализа"
            )
        
        # Сохраняем выбранный диапазон в session state
        st.session_state.selected_start_period = start_period
        st.session_state.selected_end_period = end_period
        
        # Фильтрация данных по выбранному диапазону
        filtered_orders_data = {}
        start_selected = False
        
        for key, value in st.session_state.orders_data.items():
            month, week = key.split('_')
            period_str = f"{month} - неделя {week}"
            
            if period_str == start_period:
                start_selected = True
            
            if start_selected:
                filtered_orders_data[key] = value
            
            if period_str == end_period:
                break
        
        # Пересчитываем баланс для выбранного диапазона
        if filtered_orders_data:
            # Временно заменяем данные для расчета
            original_data = st.session_state.orders_data.copy()
            st.session_state.orders_data = filtered_orders_data
            
            filtered_min_balance = calculate_min_balance()
            filtered_final_balance = calculate_final_balance()
            
            # Возвращаем оригинальные данные
            st.session_state.orders_data = original_data
            
            # Красивое отображение информации о выбранном периоде
            st.markdown("---")
            st.markdown("**📊 Результаты анализа выбранного периода:**")
            
            col_result1, col_result2, col_result3 = st.columns(3)
            
            with col_result1:
                st.success(f"**📅 Период:**\n{start_period} → {end_period}")
            
            with col_result2:
                st.info(f"**📦 Недель в анализе:** {len(filtered_orders_data)}")
            
            with col_result3:
                st.info(f"**📊 Заказов в периоде:** {sum(filtered_orders_data.values()):,}")
        
        # Визуализация выбранного периода
        st.markdown("**📈 Визуализация выбранного периода:**")
        
        # Создаем данные для графика
        chart_data = []
        for key, value in filtered_orders_data.items():
            month, week = key.split('_')
            chart_data.append({
                'Период': f"{month} - неделя {week}",
                'Заказы': value
            })
        
        if chart_data:
            # Сортируем данные для правильного отображения
            chart_df = pd.DataFrame(chart_data)
            
            # Создаем график
            fig = go.Figure()
            fig.add_trace(go.Bar(
                x=chart_df['Период'],
                y=chart_df['Заказы'],
                marker_color='lightblue',
                name='Заказы'
            ))
            
            fig.update_layout(
                title='Заказы по неделям в выбранном периоде',
                xaxis_title='Период',
                yaxis_title='Количество заказов',
                height=400,
                showlegend=False
            )
            
            st.plotly_chart(fig, use_container_width=True)
        else:
            filtered_min_balance = min_required_balance
            filtered_final_balance = final_balance
            
            st.markdown("---")
            st.markdown("**📊 Результаты анализа:**")
            
            col_result1, col_result2, col_result3 = st.columns(3)
            
            with col_result1:
                st.info(f"**📦 Недель в анализе:** {len(st.session_state.orders_data)}")
            
            with col_result2:
                st.info(f"**📊 Заказов в периоде:** {sum(st.session_state.orders_data.values()):,}")
            
            with col_result3:
                st.info(f"**📅 Анализ:** Полный период")
    else:
        filtered_min_balance = min_required_balance
        filtered_final_balance = final_balance
        
        st.markdown("---")
        st.warning("**⚠️ Нет данных для анализа**")
        st.info("Загрузите данные из основного приложения для начала работы")
    
    # Подсказка о минимальном балансе
    if min_required_balance['simple'] > 0:
        st.markdown("---")
        st.subheader("💡 Анализ необходимого баланса")
        
        # Создаем колонки для разных типов расчетов
        col_analysis1, col_analysis2, col_analysis3, col_analysis4, col_analysis5 = st.columns(5)
        
        with col_analysis1:
            st.metric(
                "📊 Простой расчет", 
                f"{filtered_min_balance['simple']:,}",
                help="Общее количество всех заказов"
            )
        
        with col_analysis2:
            st.metric(
                "🔄 С учетом возвратов", 
                f"{filtered_min_balance['with_returns']:,}",
                help="Заказы минус выкупленные товары"
            )
        
        with col_analysis3:
            st.metric(
                "⚡ Оптимальный баланс", 
                f"{filtered_min_balance['optimal']:,}",
                help="Минимальный баланс для избежания дефицита"
            )
        
        with col_analysis4:
            st.metric(
                "📈 Баланс на конец периода", 
                f"{filtered_final_balance:,}",
                help="Ожидаемый баланс после всех операций"
            )
        
        with col_analysis5:
            if filtered_min_balance['optimal'] > 0:
                st.metric(
                    "💰 Экономия", 
                    f"{filtered_min_balance['simple'] - filtered_min_balance['optimal']:,}",
                    help="Разница между простым и оптимальным расчетом"
                )
            else:
                st.metric(
                    "💰 Экономия", 
                    "0",
                    help="Нет экономии"
                )
        
        # Рекомендация
        if filtered_min_balance['optimal'] > 0:
            recommended_balance = filtered_min_balance['optimal']
            recommendation_type = "оптимальный"
        else:
            recommended_balance = filtered_min_balance['simple']
            recommendation_type = "минимальный"
        
        col_recommendation1, col_recommendation2 = st.columns(2)
        
        with col_recommendation1:
            if initial_balance >= recommended_balance:
                st.success(f"✅ Баланс достаточен! {recommendation_type.capitalize()}: {recommended_balance:,}")
            else:
                st.warning(f"⚠️ {recommendation_type.capitalize()} баланс: **{recommended_balance:,}** товаров")
        
        with col_recommendation2:
            if initial_balance < recommended_balance:
                st.info(f"📊 Текущий: {initial_balance:,} | Необходимо добавить: {recommended_balance - initial_balance:,}")
            else:
                st.info(f"📊 Текущий баланс: {initial_balance:,} товаров")
        
        # Детализация по неделям
        with st.expander("📋 Детализация по неделям"):
            weekly_analysis = []
            cumulative_orders = 0
            
            for month in months:
                for week in range(1, 6):
                    orders_key = f"{month}_{week}"
                    orders = st.session_state.orders_data.get(orders_key, 0)
                    if orders > 0:
                        cumulative_orders += orders
                        weekly_analysis.append({
                            "Месяц": month,
                            "Неделя": week,
                            "Заказы на неделю": orders,
                            "Накопительные заказы": cumulative_orders,
                            "Необходимый баланс": cumulative_orders
                        })
            
            if weekly_analysis:
                df_analysis = pd.DataFrame(weekly_analysis)
                st.dataframe(df_analysis, width='stretch')
                
                # График потребности в товарах
                if len(weekly_analysis) > 1:
                    st.line_chart(df_analysis.set_index("Неделя")["Накопительные заказы"])
            else:
                st.info("Нет данных для анализа. Добавьте заказы в форму ниже.")
    else:
        st.info("💡 Введите заказы ниже, чтобы увидеть анализ необходимого баланса")
    
    # Информация о данных (свернуто)
    with st.expander("📊 Данные заказов из основного приложения", expanded=False):
        if st.session_state.orders_data:
            # Показываем данные в виде таблицы
            data_display = []
            for key, value in st.session_state.orders_data.items():
                month, week = key.split('_')
                data_display.append({
                    "Месяц": month,
                    "Неделя": week,
                    "Заказы": f"{value:,}"
                })
            
            df_display = pd.DataFrame(data_display)
            st.dataframe(df_display, width='stretch')
            
            # Статистика по месяцам
            monthly_stats = {}
            for key, value in st.session_state.orders_data.items():
                month = key.split('_')[0]
                if month not in monthly_stats:
                    monthly_stats[month] = 0
                monthly_stats[month] += value
            
            st.write("**Статистика по месяцам:**")
            for month, total in monthly_stats.items():
                st.write(f"- {month}: {total:,} заказов")
        else:
            st.info("💡 Нажмите 'Обновить данные' для загрузки данных из основного приложения")
    
    # Общая статистика (свернуто)
    if st.session_state.orders_data:
        with st.expander("📈 Общая статистика", expanded=False):
            col_stats1, col_stats2, col_stats3, col_stats4 = st.columns(4)
            
            # Подсчет общей статистики
            total_orders = sum(st.session_state.orders_data.values())
            total_purchased = total_orders * (return_percentage / 100)
            total_returned = total_orders - total_purchased
            
            with col_stats1:
                st.metric("Всего заказов", f"{total_orders:,}")
            
            with col_stats2:
                st.metric("Выкуплено", f"{round(total_purchased):,}")
            
            with col_stats3:
                st.metric("Возвращено", f"{round(total_returned):,}")
            
            with col_stats4:
                st.metric("Процент выкупа", f"{return_percentage:.1f}%")
    
    # Создание таблицы с данными
    st.markdown("---")
    st.subheader("📋 Таблица заказов и баланса (на основе данных из основного приложения)")
    
    def create_orders_table(orders_data=None):
        if orders_data is None:
            orders_data = st.session_state.orders_data
            
        data = []
        current_balance = st.session_state.initial_balance
        returned_from_previous_week = 0  # Возвращенные товары с предыдущей недели
        
        # Сортируем данные по месяцам и неделям для правильного порядка
        sorted_orders = []
        for key, value in orders_data.items():
            month, week = key.split('_')
            month_index = months.index(month)
            sorted_orders.append((month_index, int(week), month, int(week), value))
        
        # Сортируем по индексу месяца и номеру недели
        sorted_orders.sort(key=lambda x: (x[0], x[1]))
        
        for month_index, week_num, month_name, week, orders in sorted_orders:
            if orders > 0:  # Показываем только строки с заказами
                # Расчет выкупленных товаров
                purchased = orders * (return_percentage / 100)
                returned = orders - purchased
                
                # Обновление баланса: списываем заказы, добавляем возвращенные с предыдущей недели
                balance_before = current_balance
                current_balance = current_balance - orders + returned_from_previous_week
                
                # Формируем отображение баланса
                if returned_from_previous_week > 0:
                    balance_display = f"{round(balance_before)}-{orders}+{round(returned_from_previous_week)}={round(current_balance)}"
                else:
                    balance_display = f"{round(balance_before)}-{orders}={round(current_balance)}"
                
                data.append({
                    "Месяц": month_name,
                    "Неделя": week,
                    "Заказано товаров": orders,
                    "Выкуплено": f"{round(purchased):,}",
                    "Возвращено с прошлой недели": f"{round(returned_from_previous_week):,}",
                    "Баланс": balance_display
                })
                
                # Возвращенные товары придут на следующую неделю
                returned_from_previous_week = returned
        
        return pd.DataFrame(data)
    
    # Отображение таблицы
    if st.session_state.orders_data:
        # Определяем какие данные использовать для таблицы
        table_data = filtered_orders_data if 'filtered_orders_data' in locals() and filtered_orders_data else st.session_state.orders_data
        
        df = create_orders_table(table_data)
        if not df.empty:
            st.dataframe(df, width='stretch')
            
            # Экспорт данных
            csv = df.to_csv(index=False, encoding='utf-8-sig')
            st.download_button(
                label="📥 Скачать таблицу (CSV)",
                data=csv,
                file_name=f"заказы_и_баланс_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime="text/csv"
            )
            
            # Дополнительная информация
            if 'filtered_orders_data' in locals() and filtered_orders_data:
                st.info(f"💡 Таблица отображает данные для выбранного периода: {start_period} → {end_period}")
            else:
                st.info("💡 Таблица создана на основе всех данных заказов из основного приложения")
        else:
            st.info("Нет данных для отображения. Нажмите 'Обновить данные' для загрузки.")
    else:
        st.info("Данные отсутствуют. Нажмите 'Обновить данные' для загрузки данных из основного приложения.")
