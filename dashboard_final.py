# -*- coding: utf-8 -*-
import os
import json
import base64
from io import BytesIO
import urllib.parse as _urlparse
import locale

import pandas as pd
import numpy as np
import streamlit as st
import requests
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timedelta
import warnings
warnings.filterwarnings('ignore')

# Импорт Prophet с обработкой ошибок
try:
    from prophet import Prophet
    from prophet.plot import plot_plotly, plot_components_plotly
    PROPHET_AVAILABLE = True
except ImportError:
    PROPHET_AVAILABLE = False
    # Prophet не установлен - функциональность прогнозирования недоступна

# Настройка локали для правильного отображения чисел с пробелами
try:
    locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')
except:
    try:
        locale.setlocale(locale.LC_ALL, 'Russian_Russia.1251')
    except:
        pass

try:
    from PIL import Image
except Exception:
    Image = None

st.set_page_config(page_title="WB Dashboard — Анализ товаров", layout="wide")

# ================= ФУНКЦИИ ДЛЯ РАБОТЫ С ПАРАМЕТРАМИ =================

def save_param_value(sku: str, param: str, value: str):
    """Сохраняет значение параметра для товара"""
    if "param_values" not in st.session_state:
        st.session_state["param_values"] = {}
    if sku not in st.session_state["param_values"]:
        st.session_state["param_values"][sku] = {}
    st.session_state["param_values"][sku][param] = value

def get_param_values():
    """Получает все сохраненные значения параметров"""
    return st.session_state.get("param_values", {})

def save_param_values_to_file():
    """Сохраняет параметры в файл"""
    param_values = get_param_values()
    if param_values:
        try:
            with open("param_values.json", "w", encoding="utf-8") as f:
                json.dump(param_values, f, ensure_ascii=False, indent=2)
            return True
        except Exception:
            return False
    return False

def load_param_values_from_file():
    """Загружает параметры из файла"""
    try:
        if os.path.exists("param_values.json"):
            with open("param_values.json", "r", encoding="utf-8") as f:
                data = json.load(f)
                st.session_state["param_values"] = data
                
                # Создаем param_options на основе загруженных данных
                if "param_options" not in st.session_state:
                    st.session_state["param_options"] = {}
                
                for param_name, param_data in data.items():
                    # Собираем все уникальные значения для каждого параметра
                    unique_values = list(set([v for v in param_data.values() if v and v.strip()]))
                    if unique_values:
                        st.session_state["param_options"][param_name] = sorted(unique_values)
                    elif param_name == "Крой":
                        # Для параметра "Крой" добавляем стандартные варианты, если нет данных
                        st.session_state["param_options"][param_name] = ["Классический", "Приталенный", "Свободный", "Оверсайз"]
                
                return True
    except Exception:
        pass
    return False

# ================= ФУНКЦИИ ДЛЯ ПРОГНОЗИРОВАНИЯ С PROPHET =================

def prepare_data_for_prophet(df, metric_column, date_column=None):
    """Подготавливает данные для Prophet"""
    if not PROPHET_AVAILABLE:
        return None
    
    # Если нет колонки с датами, создаем искусственную временную последовательность
    if date_column is None or date_column not in df.columns:
        # Создаем даты на основе индекса
        start_date = datetime.now() - timedelta(days=len(df)-1)
        dates = [start_date + timedelta(days=i) for i in range(len(df))]
        df_prophet = pd.DataFrame({
            'ds': dates,
            'y': df[metric_column].values
        })
    else:
        # Используем существующую колонку с датами
        df_prophet = pd.DataFrame({
            'ds': pd.to_datetime(df[date_column]),
            'y': df[metric_column].values
        })
    
    # Удаляем строки с NaN значениями
    df_prophet = df_prophet.dropna()
    
    return df_prophet

def create_prophet_forecast(df_prophet, periods=30, seasonality_mode='additive'):
    """Создает прогноз с помощью Prophet"""
    if not PROPHET_AVAILABLE or df_prophet is None or len(df_prophet) < 2:
        return None, None, None
    
    try:
        # Создаем модель Prophet
        model = Prophet(
            seasonality_mode=seasonality_mode,
            daily_seasonality=True,
            weekly_seasonality=True,
            yearly_seasonality=True,
            changepoint_prior_scale=0.05
        )
        
        # Обучаем модель
        model.fit(df_prophet)
        
        # Создаем будущие даты
        future = model.make_future_dataframe(periods=periods)
        
        # Делаем прогноз
        forecast = model.predict(future)
        
        return model, forecast, future
        
    except Exception as e:
        st.error(f"Ошибка при создании прогноза: {e}")
        return None, None, None

def plot_prophet_forecast(model, forecast, title="Прогноз Prophet"):
    """Создает график прогноза с помощью plotly"""
    if not PROPHET_AVAILABLE or model is None or forecast is None:
        return None
    
    try:
        # Создаем график с помощью Prophet
        fig = plot_plotly(model, forecast)
        
        # Обновляем заголовок
        fig.update_layout(
            title=title,
            xaxis_title="Дата",
            yaxis_title="Значение",
            width=1000,
            height=600
        )
        
        return fig
        
    except Exception as e:
        st.error(f"Ошибка при создании графика: {e}")
        return None

def plot_prophet_components(model, forecast, title="Компоненты прогноза"):
    """Создает график компонентов прогноза"""
    if not PROPHET_AVAILABLE or model is None or forecast is None:
        return None
    
    try:
        # Создаем график компонентов
        fig = plot_components_plotly(model, forecast)
        
        # Обновляем заголовок
        fig.update_layout(title=title)
        
        return fig
        
    except Exception as e:
        st.error(f"Ошибка при создании графика компонентов: {e}")
        return None

def save_main_page_data_to_file():
    """Сохраняет данные главной страницы в файл"""
    try:
        main_page_data = {
            "search": st.session_state.get("search", ""),
            "spp": st.session_state.get("spp", 25),
            "buyout_pct": st.session_state.get("buyout_pct", 25),
            "revenue_min": st.session_state.get("revenue_min", 0),
            "revenue_max": st.session_state.get("revenue_max", 1000000),
            "price_min": st.session_state.get("price_min", 0),
            "price_max": st.session_state.get("price_max", 10000),
            "show_images": st.session_state.get("show_images", False),
            "sort_column": st.session_state.get("sort_column", "Выручка"),
            "sort_descending": st.session_state.get("sort_descending", True)
        }
        
        with open("main_page_data.json", "w", encoding="utf-8") as f:
            json.dump(main_page_data, f, ensure_ascii=False, indent=2)
        return True
    except Exception:
        return False

def load_main_page_data_from_file():
    """Загружает данные главной страницы из файла"""
    try:
        if os.path.exists("main_page_data.json"):
            with open("main_page_data.json", "r", encoding="utf-8") as f:
                data = json.load(f)
                
                # Загружаем данные в session_state
                st.session_state["search"] = data.get("search", "")
                st.session_state["spp"] = data.get("spp", 25)
                st.session_state["buyout_pct"] = data.get("buyout_pct", 25)
                st.session_state["revenue_min"] = data.get("revenue_min", 0)
                st.session_state["revenue_max"] = data.get("revenue_max", 1000000)
                st.session_state["price_min"] = data.get("price_min", 0)
                st.session_state["price_max"] = data.get("price_max", 10000)
                st.session_state["show_images"] = data.get("show_images", False)
                st.session_state["sort_column"] = data.get("sort_column", "Выручка")
                st.session_state["sort_descending"] = data.get("sort_descending", True)
                
                return True
    except Exception:
        pass
    return False

# Автоматическая загрузка параметров при запуске приложения
if "param_values" not in st.session_state:
    load_param_values_from_file()

# Автоматическая загрузка данных главной страницы при запуске приложения
if "main_page_data_loaded" not in st.session_state:
    load_main_page_data_from_file()
    st.session_state["main_page_data_loaded"] = True

def _cache_root():
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), "wb_cache")
def _cache_dir():
    d = _cache_root()
    os.makedirs(d, exist_ok=True)
    os.makedirs(os.path.join(d, "imgs"), exist_ok=True)
    return d
def _url_cache_path():
    return os.path.join(_cache_dir(), "image_cache.json")
def load_url_cache():
    p = _url_cache_path()
    if os.path.exists(p):
        try:
            with open(p, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}
def save_url_cache(m: dict):
    try:
        with open(_url_cache_path(), "w", encoding="utf-8") as f:
            json.dump(m, f, ensure_ascii=False, indent=2)
    except Exception:
        pass
def get_url_cache():
    if "img_url_cache" not in st.session_state:
        st.session_state["img_url_cache"] = load_url_cache()
    return st.session_state["img_url_cache"]

def img_path_for(nm: str, fmt: str = "JPEG"):
    nm = str(nm).replace(".0", "")
    ext = "jpg" if fmt.upper() == "JPEG" else "png"
    return os.path.join(_cache_dir(), "imgs", f"{nm}.{ext}")
def get_cached_image_path(nm: str):
    nm = str(nm).replace(".0", "")
    for ext in ("jpg", "png", "jpeg", "webp"):
        p = os.path.join(_cache_dir(), "imgs", f"{nm}.{ext}")
        if os.path.exists(p):
            return p
    return ""
def ensure_image_cached(nm: str, url: str, fmt: str = "JPEG", timeout: int = 25) -> str:
    try:
        p_exist = get_cached_image_path(nm)
        if p_exist:
            return p_exist
        if not url:
            return ""
        path = img_path_for(nm, fmt)
        headers = {"User-Agent": "WB-Dashboard/1.0"}
        with requests.get(url, headers=headers, timeout=timeout, stream=True) as r:
            if r.status_code != 200:
                return ""
            with open(path, "wb") as f:
                for chunk in r.iter_content(8192):
                    if chunk:
                        f.write(chunk)
            return path
    except Exception:
        return ""
def load_image_bytes(path: str, max_w: int | None = None) -> bytes:
    if not path or not os.path.exists(path):
        return b""
    if Image is None or max_w is None:
        try:
            with open(path, "rb") as f:
                return f.read()
        except Exception:
            return b""
    try:
        im = Image.open(path)
        if im.mode not in ("RGB", "RGBA"):
            im = im.convert("RGB")
        if max_w and im.width > max_w:
            ratio = max_w / float(im.width)
            im = im.resize((int(im.width * ratio), int(im.height * ratio)))
        bio = BytesIO()
        im.save(bio, format="JPEG", quality=85)
        return bio.getvalue()
    except Exception:
        try:
            with open(path, "rb") as f:
                return f.read()
        except Exception:
            return b""
def img_data_uri(nm: str, max_w: int | None = None) -> str:
    p = get_cached_image_path(nm)
    if not p:
        return ""
    try:
        data = load_image_bytes(p, max_w=max_w)
        if not data:
            return ""
        b64 = base64.b64encode(data).decode("ascii")
        return f"data:image/jpeg;base64,{b64}"
    except Exception:
        return ""

def build_wb_product_url(nm, host="https://global.wildberries.ru"):
    return f"{host.rstrip('/')}/catalog/{str(nm).replace('.0','')}/detail.aspx"
def build_screenshot_url(page_url: str, key: str,
                         w: int = 400, h: int = 600,
                         fmt: str = "JPEG", profile: str = "D4",
                         base: str = "https://api.s-shot.ru"):
    q = _urlparse.quote(page_url, safe="")
    return f"{base.rstrip('/')}/{int(w)}x{int(h)}/{fmt}/{key}/{profile}/?{q}"
def screenshot_for_article(nm, conf):
    if not conf.get("key"):
        return ""
    page = build_wb_product_url(nm, conf.get("wb_host", "https://global.wildberries.ru"))
    return build_screenshot_url(
        page, conf.get("key", ""), conf.get("w", 400), conf.get("h", 600),
        conf.get("fmt", "JPEG"), conf.get("profile", "D4"), conf.get("base", "https://api.s-shot.ru")
    )

@st.cache_data(show_spinner=False)
def read_table(file_bytes: bytes, filename: str):
    try:
        if filename.lower().endswith((".xlsx", ".xls")):
            df_raw = pd.read_excel(BytesIO(file_bytes), sheet_name=0, header=None)
        else:
            df_raw = pd.read_csv(BytesIO(file_bytes), header=None, sep=None, engine="python")
    except Exception as e:
        st.error(f"Ошибка чтения файла: {e}")
        return None, None, {}
    key_candidates = ["Артикул", "Выручка", "Заказы", "Название"]
    header_row = None
    for i in range(min(30, len(df_raw))):
        vals = df_raw.iloc[i].astype(str).str.strip().tolist()
        if any(k in vals for k in key_candidates):
            header_row = i
            break
    if header_row is None:
        header_row = 0
    if filename.lower().endswith((".xlsx", ".xls")):
        df = pd.read_excel(BytesIO(file_bytes), sheet_name=0, header=header_row)
    else:
        df = pd.read_csv(BytesIO(file_bytes), header=header_row, sep=None, engine="python")
    df = df.loc[:, ~df.columns.astype(str).str.startswith("Unnamed")]
    df = df.loc[:, df.columns.notna()]
    df.columns = [str(c).strip() for c in df.columns]
    rename_map = {
        "Средняя цена без СПП": "Средняя цена",
        "Средняя цена без СПП, ₽": "Средняя цена",
        "Цена": "Средняя цена",
        "Выручка, ₽": "Выручка",
        "Orders": "Заказы",
        "Brand": "Бренд",
        "Supplier": "Поставщик",
        "Subject": "Предмет",
        "Creation date": "Дата создания",
        "Дата": "Дата создания",
        "Позиция": "Позиция в выдаче",
        "CPM": "Стоимость за 1000 показов",
        "Упущенная выручка, ₽": "Упущенная выручка",
    }
    df = df.rename(columns={k: v for k, v in rename_map.items() if k in df.columns})
    num_cols = ["Выручка","Заказы","Средняя цена","Упущенная выручка",
                "Позиция в выдаче","Стоимость за 1000 показов","Буст на позицию","Буст с позиции"]
    for c in num_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(
                df[c].astype(str).str.replace(r"[^\d,.-]", "", regex=True).str.replace(",", ".", regex=False),
                errors="coerce",
            )
    if "Дата создания" in df.columns:
        df["Дата создания"] = pd.to_datetime(df["Дата создания"], errors="coerce")
    if "Тип рекламы" in df.columns:
        df["Тип рекламы"] = df["Тип рекламы"].replace({"b": "Поиск", "c": "Автомат"})
    if ("Буст на позицию" in df.columns) and ("Буст с позиции" in df.columns) and ("Дельта" not in df.columns):
        df["Дельта"] = df["Буст с позиции"] - df["Буст на позицию"]
    return df, df_raw, {"header_row": header_row, "columns": list(df.columns)}

def format_thousands(x, decimals=0):
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return ""
    try:
        xf = float(x)
    except Exception:
        return str(x) if x is not None else ""
    if decimals == 0:
        return f"{int(round(xf))}"
    return f"{xf:.{decimals}f}"

def format_thousands_with_spaces(x, decimals=0):
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return ""
    try:
        xf = float(x)
    except Exception:
        return str(x) if x is not None else ""
    if decimals == 0:
        return f"{int(round(xf)):,}".replace(",", " ")
    return f"{xf:,.{decimals}f}".replace(",", " ").replace(".", ",")
def fmt_rub(x, decimals=0):
    s = format_thousands(x, decimals=decimals)
    return (s + " ₽") if s != "" else ""
def fmt_units(x, unit="шт."):
    s = format_thousands(x, decimals=0)
    return (s + f" {unit}") if s != "" else ""

def fmt_rub_kpi(x, decimals=0):
    s = format_thousands_with_spaces(x, decimals=decimals)
    return (s + " ₽") if s != "" else ""
def fmt_units_kpi(x, unit="шт."):
    s = format_thousands_with_spaces(x, decimals=0)
    return (s + f" {unit}") if s != "" else ""
def fmt_date(d):
    if d is None or (isinstance(d, float) and np.isnan(d)):
        return ""
    try:
        dt = pd.to_datetime(d)
        # Русские названия месяцев
        months = {
            1: "января", 2: "февраля", 3: "марта", 4: "апреля", 5: "мая", 6: "июня",
            7: "июля", 8: "августа", 9: "сентября", 10: "октября", 11: "ноября", 12: "декабря"
        }
        return f"{dt.day} {months[dt.month]} {dt.year}"
    except Exception:
        return str(d) if d is not None else ""
def parse_thousands_input(s, default_val):
    if s is None or str(s).strip() == "":
        return default_val
    try:
        cleaned = (str(s).replace("\\xa0"," ").replace("\\u00a0"," ").replace(" ", " "))
        cleaned = cleaned.replace(" ", "").replace(",", "").strip()
        return int(float(cleaned))
    except Exception:
        return default_val
def sort_df(df, col, asc):
    if col not in df.columns:
        return df
    if pd.api.types.is_numeric_dtype(df[col]):
        return df.sort_values(by=col, ascending=asc, na_position="last", kind="mergesort")
    return df.sort_values(by=col, ascending=asc, na_position="last", kind="mergesort",
                          key=lambda s: s.astype(str).str.lower())

def get_param_schemas():
    if "param_schemas" not in st.session_state:
        st.session_state["param_schemas"] = {}
    return st.session_state["param_schemas"]
def get_param_values():
    if "param_values" not in st.session_state:
        st.session_state["param_values"] = {}
    return st.session_state["param_values"]

def kpi_row(df):
    total_rev = float(df["Выручка"].sum()) if "Выручка" in df.columns else float('nan')
    total_orders = df["Заказы"].sum() if "Заказы" in df.columns else np.nan
    avg_check = (df["Выручка"].sum() / df["Заказы"].sum()) if ("Выручка" in df.columns and "Заказы" in df.columns and df["Заказы"].sum() > 0) else np.nan
    lost_rev = df["Упущенная выручка"].sum() if "Упущенная выручка" in df.columns else np.nan
    sku_count = (df["Артикул"].nunique() if "Артикул" in df.columns else len(df)) if len(df) > 0 else 0
    rev_per_sku = (total_rev / sku_count) if (isinstance(total_rev, (int,float,np.floating)) and not pd.isna(total_rev) and sku_count > 0) else np.nan
    k1, k2, k3, k4, k5, k6 = st.columns(6)
    k1.metric("Выручка (в выборке)", fmt_rub_kpi(total_rev))
    k2.metric("Заказы (в выборке)", fmt_units_kpi(total_orders, "шт."))
    k3.metric("Средний чек", fmt_rub_kpi(avg_check))
    k4.metric("Упущенная выручка", fmt_rub_kpi(lost_rev))
    k5.metric("Выручка / Кол-во товаров", fmt_rub_kpi(rev_per_sku))
    k6.metric("Количество артикулов", fmt_units_kpi(sku_count, "шт."))

# Параметры уже загружены выше при инициализации

# Автоматическая загрузка последней таблицы параметров при запуске
if "table_loaded" not in st.session_state:
    try:
        import json
        import os
        if os.path.exists("table_cache.json"):
            with open("table_cache.json", "r", encoding="utf-8") as f:
                table_cache_data = json.load(f)
            
            # Восстанавливаем данные
            st.session_state["param_values"] = table_cache_data.get("param_values", {})
            st.session_state["param_options"] = table_cache_data.get("param_options", {})
            st.session_state["param_ratings"] = table_cache_data.get("param_ratings", {})
            
            # Удаляем параметр "крой" если он есть в загруженных данных
            if "крой" in st.session_state["param_options"]:
                del st.session_state["param_options"]["крой"]
            if "крой" in st.session_state["param_values"]:
                del st.session_state["param_values"]["крой"]
            
            # Отмечаем, что таблица загружена
            st.session_state["table_loaded"] = True
            
            # Показываем уведомление о загрузке
            st.sidebar.success(f"📂 Автозагрузка: таблица параметров восстановлена")
    except Exception as e:
        st.session_state["table_loaded"] = True  # Отмечаем, что попытка загрузки была

# Автоматическая загрузка последнего файла при запуске
if "file_auto_loaded" not in st.session_state:
    try:
        import json
        import os
        
        # Проверяем наличие кешированного файла
        if os.path.exists("file_cache_meta.json"):
            with open("file_cache_meta.json", "r", encoding="utf-8") as f:
                meta_data = json.load(f)
            
            filename = meta_data.get("filename")
            cache_path = os.path.join("file_cache", filename)
            
            # Если файл существует, устанавливаем флаг для автозагрузки
            if os.path.exists(cache_path):
                st.session_state["auto_load_file"] = True
                st.session_state["file_auto_loaded"] = True
                
                # Показываем уведомление
                st.sidebar.info(f"📂 Найден кешированный файл: {filename}")
            else:
                st.session_state["file_auto_loaded"] = True
        else:
            st.session_state["file_auto_loaded"] = True
    except Exception as e:
        st.session_state["file_auto_loaded"] = True  # Отмечаем, что попытка загрузки была

# Инициализация session_state
if "schemas" not in st.session_state:
    st.session_state["schemas"] = {}

# Функции для работы с кешем файлов
def save_file_cache(file_data, filename):
    """Сохраняет файл в кеш"""
    try:
        import os
        cache_dir = "file_cache"
        if not os.path.exists(cache_dir):
            os.makedirs(cache_dir)
        
        # Сохраняем файл
        cache_path = os.path.join(cache_dir, filename)
        with open(cache_path, "wb") as f:
            f.write(file_data)
        
        # Сохраняем метаданные
        import json
        meta_data = {
            "filename": filename,
            "timestamp": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
            "size": len(file_data)
        }
        
        with open("file_cache_meta.json", "w", encoding="utf-8") as f:
            json.dump(meta_data, f, ensure_ascii=False, indent=2)
        
        return True
    except Exception as e:
        st.error(f"Ошибка сохранения файла в кеш: {e}")
        return False

def get_analysis_period(df, df_raw=None, header_row=None):
    """Извлекает анализируемый период из заголовка таблицы"""
    try:
        # Сначала пытаемся найти период в заголовке таблицы
        if df_raw is not None and header_row is not None:
            # Ищем строку с "Анализируемый период" в заголовке
            for i in range(max(0, header_row - 5), min(len(df_raw), header_row + 1)):
                row_values = df_raw.iloc[i].astype(str).str.strip().tolist()
                
                # Ищем "Анализируемый период" в строке
                for j, cell_value in enumerate(row_values):
                    if "анализируемый период" in cell_value.lower():
                        # Ищем даты в соседних ячейках
                        for k in range(j + 1, min(len(row_values), j + 5)):
                            period_value = row_values[k]
                            if period_value and period_value != "nan":
                                # Пытаемся извлечь даты из строки вида "01.01.2025 - 30.04.2025"
                                import re
                                date_pattern = r'(\d{2}\.\d{2}\.\d{4})\s*-\s*(\d{2}\.\d{2}\.\d{4})'
                                match = re.search(date_pattern, period_value)
                                
                                if match:
                                    start_date_str = match.group(1)
                                    end_date_str = match.group(2)
                                    
                                    try:
                                        # Парсим даты
                                        start_date = pd.to_datetime(start_date_str, format='%d.%m.%Y')
                                        end_date = pd.to_datetime(end_date_str, format='%d.%m.%Y')
                                        
                                        # Форматируем даты для отображения
                                        months = {
                                            1: "января", 2: "февраля", 3: "марта", 4: "апреля", 5: "мая", 6: "июня",
                                            7: "июля", 8: "августа", 9: "сентября", 10: "октября", 11: "ноября", 12: "декабря"
                                        }
                                        
                                        start_date_formatted = f"{start_date.day} {months[start_date.month]} {start_date.year}"
                                        end_date_formatted = f"{end_date.day} {months[end_date.month]} {end_date.year}"
                                        
                                        # Вычисляем количество дней
                                        days_diff = (end_date - start_date).days
                                        
                                        return {
                                            "start_date": start_date,
                                            "end_date": end_date,
                                            "start_date_str": start_date_formatted,
                                            "end_date_str": end_date_formatted,
                                            "days_count": days_diff + 1,
                                            "period_str": f"{start_date_formatted} - {end_date_formatted} ({days_diff + 1} дней)",
                                            "source": "header"
                                        }
                                    except Exception as e:
                                        st.warning(f"Не удалось распарсить даты из заголовка: {e}")
                                        break
                                break
                        break
        
        # Если не нашли в заголовке, пытаемся извлечь из колонки "Дата создания"
        if "Дата создания" in df.columns and not df["Дата создания"].isna().all():
            # Получаем минимальную и максимальную даты
            min_date = df["Дата создания"].dropna().min()
            max_date = df["Дата создания"].dropna().max()
            
            if pd.notna(min_date) and pd.notna(max_date):
                # Форматируем даты для отображения
                months = {
                    1: "января", 2: "февраля", 3: "марта", 4: "апреля", 5: "мая", 6: "июня",
                    7: "июля", 8: "августа", 9: "сентября", 10: "октября", 11: "ноября", 12: "декабря"
                }
                
                min_date_str = f"{min_date.day} {months[min_date.month]} {min_date.year}"
                max_date_str = f"{max_date.day} {months[max_date.month]} {max_date.year}"
                
                # Вычисляем количество дней
                days_diff = (max_date - min_date).days
                
                return {
                    "start_date": min_date,
                    "end_date": max_date,
                    "start_date_str": min_date_str,
                    "end_date_str": max_date_str,
                    "days_count": days_diff + 1,
                    "period_str": f"{min_date_str} - {max_date_str} ({days_diff + 1} дней)",
                    "source": "date_column"
                }
        
        return None
    except Exception as e:
        st.error(f"Ошибка при определении периода анализа: {e}")
        return None

def get_file_statistics(df):
    """Получает статистику по загруженному файлу"""
    try:
        stats = {
            "total_rows": len(df),
            "total_products": df["Артикул"].nunique() if "Артикул" in df.columns else len(df),
            "total_revenue": df["Выручка"].sum() if "Выручка" in df.columns else 0,
            "total_orders": df["Заказы"].sum() if "Заказы" in df.columns else 0,
            "avg_price": df["Средняя цена"].mean() if "Средняя цена" in df.columns else 0,
            "columns_count": len(df.columns)
        }
        
        # Добавляем информацию о колонках
        stats["available_columns"] = list(df.columns)
        
        return stats
    except Exception as e:
        st.error(f"Ошибка при получении статистики файла: {e}")
        return None

def load_file_cache():
    """Загружает файл из кеша"""
    try:
        import json
        import os
        
        # Проверяем наличие метаданных
        if not os.path.exists("file_cache_meta.json"):
            return None, None
        
        # Загружаем метаданные
        with open("file_cache_meta.json", "r", encoding="utf-8") as f:
            meta_data = json.load(f)
        
        filename = meta_data.get("filename")
        cache_path = os.path.join("file_cache", filename)
        
        # Проверяем наличие файла
        if not os.path.exists(cache_path):
            return None, None
        
        # Загружаем файл
        with open(cache_path, "rb") as f:
            file_data = f.read()
        
        return file_data, meta_data
    except Exception as e:
        return None, None

def get_file_cache_info():
    """Получает информацию о кешированном файле"""
    try:
        import json
        import os
        
        if os.path.exists("file_cache_meta.json"):
            with open("file_cache_meta.json", "r", encoding="utf-8") as f:
                meta_data = json.load(f)
            return meta_data
        return None
    except:
        return None

def get_all_cached_files():
    """Получает список всех кешированных файлов"""
    try:
        import os
        cache_dir = "file_cache"
        if not os.path.exists(cache_dir):
            return []
        
        cached_files = []
        for filename in os.listdir(cache_dir):
            if filename.endswith(('.xlsx', '.xls', '.csv')):
                file_path = os.path.join(cache_dir, filename)
                file_size = os.path.getsize(file_path)
                file_time = os.path.getmtime(file_path)
                
                cached_files.append({
                    "filename": filename,
                    "size": file_size,
                    "timestamp": pd.Timestamp.fromtimestamp(file_time).strftime("%Y-%m-%d %H:%M:%S"),
                    "path": file_path
                })
        
        # Сортируем по времени изменения (новые сначала)
        cached_files.sort(key=lambda x: x["timestamp"], reverse=True)
        return cached_files
    except Exception as e:
        st.error(f"Ошибка при получении списка кешированных файлов: {e}")
        return []

def save_file_to_cache(file_data, filename):
    """Сохраняет файл в кеш с улучшенной системой"""
    try:
        import os
        cache_dir = "file_cache"
        if not os.path.exists(cache_dir):
            os.makedirs(cache_dir)
        
        # Сохраняем файл
        cache_path = os.path.join(cache_dir, filename)
        with open(cache_path, "wb") as f:
            f.write(file_data)
        
        # Обновляем метаданные
        meta_data = {
            "filename": filename,
            "timestamp": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
            "size": len(file_data),
            "last_used": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        with open("file_cache_meta.json", "w", encoding="utf-8") as f:
            json.dump(meta_data, f, ensure_ascii=False, indent=2)
        
        return True
    except Exception as e:
        st.error(f"Ошибка сохранения файла в кеш: {e}")
        return False

# --- UI (урезанный пример, ключевые места с прибыль и миниатюрами) ---
with st.sidebar.expander("Загрузка файла", expanded=True):
    # Получаем список всех кешированных файлов
    cached_files = get_all_cached_files()
    
    if cached_files:
        st.write("**📂 Кешированные файлы:**")
        
        # Показываем последний использованный файл
        current_file_info = get_file_cache_info()
        if current_file_info:
            st.info(f"🔄 Текущий: {current_file_info['filename']}\n🕒 {current_file_info['timestamp']}")
        
        # Список всех файлов
        for i, file_info in enumerate(cached_files[:5]):  # Показываем только 5 последних
            col_file, col_load, col_del = st.columns([3, 1, 1])
            
            with col_file:
                file_size_mb = file_info["size"] / (1024 * 1024)
                st.caption(f"📄 {file_info['filename']}\n💾 {file_size_mb:.1f} MB • {file_info['timestamp']}")
            
            with col_load:
                if st.button("📂", key=f"load_{i}"):
                    # Загружаем выбранный файл
                    try:
                        with open(file_info["path"], "rb") as f:
                            file_data = f.read()
                        
                        # Обновляем метаданные
                        meta_data = {
                            "filename": file_info["filename"],
                            "timestamp": file_info["timestamp"],
                            "size": file_info["size"],
                            "last_used": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
                        }
                        
                        with open("file_cache_meta.json", "w", encoding="utf-8") as f:
                            json.dump(meta_data, f, ensure_ascii=False, indent=2)
                        
                        # Сохраняем в session_state
                        st.session_state["cached_file_data"] = file_data
                        st.session_state["cached_file_name"] = file_info["filename"]
                        st.session_state["file_loaded_from_cache"] = True
                        st.session_state["load_from_cache"] = True
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Ошибка загрузки: {e}")
            
            with col_del:
                if st.button("🗑️", key=f"del_{i}"):
                    try:
                        import os
                        os.remove(file_info["path"])
                        st.success(f"✅ Файл {file_info['filename']} удален из кеша")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Ошибка удаления: {e}")
        
        # Кнопка очистки всего кеша
        if st.button("🗑️ Очистить весь кеш файлов", type="secondary"):
            try:
                import os
                import shutil
                
                # Удаляем все файлы кеша
                cache_dir = "file_cache"
                if os.path.exists(cache_dir):
                    shutil.rmtree(cache_dir)
                
                # Удаляем метаданные
                if os.path.exists("file_cache_meta.json"):
                    os.remove("file_cache_meta.json")
                
                # Очищаем session_state
                if "cached_file_data" in st.session_state:
                    del st.session_state["cached_file_data"]
                if "cached_file_name" in st.session_state:
                    del st.session_state["cached_file_name"]
                if "file_loaded_from_cache" in st.session_state:
                    del st.session_state["file_loaded_from_cache"]
                
                st.success("✅ Весь кеш файлов очищен!")
                st.rerun()
            except Exception as e:
                st.error(f"❌ Ошибка очистки кеша: {e}")
    else:
        st.info("📂 Кеш файлов пуст")
    
    # Кнопка для сброса текущего файла (если файл загружен)
    if st.session_state.get("file_loaded_from_cache", False):
        if st.button("🔄 Загрузить новый файл", type="secondary"):
            # Очищаем session_state
            if "cached_file_data" in st.session_state:
                del st.session_state["cached_file_data"]
            if "cached_file_name" in st.session_state:
                del st.session_state["cached_file_name"]
            if "file_loaded_from_cache" in st.session_state:
                del st.session_state["file_loaded_from_cache"]
            st.rerun()
    
    # Всегда показываем кнопку загрузки файла
    st.markdown("---")
    st.markdown("**📤 Загрузить новый файл:**")
    uploaded = st.file_uploader("Excel/CSV с товарами", type=["xlsx","xls","csv"], key="main_uploader")
    
    # Автоматическое сохранение загруженного файла в кеш
    if uploaded is not None:
        file_data = uploaded.read()
        # Возвращаем указатель в начало для дальнейшего чтения
        uploaded.seek(0)
        
        # Сохраняем в кеш с улучшенной системой
        if save_file_to_cache(file_data, uploaded.name):
            st.success(f"💾 Файл автоматически сохранен в кеш: {uploaded.name}")
            
            # Сохраняем информацию в session_state
            st.session_state["cached_file_data"] = file_data
            st.session_state["cached_file_name"] = uploaded.name
            st.session_state["file_loaded_from_cache"] = True
            st.session_state["uploaded_file"] = uploaded
            st.rerun()
    
    # Обработка загрузки из кеша (ручная или автоматическая)
    if st.session_state.get("load_from_cache", False) or st.session_state.get("auto_load_file", False):
        file_data, meta_data = load_file_cache()
        if file_data and meta_data:
            # Создаем объект, похожий на uploaded file
            class CachedFile:
                def __init__(self, data, name):
                    self.data = data
                    self.name = name
                
                def read(self):
                    return self.data
                
                def seek(self, pos):
                    pass  # Заглушка для совместимости
            
            uploaded = CachedFile(file_data, meta_data["filename"])
            
            # Сохраняем информацию о загруженном файле в session_state
            st.session_state["cached_file_data"] = file_data
            st.session_state["cached_file_name"] = meta_data["filename"]
            st.session_state["file_loaded_from_cache"] = True
            
            if st.session_state.get("auto_load_file", False):
                st.success(f"🔄 Автозагрузка: файл восстановлен из кеша - {meta_data['filename']}")
                st.session_state["auto_load_file"] = False
            else:
                st.success(f"✅ Файл загружен из кеша: {meta_data['filename']}")
            
            # Сбрасываем флаги
            st.session_state["load_from_cache"] = False
        else:
            st.error("❌ Не удалось загрузить файл из кеша")
            st.session_state["load_from_cache"] = False
            st.session_state["auto_load_file"] = False
            uploaded = None
    else:
        # Проверяем, есть ли сохраненный файл в session_state
        if st.session_state.get("file_loaded_from_cache", False) and st.session_state.get("cached_file_data"):
            # Восстанавливаем файл из session_state
            class CachedFile:
                def __init__(self, data, name):
                    self.data = data
                    self.name = name
                
                def read(self):
                    return self.data
                
                def seek(self, pos):
                    pass  # Заглушка для совместимости
            
            uploaded = CachedFile(st.session_state["cached_file_data"], st.session_state["cached_file_name"])
        else:
            # Используем файл, загруженный через основную кнопку загрузки
            uploaded = st.session_state.get("uploaded_file", None)
with st.sidebar.expander("Скриншоты страниц (s-shot.ru)", expanded=True):
    sc_key = st.text_input("Ключ s-shot", value="KEYSV7S9IWCFGI50SA8")
    sc_base = st.text_input("Базовый URL", value="https://api.s-shot.ru")
    sc_host = st.text_input("Домен карточки WB", value="https://global.wildberries.ru")
    sc_w = st.number_input("Ширина", 100, 2000, 400, 10)
    sc_h = st.number_input("Высота", 100, 2000, 600, 10)
    sc_fmt = st.selectbox("Формат", ["JPEG","PNG"], 0)
    sc_profile = st.text_input("Профиль", value="D4")
    
    # Информация о кеше
    url_cache = get_url_cache()
    cached_count = len(url_cache)
    st.info(f"📦 В кеше: {cached_count} изображений")
    
    # Кнопки управления кешем
    col_cache1, col_cache2 = st.columns(2)
    if col_cache1.button("🗑️ Очистить кеш"):
        st.session_state["img_url_cache"] = {}
        save_url_cache({})
        st.rerun()
    
    if col_cache2.button("💾 Сохранить параметры"):
        if save_param_values_to_file():
            st.success("✅ Параметры сохранены!")
        else:
            st.error("❌ Ошибка сохранения параметров")
    
    # Кнопка сохранения данных главной страницы
    if st.button("💾 Сохранить настройки главной страницы", width='stretch'):
        # Сохраняем текущие значения в session_state
        st.session_state["search"] = st.session_state.get("search_input", "")
        st.session_state["spp"] = st.session_state.get("spp_input", 25)
        st.session_state["buyout_pct"] = st.session_state.get("buyout_input", 25)
        st.session_state["revenue_min"] = st.session_state.get("revenue_min_input", 0)
        st.session_state["revenue_max"] = st.session_state.get("revenue_max_input", 1000000)
        st.session_state["price_min"] = st.session_state.get("price_min_input", 0)
        st.session_state["price_max"] = st.session_state.get("price_max_input", 10000)
        
        if save_main_page_data_to_file():
            st.success("✅ Настройки главной страницы сохранены!")
        else:
            st.error("❌ Ошибка сохранения настроек")
    
    st.divider()
    
    # Управление таблицей параметров
    st.write("**Управление таблицей параметров:**")
    
    # Кнопка сохранения
    if st.button("💾 Сохранить таблицу в кеш", width='stretch'):
            # Сохраняем текущую таблицу параметров
            table_cache_data = {
                "param_values": st.session_state.get("param_values", {}),
                "param_options": st.session_state.get("param_options", {}),
                "timestamp": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            
            # Сохраняем в файл
            try:
                import json
                with open("table_cache.json", "w", encoding="utf-8") as f:
                    json.dump(table_cache_data, f, ensure_ascii=False, indent=2)
                st.success("✅ Таблица сохранена в кеш!")
            except Exception as e:
                st.error(f"❌ Ошибка сохранения: {e}")
    
    # Кнопка загрузки
    if st.button("📂 Загрузить таблицу из кеша", width='stretch'):
            # Загружаем последнюю сохраненную таблицу
            try:
                import json
                import os
                if os.path.exists("table_cache.json"):
                    with open("table_cache.json", "r", encoding="utf-8") as f:
                        table_cache_data = json.load(f)
                    
                    # Восстанавливаем данные
                    st.session_state["param_values"] = table_cache_data.get("param_values", {})
                    st.session_state["param_options"] = table_cache_data.get("param_options", {})
                    st.session_state["param_ratings"] = table_cache_data.get("param_ratings", {})
                    
                    # Удаляем параметр "крой" если он есть в загруженных данных
                    if "крой" in st.session_state["param_options"]:
                        del st.session_state["param_options"]["крой"]
                    if "крой" in st.session_state["param_values"]:
                        del st.session_state["param_values"]["крой"]
                    
                    timestamp = table_cache_data.get("timestamp", "неизвестно")
                    st.success(f"✅ Таблица загружена! (сохранена: {timestamp})")
                    st.rerun()
                else:
                    st.warning("Кеш таблицы не найден")
            except Exception as e:
                st.error(f"❌ Ошибка загрузки: {e}")
    
    # Кнопка очистки
    if st.button("🗑️ Очистить кеш таблицы", width='stretch'):
            # Очищаем кеш таблицы параметров
            try:
                import os
                if os.path.exists("table_cache.json"):
                    os.remove("table_cache.json")
                    st.success("✅ Кеш таблицы очищен!")
                else:
                    st.warning("Кеш таблицы не найден")
                st.rerun()
            except Exception as e:
                st.error(f"❌ Ошибка очистки кеша: {e}")
    
    # Информация о кеше таблицы
    try:
        import json
        import os
        if os.path.exists("table_cache.json"):
            with open("table_cache.json", "r", encoding="utf-8") as f:
                table_cache_data = json.load(f)
            timestamp = table_cache_data.get("timestamp", "неизвестно")
            param_count = len(table_cache_data.get("param_options", {}))
            st.info(f"📦 Кеш таблицы: {param_count} параметров, сохранен {timestamp}")
    except:
        pass

if uploaded is None:
    st.info("Загрузите файл с данными.")
else:
    df, raw, meta = read_table(uploaded.read(), uploaded.name)
    if df is None or df.empty:
        st.error("Не удалось прочитать таблицу.")
    else:
        st.title("📊 Дашборд WB")
        
        # Получаем анализируемый период (для внутреннего использования)
        analysis_period = get_analysis_period(df, raw, meta.get("header_row"))
        
        # Отображаем информацию о периоде анализа
        if analysis_period:
            source_text = "из заголовка таблицы" if analysis_period.get("source") == "header" else "из колонки 'Дата создания'"
            st.success(f"📅 **Анализируемый период:** {analysis_period['period_str']} ({source_text})")
        
        # Импортируем модуль анализа сезонности
        try:
            from seasonality_module import (
                load_seasonality_data, clean_seasonality_data, create_seasonality_graph,
                get_status_stats, style_dataframe, load_custom_data, create_manual_entry_data
            )
            seasonality_available = True
        except ImportError:
            seasonality_available = False
        
        # Создаем вкладки
        if seasonality_available:
            if PROPHET_AVAILABLE:
                tab1, tab2, tab3, tab4, tab5 = st.tabs(["📊 Анализ данных", "⚙️ Установка параметров", "📈 Аналитика по параметрам", "📅 Анализ сезонности", "🔮 Прогнозирование"])
            else:
                tab1, tab2, tab3, tab4 = st.tabs(["📊 Анализ данных", "⚙️ Установка параметров", "📈 Аналитика по параметрам", "📅 Анализ сезонности"])
        else:
            if PROPHET_AVAILABLE:
                tab1, tab2, tab3, tab4 = st.tabs(["📊 Анализ данных", "⚙️ Установка параметров", "📈 Аналитика по параметрам", "🔮 Прогнозирование"])
            else:
                tab1, tab2, tab3 = st.tabs(["📊 Анализ данных", "⚙️ Установка параметров", "📈 Аналитика по параметрам"])
        
        with tab1:
            # Основные фильтры
            col1, col2, col3, col4 = st.columns(4)
        
            search = col1.text_input("🔍 Поиск", value=st.session_state.get("search", ""), key="search_input")
            spp = col2.number_input("💰 СПП, %", 0, 100, st.session_state.get("spp", 25), 1, key="spp_input")
            buyout_pct = col3.number_input("📈 Процент выкупа, %", 1, 100, st.session_state.get("buyout_pct", 25), 1, key="buyout_input")
            
            # Кнопка обновления данных
            col4.markdown("🔄 Обновить данные")
            if col4.button("Обновить", type="primary"):
                st.rerun()
            
            # Фильтр по предмету на отдельной строчке
            if "Предмет" in df.columns:
                subjects = sorted(df["Предмет"].dropna().unique())
                selected_subjects = st.multiselect("📦 Предмет", subjects, default=subjects)
            else:
                selected_subjects = []
            
            # Фильтры по параметрам товаров
            param_values = get_param_values()
            selected_param_filters = {}
            
            if param_values:
                # Переключатель для включения/выключения фильтров по параметрам
                enable_param_filters = st.checkbox(
                    "🎨 Включить фильтры по параметрам", 
                    value=False,
                    help="Включить фильтрацию товаров по их параметрам (цвет, материал и т.д.)"
                )
                
                if enable_param_filters:
                    st.subheader("🎨 Фильтры по параметрам")
                    
                    # Создаем колонки для фильтров параметров
                    param_cols = st.columns(min(len(param_values), 4))  # Максимум 4 колонки
                    
                    for i, (param_name, param_data) in enumerate(param_values.items()):
                        col_idx = i % 4
                        
                        with param_cols[col_idx]:
                            # Получаем уникальные значения для этого параметра (убираем дубликаты)
                            unique_values = sorted(list(set([v for v in param_data.values() if v and v.strip()])))
                            
                            if unique_values:
                                # Создаем multiselect для каждого параметра
                                selected_values = st.multiselect(
                                    f"🎨 {param_name}",
                                    unique_values,
                                    default=[],  # По умолчанию ничего не выбрано
                                    help=f"Выберите значения для параметра '{param_name}'"
                                )
                                selected_param_filters[param_name] = selected_values
                            else:
                                selected_param_filters[param_name] = []
            
            # Настройки отображения
            col_img1, col_img2, col_img3 = st.columns(3)
            show_images = col_img1.checkbox("🖼️ Показывать изображения", value=False)
            if show_images:
                img_size = col_img2.number_input("📏 Размер миниатюр (px)", min_value=50, max_value=300, value=200, step=10)
                if col_img3.button("🔄 Обновить кеш изображений", type="secondary"):
                    # Очищаем кеш URL
                    st.session_state["img_url_cache"] = {}
                    save_url_cache({})
                    st.rerun()
                if not sc_key:
                    st.info("💡 Для отображения изображений товаров введите API ключ s-shot.ru в боковой панели")
            else:
                img_size = 200
            
            # Фильтр сортировки
            col_sort1, col_sort2 = st.columns(2)
            
            # Определяем доступные колонки для сортировки
            sortable_columns = []
            if "Выручка" in df.columns:
                sortable_columns.append("Выручка")
            if "Средняя цена" in df.columns:
                sortable_columns.append("Средняя цена")
            if "Заказы" in df.columns:
                sortable_columns.append("Заказы")
            if "Дата создания" in df.columns:
                sortable_columns.append("Дата создания")
            if "Прибыль" in df.columns:
                sortable_columns.append("Прибыль")
            if "Упущенная выручка" in df.columns:
                sortable_columns.append("Упущенная выручка")
            if "Позиция в выдаче" in df.columns:
                sortable_columns.append("Позиция в выдаче")
            
            # Добавляем опцию "Без сортировки"
            sortable_columns.insert(0, "Без сортировки")
            
            # Находим индекс "Выручка" для установки по умолчанию
            default_index = 0  # По умолчанию "Без сортировки"
            if "Выручка" in sortable_columns:
                default_index = sortable_columns.index("Выручка")
            
            sort_column = col_sort1.selectbox("📊 Сортировка по", sortable_columns, index=default_index)
            sort_ascending = col_sort2.selectbox("🔽 Направление", ["По убыванию", "По возрастанию"], index=0) == "По возрастанию"
            
            st.divider()
            
            # Выручка
            col7, col8 = st.columns(2)
            
            if "Выручка" in df.columns:
                default_revenue_max = int(df["Выручка"].max()) if not df["Выручка"].isna().all() else 1000000
                revenue_min = col7.number_input("Выручка от", min_value=0, value=st.session_state.get("revenue_min", 0), step=1000, key="revenue_min_input")
                revenue_max = col8.number_input("Выручка до", min_value=0, value=st.session_state.get("revenue_max", default_revenue_max), step=1000, key="revenue_max_input")
            else:
                revenue_min = st.session_state.get("revenue_min", 0)
                revenue_max = st.session_state.get("revenue_max", 1000000)
            
            # Цена
            col9, col10 = st.columns(2)
            
            if "Средняя цена" in df.columns:
                default_price_max = int(df["Средняя цена"].max()) if not df["Средняя цена"].isna().all() else 10000
                price_min = col9.number_input("Цена (до СПП) от", min_value=0, value=st.session_state.get("price_min", 0), step=100, key="price_min_input")
                price_max = col10.number_input("Цена (до СПП) до", min_value=0, value=st.session_state.get("price_max", default_price_max), step=100, key="price_max_input")
            else:
                price_min = st.session_state.get("price_min", 0)
                price_max = st.session_state.get("price_max", 10000)
            
            # Дата создания
            
            # Определяем диапазон дат для фильтрации
            # Всегда используем полный диапазон дат создания для фильтра
            if "Дата создания" in df.columns:
                # Используем даты из колонки "Дата создания"
                date_range = df["Дата создания"].dropna()
                if not date_range.empty:
                    min_date = date_range.min().date()
                    max_date = date_range.max().date()
                    date_source = "колонка 'Дата создания'"
                else:
                    min_date = pd.Timestamp.now().date()
                    max_date = pd.Timestamp.now().date()
                    date_source = "не найдены"
                    st.warning("⚠️ В данных нет информации о датах")
            else:
                min_date = pd.Timestamp.now().date()
                max_date = pd.Timestamp.now().date()
                date_source = "не найдены"
                st.warning("⚠️ Колонка 'Дата создания' не найдена в данных")
            

            
            # Создаем ползунок для выбора диапазона дат
            date_range_days = (max_date - min_date).days
            if date_range_days > 0:
                date_slider = st.slider(
                    "📅 Выберите диапазон дат для фильтрации",
                    min_value=min_date,
                    max_value=max_date,
                    value=(min_date, max_date),
                    help="Выберите период для анализа. По умолчанию показан полный диапазон дат создания."
                )
                date_min, date_max = date_slider
            else:
                date_min = min_date
                date_max = max_date
                st.info(f"📅 Данные за один день: {min_date.strftime('%d.%m.%Y')}")
            fdf = df.copy()
            
            # Применяем фильтры
            if search:
                mask = fdf.apply(lambda x: x.astype(str).str.contains(search, case=False, na=False)).any(axis=1)
                fdf = fdf[mask]
            
            if selected_subjects and "Предмет" in fdf.columns:
                fdf = fdf[fdf["Предмет"].isin(selected_subjects)]
            
            # Фильтр по выручке
            if "Выручка" in fdf.columns:
                fdf = fdf[(fdf["Выручка"] >= revenue_min) & (fdf["Выручка"] <= revenue_max)]
            
            # Фильтр по цене
            if "Средняя цена" in fdf.columns:
                fdf = fdf[(fdf["Средняя цена"] >= price_min) & (fdf["Средняя цена"] <= price_max)]
            
            # Фильтр по дате создания
            if "Дата создания" in fdf.columns:
                fdf = fdf[(fdf["Дата создания"].dt.date >= date_min) & (fdf["Дата создания"].dt.date <= date_max)]
            
            # Фильтры по параметрам товаров
            if param_values and selected_param_filters and enable_param_filters:
                for param_name, selected_values in selected_param_filters.items():
                    if selected_values:  # Если выбраны значения для фильтрации
                        # Находим артикулы, которые соответствуют выбранным значениям параметра
                        matching_skus = []
                        if param_name in param_values:
                            for sku, value in param_values[param_name].items():
                                if value in selected_values:
                                    matching_skus.append(sku)
                        
                        if matching_skus:
                            # Фильтруем данные по найденным артикулам
                            mask = fdf["Артикул"].astype(str).str.replace(".0", "").isin(matching_skus)
                            fdf = fdf[mask]
            
            if "Средняя цена" in fdf.columns:
                fdf["Цена (с СПП)"] = fdf["Средняя цена"] * (1 - float(spp)/100.0)
            buyout_k = float(buyout_pct)/100.0 if buyout_pct else 0.0
            if "Заказы" in fdf.columns:
                fdf["Выкупы"] = pd.to_numeric(fdf["Заказы"], errors="coerce") * buyout_k
            else:
                fdf["Выкупы"] = np.nan
            # === FIX: Прибыль = Выручка * (процент выкупа) ===
            if "Выручка" in fdf.columns and buyout_k > 0:
                fdf["Прибыль"] = pd.to_numeric(fdf["Выручка"], errors="coerce") * buyout_k
            else:
                fdf["Прибыль"] = np.nan
            
            # Применяем сортировку
            if sort_column and sort_column != "Без сортировки" and sort_column in fdf.columns:
                fdf = sort_df(fdf, sort_column, sort_ascending)
            
            kpi_row(fdf)
            st.divider()

            # Миниатюры кэш
            url_cache = get_url_cache()
            
            # Создаем копию для отображения
            display_df = fdf.copy()
            
            # Добавляем изображения
            if show_images and "Артикул" in display_df.columns:
                imgs = []
                loaded_count = 0
                total_items = len(display_df)
                
                # Создаем прогресс-бар если много товаров
                if total_items > 10:
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                
                for i, a in enumerate(display_df["Артикул"].astype(str)):
                    k = a.replace(".0","")
                    url = url_cache.get(k, "")
                    if not url and sc_key:
                        url = screenshot_for_article(k, {"key": sc_key,"w": sc_w,"h": sc_h,"fmt": sc_fmt,"profile": sc_profile,"base": sc_base,"wb_host": sc_host})
                        if url:
                            url_cache[k] = url
                            save_url_cache(url_cache)
                    
                    # Проверяем кеш изображений
                    cached_path = get_cached_image_path(k)
                    if cached_path and os.path.exists(cached_path):
                        path = cached_path
                    elif url:
                        path = ensure_image_cached(k, url, sc_fmt)
                    else:
                        path = ""
                    
                    if path and os.path.exists(path):
                        # Создаем data URI для Streamlit
                        img_bytes = load_image_bytes(path, img_size)
                        if img_bytes:
                            b64_data = base64.b64encode(img_bytes).decode()
                            data_uri = f"data:image/jpeg;base64,{b64_data}"
                            imgs.append(data_uri)
                            loaded_count += 1
                        else:
                            imgs.append("")
                    else:
                        imgs.append("")
                    
                    # Обновляем прогресс
                    if total_items > 10:
                        progress = (i + 1) / total_items
                        progress_bar.progress(progress)
                        status_text.text(f"Загружаем изображения: {i + 1}/{total_items}")
                
                # Очищаем прогресс-бар
                if total_items > 10:
                    progress_bar.empty()
                    status_text.empty()
                
                display_df.insert(1, "Изображение", imgs)
                
                # Показываем статистику загрузки
                st.success(f"📊 Загружено изображений: {loaded_count} из {len(display_df)} товаров")
            
            # Добавляем отдельный столбец со ссылками на артикулы
            if "Артикул" in display_df.columns:
                # Создаем специальный формат для отображения "Открыть" в Streamlit
                display_df["Ссылка"] = "Открыть"
                # Но фактические ссылки будут настроены через column_config
            
            # Форматирование даты для Streamlit таблицы
            if "Дата создания" in display_df.columns:
                display_df["Дата создания"] = display_df["Дата создания"].apply(fmt_date)
            
            # Оставляем числовые данные как есть для корректной сортировки в Streamlit таблице
            
            # Изменение порядка столбцов
            desired_order = [
                "Артикул", "Ссылка", "Дата создания", "Выручка", "Заказы", "Выкупы", 
                "Средняя цена", "Цена (с СПП)", "Упущенная выручка", "Прибыль",
                "Предмет", "Позиция в выдаче", "Стоимость за 1000 показов", 
                "Тип рекламы", "Буст на позицию", "Буст с позиции", "Дельта",
                "Название", "Поставщик", "Бренд"
            ]
            
            # Добавляем изображения в начало если нужно
            if show_images and "Изображение" in display_df.columns:
                desired_order.insert(1, "Изображение")
            
            # Информация о загруженных параметрах
            if param_values:
                total_params = sum(len(param_data) for param_data in param_values.values())
                st.info(f"📊 Загружено параметров: {list(param_values.keys())} ({total_params} значений)")
            
            # Получаем список всех параметров (исключаем "крой")
            all_params = list(param_values.keys())
            if "param_options" in st.session_state:
                all_params.extend([p for p in st.session_state["param_options"].keys() if p not in all_params])
            
            # Исключаем столбец "крой" из отображения
            all_params = [param for param in all_params if param.lower() != "крой"]
            
            # Добавляем столбцы параметров в DataFrame
            for param in all_params:
                param_column_data = []
                for sku in display_df["Артикул"].astype(str):
                    sku_clean = sku.replace(".0", "")
                    param_value = param_values.get(param, {}).get(sku_clean, "")
                    param_column_data.append(param_value)
                display_df[param] = param_column_data
            
            # Переупорядочиваем столбцы - параметры после "Дата создания"
            existing_cols = [col for col in desired_order if col in display_df.columns]
            
            # Находим позицию "Дата создания" в desired_order
            date_creation_index = -1
            if "Дата создания" in desired_order:
                date_creation_index = desired_order.index("Дата создания")
            
            # Разделяем столбцы на основные и параметры
            main_cols = [col for col in existing_cols if col not in all_params]
            param_cols = [col for col in all_params if col in display_df.columns]
            other_cols = [col for col in display_df.columns if col not in existing_cols and col not in all_params]
            
            # Если "Дата создания" найдена, вставляем параметры после неё
            if date_creation_index >= 0 and "Дата создания" in main_cols:
                date_index = main_cols.index("Дата создания")
                # Вставляем параметры после "Дата создания"
                final_order = main_cols[:date_index+1] + param_cols + main_cols[date_index+1:] + other_cols
            else:
                # Если "Дата создания" не найдена, добавляем параметры в конец основных столбцов
                final_order = main_cols + param_cols + other_cols
            
            display_df = display_df[final_order]

            from streamlit import column_config as cc
            
            # Настройка конфигурации столбцов для лучшего отображения
            col_cfg = {}
            
            # Конфигурация для изображений
            if "Изображение" in display_df.columns:
                col_cfg["Изображение"] = cc.ImageColumn("Изображение", width=img_size + 20)
            
            # Конфигурация для артикула (числовой тип)
            if "Артикул" in display_df.columns:
                col_cfg["Артикул"] = cc.NumberColumn("Артикул", format="%.0f", width=120)
            
            # Конфигурация для ссылки на товар с динамическими URL
            if "Ссылка" in display_df.columns and "Артикул" in display_df.columns:
                # Создаем ссылки на основе артикулов
                links_data = []
                for sku in display_df["Артикул"].astype(str):
                    sku_clean = sku.replace(".0", "")
                    links_data.append(f"https://global.wildberries.ru/catalog/{sku_clean}/detail.aspx")
                display_df["Ссылка"] = links_data
                # Явно указываем тип данных как строковый
                display_df["Ссылка"] = display_df["Ссылка"].astype(str)
                col_cfg["Ссылка"] = cc.LinkColumn("Ссылка", display_text="🔗", width=60)
            
            # Конфигурация для числовых столбцов (NumberColumn для корректной сортировки)
            money_columns = ["Выручка", "Средняя цена", "Цена (с СПП)", "Упущенная выручка", "Прибыль"]
            for col in money_columns:
                if col in display_df.columns:
                    col_cfg[col] = cc.NumberColumn(col, format="%.0f ₽", width=120)
            
            # Конфигурация для числовых столбцов с единицами
            if "Заказы" in display_df.columns:
                col_cfg["Заказы"] = cc.NumberColumn("Заказы", format="%.0f шт.", width=120)
            if "Выкупы" in display_df.columns:
                col_cfg["Выкупы"] = cc.NumberColumn("Выкупы", format="%.0f шт.", width=120)
            
            # Конфигурация для даты - отключаем редактирование чтобы избежать проблем с типами
            if "Дата создания" in display_df.columns:
                col_cfg["Дата создания"] = cc.TextColumn("Дата создания", width=150, disabled=True)
            
    # Конфигурация для параметров товаров
    for param in all_params:
        if param in display_df.columns:
            if param in st.session_state.get("param_options", {}):
                # Selectbox для параметров с вариантами
                options = [""] + st.session_state["param_options"][param]
                col_cfg[param] = cc.SelectboxColumn(
                    param, 
                    options=options, 
                    width=150
                )
            else:
                # Обычный текст для свободных параметров
                col_cfg[param] = cc.TextColumn(param, width=150)
            
    # Отображаем редактируемую таблицу с возможностью сортировки
    edited_df = st.data_editor(
                display_df, 
        width='stretch', 
                hide_index=True, 
                column_config=col_cfg,
        column_order=None,  # Позволяет пользователю переупорядочивать столбцы
        num_rows="fixed",
        key="main_table_editor"
    )
            
    # Сохраняем изменения параметров обратно в param_values
    if all_params:
        changes_made = False
        for index, row in edited_df.iterrows():
            # Обрабатываем артикул как число, но конвертируем в строку для ключей
            sku_raw = row["Артикул"]
            if pd.isna(sku_raw):
                continue
            sku = str(int(sku_raw)) if isinstance(sku_raw, (int, float)) else str(sku_raw)
            
            for param in all_params:
                if param in row and row[param]:
                    if param not in param_values:
                        param_values[param] = {}
                    if sku not in param_values[param] or param_values[param][sku] != str(row[param]):
                        param_values[param][sku] = str(row[param])
                        changes_made = True
                elif param in param_values and sku in param_values[param]:
                    # Удаляем пустые значения
                    if not row.get(param):
                        del param_values[param][sku]
                        changes_made = True
        
        # Показываем уведомление об изменениях
        if changes_made:
            st.success("✅ Изменения параметров сохранены!")
                    
    # Кнопка для сохранения в файл
    col_save1, col_save2 = st.columns([1, 4])
    with col_save1:
        if st.button("💾 Сохранить в файл", type="primary"):
            if save_param_values_to_file():
                st.success("✅ Параметры сохранены в файл!")
            else:
                st.error("❌ Ошибка сохранения в файл")
        
        with tab2:
            st.subheader("⚙️ Установка параметров товаров")
            
            # Кнопка для принудительной загрузки параметров
            if st.button("🔄 Перезагрузить параметры", key="reload_params_tab2"):
                load_param_values_from_file()
                st.rerun()
            
            # Получаем данные
            param_values = get_param_values()
            
            # Инициализация session_state для автосохранения (отключено)
            if "last_autosave" not in st.session_state:
                st.session_state["last_autosave"] = 0
            
            # Автосохранение отключено
            # import time
            # current_time = time.time()
            # if current_time - st.session_state["last_autosave"] > 60:  # 60 секунд
            #     save_param_values_to_file()
            #     st.session_state["last_autosave"] = current_time
            #     st.info("🔄 Автосохранение выполнено")
            
            # Добавление параметров с отдельными полями
            st.write("**Добавить параметр:**")
            
            # Готовые шаблоны параметров
            param_templates = {
                "Цвет": ["Красный", "Синий", "Зеленый", "Черный", "Белый", "Желтый", "Оранжевый", "Фиолетовый", "Розовый", "Серый"],
                "Длина": ["Короткая", "Средняя", "Длинная", "Мини", "Макси", "Миди", "Анкл"],
                "Пуговицы": ["С пуговицами", "Без пуговиц", "На молнии", "На липучке", "На кнопках", "На шнурке"],
                "Материал": ["Хлопок", "Полиэстер", "Шерсть", "Лен", "Джинс", "Кожа", "Замша", "Трикотаж"],
                "Размер": ["XS", "S", "M", "L", "XL", "XXL", "XXXL"],
                "Сезон": ["Лето", "Зима", "Весна", "Осень", "Демисезон"],
                "Стиль": ["Классический", "Спортивный", "Повседневный", "Деловой", "Вечерний", "Романтический"]
            }
            
            # Выбор шаблона или создание нового
            template_choice = st.selectbox(
                "Выберите готовый шаблон или создайте новый",
                ["Создать новый параметр"] + list(param_templates.keys()),
                index=0,
                key="template_selector"
            )
            
            # Автоматическое заполнение при выборе шаблона
            if template_choice != "Создать новый параметр":
                # Автоматически заполняем поля при выборе шаблона
                if "current_template" not in st.session_state or st.session_state["current_template"] != template_choice:
                    st.session_state["current_template"] = template_choice
                    st.session_state["temp_param_name"] = template_choice
                    st.session_state["temp_param_options"] = " / ".join(param_templates[template_choice])
            else:
                # Очищаем поля при выборе "Создать новый параметр"
                if "current_template" in st.session_state and st.session_state["current_template"] != "Создать новый параметр":
                    st.session_state["current_template"] = "Создать новый параметр"
                    if "temp_param_name" in st.session_state:
                        del st.session_state["temp_param_name"]
                    if "temp_param_options" in st.session_state:
                        del st.session_state["temp_param_options"]
            
            # Поля для ввода параметра
            col_param1, col_param2 = st.columns(2)
            
            with col_param1:
                param_name = st.text_input(
                    "Название параметра",
                    value=st.session_state.get("temp_param_name", ""),
                    placeholder="Например: Цвет, Длина, Материал",
                    key="param_name_input"
                )
            
            with col_param2:
                param_options = st.text_area(
                    "Варианты (через слэш /)",
                    value=st.session_state.get("temp_param_options", ""),
                    placeholder="Красный / Синий / Зеленый",
                    height=100,
                    key="param_options_input"
                )
            
            # Кнопки управления
            col_add, col_clear = st.columns([2, 1])
            
            with col_add:
                if st.button("➕ Добавить параметр", type="primary"):
                    if param_name and param_options:
                        try:
                            # Очищаем и разбираем варианты
                            options = [opt.strip() for opt in param_options.split("/") if opt.strip()]
                            
                            if options:
                                # Создаем новый параметр
                                if param_name not in param_values:
                                    param_values[param_name] = {}
                                
                                st.success(f"✅ Параметр '{param_name}' добавлен с вариантами: {', '.join(options)}")
                                
                                # Сохраняем варианты в session_state для использования в таблице
                                if "param_options" not in st.session_state:
                                    st.session_state["param_options"] = {}
                                st.session_state["param_options"][param_name] = options
                                
                                # Очищаем временные поля
                                if "temp_param_name" in st.session_state:
                                    del st.session_state["temp_param_name"]
                                if "temp_param_options" in st.session_state:
                                    del st.session_state["temp_param_options"]
                            else:
                                st.warning("Добавьте хотя бы один вариант")
                        except Exception as e:
                            st.error(f"❌ Ошибка при создании параметра: {e}")
                    else:
                        st.warning("Заполните название параметра и варианты")
            
            with col_clear:
                if st.button("🗑️ Очистить все"):
                    st.session_state["param_values"] = {}
                    st.session_state["param_options"] = {}
                    if "temp_param_name" in st.session_state:
                        del st.session_state["temp_param_name"]
                    if "temp_param_options" in st.session_state:
                        del st.session_state["temp_param_options"]
                    save_param_values_to_file()
                    st.success("✅ Все параметры удалены!")
            
            st.divider()
            
            # Отображение текущих параметров
            if "param_options" in st.session_state and st.session_state["param_options"]:
                st.write("**Текущие параметры:**")
                
            # Создаем колонки для отображения параметров (максимум 4 в ряд)
            num_params = len(st.session_state["param_options"])
            
            # Разбиваем параметры на группы по 4
            param_items = list(st.session_state["param_options"].items())
            for row_start in range(0, num_params, 4):
                row_params = param_items[row_start:row_start + 4]
                param_cols = st.columns(len(row_params))
                
                for col_idx, (param_name, options) in enumerate(row_params):
                    with param_cols[col_idx]:
                        with st.expander(f"📋 {param_name} ({len(options)} вариантов)"):
                            st.write("**Варианты:**")
                            for j, option in enumerate(options):
                                col_opt, col_edit, col_del = st.columns([3, 1, 1])
                                with col_opt:
                                    st.write(f"• {option}")
                                with col_edit:
                                    if st.button(f"✏️", key=f"edit_option_{param_name}_{j}"):
                                        st.session_state[f"editing_{param_name}_{j}"] = True
                                with col_del:
                                    if st.button(f"🗑️", key=f"del_option_{param_name}_{j}"):
                                        # Удаляем вариант
                                        options.pop(j)
                                        if not options:
                                            # Если вариантов не осталось, удаляем параметр
                                            del st.session_state["param_options"][param_name]
                                
                                # Редактирование варианта
                                if st.session_state.get(f"editing_{param_name}_{j}", False):
                                    col_edit_input, col_edit_save, col_edit_cancel = st.columns([3, 1, 1])
                                    with col_edit_input:
                                        edited_option = st.text_input(
                                            "Редактировать:",
                                            value=option,
                                            key=f"edit_input_{param_name}_{j}"
                                        )
                                    with col_edit_save:
                                        if st.button("💾", key=f"save_edit_{param_name}_{j}"):
                                            if edited_option and edited_option.strip():
                                                options[j] = edited_option.strip()
                                                del st.session_state[f"editing_{param_name}_{j}"]
                                    with col_edit_cancel:
                                        if st.button("❌", key=f"cancel_edit_{param_name}_{j}"):
                                            del st.session_state[f"editing_{param_name}_{j}"]
                            
                            # Добавление нового варианта
                            st.markdown("---")
                            col_add_opt, col_add_btn = st.columns([3, 1])
                            with col_add_opt:
                                new_option = st.text_input(
                                    f"Добавить вариант в '{param_name}':",
                                    placeholder="Введите новый вариант",
                                    key=f"new_option_{param_name}"
                                )
                            with col_add_btn:
                                if st.button("➕ Добавить", key=f"add_option_{param_name}"):
                                    if new_option and new_option.strip():
                                        if new_option.strip() not in options:
                                            options.append(new_option.strip())
                                            st.success(f"✅ Вариант '{new_option.strip()}' добавлен")
                                        else:
                                            st.warning("Этот вариант уже существует")
                                    else:
                                        st.warning("Введите название варианта")
                            
                            # Кнопка удаления всего параметра
                            if st.button(f"🗑️ Удалить параметр '{param_name}'", key=f"del_param_{param_name}"):
                                del st.session_state["param_options"][param_name]
                                if param_name in param_values:
                                    del param_values[param_name]
            
            st.divider()
            
            # Таблица с параметрами
            if "Артикул" in display_df.columns:
                st.write("**Таблица параметров товаров:**")
                
                # Получаем список всех параметров (исключаем "крой")
                all_params = list(param_values.keys())
                if "param_options" in st.session_state:
                    all_params.extend([p for p in st.session_state["param_options"].keys() if p not in all_params])
                
                # Убираем фильтр исключения параметра "Крой" - теперь все параметры отображаются
                
                if all_params:
                    # Создаем DataFrame для отображения
                    table_data = []
                    
                    for sku in sorted(display_df["Артикул"].dropna().unique()):
                        sku_str = str(sku).replace(".0", "")
                        row_data = {"Артикул": sku_str}
                        
                        # Добавляем изображение
                        url_cache = get_url_cache()
                        url = url_cache.get(sku_str, "")
                        cached_path = get_cached_image_path(sku_str)
                        if cached_path and os.path.exists(cached_path):
                            with open(cached_path, "rb") as f:
                                img_data = base64.b64encode(f.read()).decode()
                                row_data["Изображение"] = f"data:image/jpeg;base64,{img_data}"
                        else:
                            row_data["Изображение"] = ""
                        
                        # Добавляем ссылку на товар в Wildberries
                        wb_url = f"https://www.wildberries.ru/catalog/{sku_str}/detail.aspx"
                        row_data["Ссылка"] = wb_url
                        
                        # Добавляем параметры
                        for param in all_params:
                            current_value = param_values.get(param, {}).get(sku_str, "")
                            row_data[param] = current_value
                        
                        table_data.append(row_data)
                    
                    # Создаем DataFrame
                    params_df = pd.DataFrame(table_data)
                    
                    # Конфигурация столбцов
                    column_config = {
                        "Артикул": st.column_config.TextColumn("Артикул", width=120),
                    "Изображение": st.column_config.ImageColumn("Изображение", width=100),
                    "Ссылка": st.column_config.LinkColumn("Ссылка", display_text="🔗", width=60)
                    }
                    
                    # Добавляем конфигурацию для параметров
                    for param in all_params:
                        if param in st.session_state.get("param_options", {}):
                            # Selectbox для параметров с вариантами
                            options = [""] + st.session_state["param_options"][param]
                            column_config[param] = st.column_config.SelectboxColumn(
                                param, 
                                options=options, 
                                width=150
                            )
                        else:
                            # Обычный текст для свободных параметров
                            column_config[param] = st.column_config.TextColumn(param, width=150)
                    
                    # Отображаем редактируемую таблицу
                    edited_df = st.data_editor(
                        params_df,
                        column_config=column_config,
                    width='stretch',
                        hide_index=True,
                        num_rows="fixed",
                        key="params_table"
                    )
                    
                # Сохраняем изменения только при нажатии кнопки "Сохранить сейчас"
                # Изменения накапливаются в edited_df, но не применяются к param_values автоматически
                    
                    # Автоматическое сохранение отключено
                    # if changes_made:
                    #     try:
                    #         table_cache_data = {
                    #             "param_values": param_values,
                    #             "param_options": st.session_state.get("param_options", {}),
                    #             "timestamp": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
                    #         }
                    #         
                    #         import json
                    #         with open("table_cache.json", "w", encoding="utf-8") as f:
                    #             json.dump(table_cache_data, f, ensure_ascii=False, indent=2)
                    #         
                    #         # Показываем уведомление об автосохранении
                    #         st.success("💾 Изменения автоматически сохранены в кеш")
                    #     except Exception as e:
                    #         st.error(f"❌ Ошибка автосохранения: {e}")
                    
                    # Кнопки управления
                    col_save, col_export, col_clear_cache, col_stats = st.columns(4)
                    
                    with col_save:
                        if st.button("💾 Сохранить сейчас", type="primary"):
                            # Применяем изменения из edited_df к param_values
                            changes_made = False
                            for index, row in edited_df.iterrows():
                                sku = row["Артикул"]
                                for param in all_params:
                                    if param in row and row[param]:
                                        if param not in param_values:
                                            param_values[param] = {}
                                        if sku not in param_values[param] or param_values[param][sku] != str(row[param]):
                                            param_values[param][sku] = str(row[param])
                                            changes_made = True
                                    elif param in param_values and sku in param_values[param]:
                                        # Удаляем пустые значения
                                        if not row.get(param):
                                            del param_values[param][sku]
                                            changes_made = True
                            
                            # Сохраняем в файл
                            if save_param_values_to_file():
                                if changes_made:
                                    st.success("✅ Параметры сохранены!")
                                else:
                                    st.info("ℹ️ Изменений не обнаружено")
                            else:
                                st.error("❌ Ошибка сохранения")
                    
                    with col_export:
                        if st.button("📥 Экспорт в CSV"):
                            if all_params:
                                # Создаем CSV с текущими данными таблицы
                                csv_data = edited_df.drop("Изображение", axis=1).to_csv(index=False, encoding='utf-8-sig')
                            # Получаем имя загруженного файла для названия экспорта
                            base_filename = "products_parameters"
                            if hasattr(uploaded, 'name') and uploaded.name:
                                # Убираем расширение и добавляем суффикс
                                name_without_ext = os.path.splitext(uploaded.name)[0]
                                base_filename = f"{name_without_ext}_parameters"
                            
                                st.download_button(
                                    label="💾 Скачать CSV",
                                    data=csv_data,
                                file_name=f"{base_filename}.csv",
                                    mime="text/csv"
                                )
                            else:
                                st.warning("Нет данных для экспорта")
                    
                    with col_clear_cache:
                        if st.button("🗑️ Очистить кеш", help="Удалить сохраненный кеш параметров"):
                            try:
                                import os
                                if os.path.exists("table_cache.json"):
                                    os.remove("table_cache.json")
                                    st.success("✅ Кеш параметров очищен!")
                                else:
                                    st.warning("Кеш не найден")
                            except Exception as e:
                                st.error(f"❌ Ошибка очистки кеша: {e}")
                    
                    with col_stats:
                        # Статистика заполнения
                        total_products = len(edited_df)
                        filled_count = 0
                        for param in all_params:
                            if param in edited_df.columns:
                                filled_count += len([v for v in edited_df[param] if v])
                        
                        st.metric(
                            "Заполнено параметров", 
                            f"{filled_count}",
                            f"из {total_products * len(all_params) if all_params else 0}"
                        )
                
                else:
                    st.info("Добавьте параметры выше, чтобы начать работу с таблицей")
                    
                # Информация об автосохранении
                st.caption("🔄 Таблица автоматически сохраняется каждую минуту")
            
            else:
                st.warning("Сначала загрузите данные с артикулами в первой вкладке")
        
        with tab3:
            st.subheader("📈 Аналитика по параметрам")
            
            # Получаем данные параметров
            param_values = get_param_values()

            if not param_values:
                st.warning("Сначала установите параметры товаров во второй вкладке")
            else:
                # Выбор параметра для анализа
                available_params = list(param_values.keys())
                if available_params:
                    # Кнопка экспорта всех параметров в Excel
                    col_export_all, col_select = st.columns([1, 2])
                    
                    with col_export_all:
                        if st.button("📊 Экспорт всех параметров в Excel", type="secondary"):
                            try:
                                # Создаем Excel файл с несколькими листами
                                import io
                                from openpyxl import Workbook
                                
                                wb = Workbook()
                                # Удаляем дефолтный лист
                                wb.remove(wb.active)
                                
                                # Получаем имя загруженного файла для названия экспорта
                                base_filename = "analytics_all_parameters"
                                if hasattr(uploaded, 'name') and uploaded.name:
                                    name_without_ext = os.path.splitext(uploaded.name)[0]
                                    base_filename = f"{name_without_ext}_analytics_all_parameters"
                                
                                # Создаем лист для каждого параметра
                                for param_name in available_params:
                                    ws = wb.create_sheet(title=param_name)
                                    
                                    # Получаем данные для этого параметра
                                    param_values_set = set()
                                    if param_name in param_values:
                                        for sku, value in param_values[param_name].items():
                                            if value:
                                                param_values_set.add(value)
                                        
                                        # Создаем аналитику для этого параметра
                                        analytics_data = []
                                        
                                        for param_value in sorted(param_values_set):
                                            matching_skus = []
                                            if param_name in param_values:
                                                for sku, value in param_values[param_name].items():
                                                    if value == param_value:
                                                        matching_skus.append(sku)
                                            
                                            if matching_skus:
                                                mask = df["Артикул"].astype(str).str.replace(".0", "").isin(matching_skus)
                                                filtered_df = df[mask]
                                                
                                                if not filtered_df.empty:
                                                    total_revenue = filtered_df["Выручка"].sum() if "Выручка" in filtered_df.columns else 0
                                                    total_orders = filtered_df["Заказы"].sum() if "Заказы" in filtered_df.columns else 0
                                                    avg_price = filtered_df["Средняя цена"].mean() if "Средняя цена" in filtered_df.columns else 0
                                                    lost_revenue = filtered_df["Упущенная выручка"].sum() if "Упущенная выручка" in filtered_df.columns else 0
                                                    avg_position = filtered_df["Позиция в выдаче (средняя)"].mean() if "Позиция в выдаче (средняя)" in filtered_df.columns else 0
                                                    avg_cpm = filtered_df["Стоимость за 1000 показов на 1 артикул"].mean() if "Стоимость за 1000 показов на 1 артикул" in filtered_df.columns else 0
                                                    
                                                    analytics_data.append({
                                                        param_value: {
                                                            'Общая выручка': total_revenue,
                                                            'Количество артикулов': len(filtered_df),
                                                            'Выручка на 1 артикул': total_revenue / len(filtered_df) if len(filtered_df) > 0 else 0,
                                                            'Средняя цена без СПП': avg_price,
                                                            'Упущенная выручка': lost_revenue,
                                                            'Упущенная выручка на 1 артикул': lost_revenue / len(filtered_df) if len(filtered_df) > 0 else 0,
                                                            'Позиция в выдаче (средняя)': avg_position,
                                                            'Стоимость за 1000 показов на 1 артикул': avg_cpm
                                                        }
                                                    })
                                            
                                            # Записываем данные в лист
                                            if analytics_data:
                                                # Заголовки
                                                ws['A1'] = 'Метрика'
                                                col = 2
                                                param_values_list = sorted([list(item.keys())[0] for item in analytics_data])
                                                
                                                for param_val in param_values_list:
                                                    ws.cell(row=1, column=col, value=param_val)
                                                    col += 1
                                                
                                                # Данные
                                                metric_names = [
                                                    "Общая выручка",
                                                    "Количество артикулов", 
                                                    "Выручка на 1 артикул",
                                                    "Средняя цена без СПП",
                                                    "Упущенная выручка",
                                                    "Упущенная выручка на 1 артикул",
                                                    "Позиция в выдаче (средняя)",
                                                    "Стоимость за 1000 показов на 1 артикул"
                                                ]
                                                
                                                for row, metric in enumerate(metric_names, 2):
                                                    ws.cell(row=row, column=1, value=metric)
                                                    
                                                    col = 2
                                                    for param_val in param_values_list:
                                                        # Находим данные для этого значения параметра
                                                        for item in analytics_data:
                                                            if param_val in item:
                                                                metrics = item[param_val]
                                                                value = metrics.get(metric, 0)
                                                                if metric in ["Общая выручка", "Выручка на 1 артикул", "Средняя цена без СПП", "Упущенная выручка", "Упущенная выручка на 1 артикул"]:
                                                                    ws.cell(row=row, column=col, value=value)
                                                                else:
                                                                    ws.cell(row=row, column=col, value=value)
                                                                break
                                                    col += 1
                                
                                # Сохраняем в байты
                                excel_buffer = io.BytesIO()
                                wb.save(excel_buffer)
                                excel_buffer.seek(0)
                                
                                st.download_button(
                                    label="💾 Скачать Excel файл",
                                    data=excel_buffer.getvalue(),
                                    file_name=f"{base_filename}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                                
                                st.success("✅ Excel файл готов к скачиванию!")
                                
                            except Exception as e:
                                st.error(f"❌ Ошибка создания Excel файла: {e}")
                    
                    with col_select:
                        selected_param = st.selectbox(
                        "Выберите параметр для анализа",
                        available_params,
                        key="analytics_param_selector"
                    )
                    
                    if selected_param:
                        st.write(f"**Аналитика по параметру: {selected_param}**")
                        
                        # Создаем DataFrame для анализа
                        analytics_data = []
                        
                        # Получаем уникальные значения параметра
                        param_values_set = set()
                        if selected_param in param_values:
                            for sku, value in param_values[selected_param].items():
                                if value:
                                    param_values_set.add(value)
                        
                        # Для каждого значения параметра собираем метрики
                        for param_value in sorted(param_values_set):
                            # Находим артикулы с этим значением параметра
                            matching_skus = []
                            if selected_param in param_values:
                                for sku, value in param_values[selected_param].items():
                                    if value == param_value:
                                        matching_skus.append(sku)
                            
                            if matching_skus:
                                # Фильтруем исходные данные по этим артикулам
                                mask = df["Артикул"].astype(str).str.replace(".0", "").isin(matching_skus)
                                filtered_df = df[mask]
                                
                                if not filtered_df.empty:
                                    # Вычисляем метрики
                                    total_revenue = filtered_df["Выручка"].sum() if "Выручка" in filtered_df.columns else 0
                                    total_orders = filtered_df["Заказы"].sum() if "Заказы" in filtered_df.columns else 0
                                    avg_price = filtered_df["Средняя цена"].mean() if "Средняя цена" in filtered_df.columns else 0
                                    lost_revenue = filtered_df["Упущенная выручка"].sum() if "Упущенная выручка" in filtered_df.columns else 0
                                    revenue_per_product = total_revenue / len(filtered_df) if len(filtered_df) > 0 else 0
                                    lost_revenue_per_product = lost_revenue / len(filtered_df) if len(filtered_df) > 0 else 0
                                    avg_position = filtered_df["Позиция в выдаче"].mean() if "Позиция в выдаче" in filtered_df.columns else 0
                                    avg_cpm = filtered_df["Стоимость за 1000 показов"].mean() if "Стоимость за 1000 показов" in filtered_df.columns else 0
                                    
                                    analytics_data.append({
                                        "Метрика": selected_param,
                                        param_value: {
                                            "Общая выручка": total_revenue,
                                            "Количество артикулов": len(filtered_df),
                                            "Выручка на 1 артикул": revenue_per_product,
                                            "Средняя цена без СПП": avg_price,
                                            "Упущенная выручка": lost_revenue,
                                            "Упущенная выручка на 1 артикул": lost_revenue_per_product,
                                            "Позиция в выдаче (средняя)": avg_position,
                                            "Стоимость за 1000 показов на 1 артикул": avg_cpm,
                                        }
                                    })
                        
                        if analytics_data:
                            # Создаем сводную таблицу
                            summary_data = {}
                            for item in analytics_data:
                                for param_val, metrics in item.items():
                                    if param_val != "Метрика":
                                        summary_data[param_val] = metrics
                            
                            # Создаем сводную таблицу в стиле как на картинке
                            if summary_data:
                                # Получаем все значения параметра (цвета) и сортируем их
                                param_values_list = sorted(summary_data.keys())
                                
                                # Создаем DataFrame с метриками по строкам и значениями параметра по столбцам
                                metric_names = [
                                    "Общая выручка",
                                    "Количество артикулов", 
                                    "Выручка на 1 артикул",
                                    "Средняя цена без СПП",
                                    "Упущенная выручка",
                                    "Упущенная выручка на 1 артикул",
                                    "Позиция в выдаче (средняя)",
                                    "Стоимость за 1000 показов на 1 артикул"
                                ]
                                
                                table_data = {"Метрика": metric_names}
                                
                                # Сохраняем числовые данные для цветового кодирования
                                numeric_data = {}
                                
                                # Добавляем данные для каждого значения параметра
                                for param_value in param_values_list:
                                    metrics = summary_data[param_value]
                                    # Сохраняем числовые значения
                                    numeric_data[param_value] = [
                                        metrics['Общая выручка'],
                                        metrics['Количество артикулов'],
                                        metrics['Выручка на 1 артикул'],
                                        metrics['Средняя цена без СПП'],
                                        metrics['Упущенная выручка'],
                                        metrics['Упущенная выручка на 1 артикул'],
                                        metrics['Позиция в выдаче (средняя)'],  # Для позиции меньше = лучше, обработаем отдельно
                                        metrics['Стоимость за 1000 показов на 1 артикул']
                                    ]
                                    
                                    # Форматированные значения для отображения
                                    table_data[param_value] = [
                                        f"₽{metrics['Общая выручка']:,.0f}".replace(",", " "),
                                        f"{metrics['Количество артикулов']:,.0f}".replace(",", " "),
                                        f"₽{metrics['Выручка на 1 артикул']:,.0f}".replace(",", " "),
                                        f"₽{metrics['Средняя цена без СПП']:,.0f}".replace(",", " "),
                                        f"₽{metrics['Упущенная выручка']:,.0f}".replace(",", " "),
                                        f"₽{metrics['Упущенная выручка на 1 артикул']:,.0f}".replace(",", " "),
                                        f"{metrics['Позиция в выдаче (средняя)']:,.0f}".replace(",", " "),
                                        f"{metrics['Стоимость за 1000 показов на 1 артикул']:,.0f}".replace(",", " ")
                                    ]
                                
                                # Создаем DataFrame
                                display_df = pd.DataFrame(table_data)
                                
                                # Функция для выделения лучших результатов
                                def highlight_best_values(values, reverse=False):
                                    """Выделяет зеленым только лучшие результаты"""
                                    if not values or all(pd.isna(v) or v == 0 for v in values):
                                        return ['background-color: white'] * len(values)
                                    
                                    # Очищаем от NaN и нулевых значений для поиска лучшего
                                    clean_values = [v for v in values if not pd.isna(v) and v != 0]
                                    if not clean_values:
                                        return ['background-color: white'] * len(values)
                                    
                                    # Находим лучшее значение
                                    if reverse:  # Для позиции: меньше = лучше
                                        best_val = min(clean_values)
                                    else:  # Для остальных метрик: больше = лучше
                                        best_val = max(clean_values)
                                    
                                    colors = []
                                    for val in values:
                                        if pd.isna(val) or val == 0:
                                            colors.append('background-color: white')
                                        elif val == best_val:
                                            colors.append('background-color: lightgreen')  # Зеленый для лучших
                                        else:
                                            colors.append('background-color: white')  # Белый для остальных
                                    
                                    return colors
                                
                                # Применяем цветовое кодирование
                                def apply_colors(df):
                                    # Создаем стили для каждой строки
                                    styles = pd.DataFrame('', index=df.index, columns=df.columns)
                                    
                                    for i, metric in enumerate(metric_names):
                                        row_values = []
                                        for param_value in param_values_list:
                                            row_values.append(numeric_data[param_value][i])
                                        
                                        # Для позиции используем обратную логику (меньше = лучше)
                                        reverse_logic = (metric == "Позиция в выдаче (средняя)")
                                        colors = highlight_best_values(row_values, reverse=reverse_logic)
                                        
                                        # Применяем цвета к соответствующим ячейкам
                                        for j, param_value in enumerate(param_values_list):
                                            styles.iloc[i, j + 1] = colors[j]  # +1 потому что первый столбец - "Метрика"
                                    
                                    return styles
                                
                                # Добавляем рейтинг
                                if "param_ratings" not in st.session_state:
                                    st.session_state["param_ratings"] = {}
                                
                                param_rating_key = f"{selected_param}_ratings"
                                if param_rating_key not in st.session_state["param_ratings"]:
                                    # Автоматически формируем рейтинг по приоритетам:
                                    # 1. Выручка на 1 артикул (больше = лучше)
                                    # 2. Средняя цена без СПП (больше = лучше) 
                                    # 3. Упущенная выручка на 1 артикул (меньше = лучше)
                                    
                                    def calculate_score(item):
                                        param_val, metrics = item
                                        # Нормализуем значения от 0 до 1
                                        revenue_per_sku = metrics["Выручка на 1 артикул"]
                                        avg_price = metrics["Средняя цена без СПП"]
                                        lost_revenue_per_sku = metrics["Упущенная выручка на 1 артикул"]
                                        
                                        # Находим мин/макс для нормализации
                                        all_revenues = [m["Выручка на 1 артикул"] for m in summary_data.values()]
                                        all_prices = [m["Средняя цена без СПП"] for m in summary_data.values()]
                                        all_lost = [m["Упущенная выручка на 1 артикул"] for m in summary_data.values()]
                                        
                                        # Нормализуем выручку (0-1, где 1 = максимум)
                                        if max(all_revenues) > min(all_revenues):
                                            norm_revenue = (revenue_per_sku - min(all_revenues)) / (max(all_revenues) - min(all_revenues))
                                        else:
                                            norm_revenue = 0.5
                                        
                                        # Нормализуем цену (0-1, где 1 = максимум)
                                        if max(all_prices) > min(all_prices):
                                            norm_price = (avg_price - min(all_prices)) / (max(all_prices) - min(all_prices))
                                        else:
                                            norm_price = 0.5
                                        
                                        # Нормализуем упущенную выручку (0-1, где 1 = минимум, т.е. лучше)
                                        if max(all_lost) > min(all_lost):
                                            norm_lost = 1 - (lost_revenue_per_sku - min(all_lost)) / (max(all_lost) - min(all_lost))
                                        else:
                                            norm_lost = 0.5
                                        
                                        # Взвешенная сумма с приоритетами
                                        score = (norm_revenue * 0.9) + (norm_price * 0.09) + (norm_lost * 0.01)
                                        return score
                                    
                                    # Сортируем по рассчитанному рейтингу
                                    sorted_by_score = sorted(
                                        summary_data.items(), 
                                        key=calculate_score,
                                        reverse=True
                                    )
                                    ratings = {param_val: i+1 for i, (param_val, _) in enumerate(sorted_by_score)}
                                    st.session_state["param_ratings"][param_rating_key] = ratings
                                
                                ratings = st.session_state["param_ratings"][param_rating_key]
                                
                                # Отображаем таблицу
                                st.write(f"**Сводная таблица по параметру: {selected_param}**")
                                st.info("💡 Цветовое выделение: 🟢 лучший результат в каждой строке. Для позиции в выдаче лучший = меньшее число. Рейтинг формируется по приоритетам: Выручка на 1 артикул (90%) → Средняя цена (9%) → Упущенная выручка на 1 артикул (1%)")
                                
                                # Создаем конфигурацию столбцов
                                column_config = {
                                    "Метрика": st.column_config.TextColumn("Метрика", width=250)
                                }
                                
                                # Настраиваем столбцы для значений параметров
                                for param_value in param_values_list:
                                    column_config[param_value] = st.column_config.TextColumn(
                                        param_value, 
                                        width=150
                                    )
                                
                                # Применяем стили и отображаем таблицу
                                styled_df = display_df.style.apply(lambda x: apply_colors(display_df), axis=None)
                                
                                # Отображаем стилизованную таблицу
                                st.dataframe(
                                    styled_df,
                                    column_config=column_config,
                                    width='stretch',
                                    hide_index=True
                                )
                                
                                # Добавляем строку с рейтингом отдельно (без цветового кодирования)
                                st.write("**Рейтинг:**")
                                rating_display_data = {"Метрика": ["Рейтинг"]}
                                for param_value in param_values_list:
                                    rating_display_data[param_value] = [str(ratings.get(param_value, len(param_values_list)+1))]
                                
                                rating_display_df = pd.DataFrame(rating_display_data)
                                st.dataframe(
                                    rating_display_df,
                                    column_config=column_config,
                                    width='stretch',
                                    hide_index=True
                                )
                                
                                st.divider()
                                
                                # Редактируемая таблица только для рейтинга
                                st.write("**Редактирование рейтинга:**")
                                
                                # Создаем DataFrame только для рейтинга
                                rating_edit_data = {}
                                for param_value in param_values_list:
                                    rating_edit_data[param_value] = [ratings.get(param_value, len(param_values_list)+1)]
                                
                                rating_edit_df = pd.DataFrame(rating_edit_data, index=["Рейтинг"])
                                
                                # Конфигурация для редактирования рейтинга
                                rating_column_config = {}
                                for param_value in param_values_list:
                                    rating_column_config[param_value] = st.column_config.NumberColumn(
                                        param_value,
                                        min_value=1,
                                        max_value=len(param_values_list),
                                        step=1,
                                        width=150
                                    )
                                
                                # Редактируемая таблица рейтинга
                                edited_rating_df = st.data_editor(
                                    rating_edit_df,
                                    column_config=rating_column_config,
                                    width='stretch',
                                    hide_index=False,
                                    key=f"rating_table_{selected_param}"
                                )
                                
                                # Сохраняем изменения рейтинга
                                if not edited_rating_df.equals(rating_edit_df):
                                    new_ratings = {}
                                    for param_value in param_values_list:
                                        new_ratings[param_value] = int(edited_rating_df.loc["Рейтинг", param_value])
                                    st.session_state["param_ratings"][param_rating_key] = new_ratings
                                    
                                    # Автосохранение рейтингов отключено
                                    # try:
                                    #     table_cache_data = {
                                    #         "param_values": st.session_state.get("param_values", {}),
                                    #         "param_options": st.session_state.get("param_options", {}),
                                    #         "param_ratings": st.session_state.get("param_ratings", {}),
                                    #         "timestamp": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
                                    #     }
                                    #     
                                    #     import json
                                    #     with open("table_cache.json", "w", encoding="utf-8") as f:
                                    #         json.dump(table_cache_data, f, ensure_ascii=False, indent=2)
                                    #     
                                    #     st.success("💾 Рейтинг автоматически сохранен")
                                    # except Exception as e:
                                    #     st.error(f"❌ Ошибка сохранения рейтинга: {e}")
                                
                                # Кнопки управления
                                col_reset, col_export_analytics = st.columns(2)
                                
                                with col_reset:
                                    if st.button("🔄 Сбросить рейтинг", type="secondary"):
                                        # Пересчитываем рейтинг по приоритетам
                                        def calculate_score(item):
                                            param_val, metrics = item
                                            # Нормализуем значения от 0 до 1
                                            revenue_per_sku = metrics["Выручка на 1 артикул"]
                                            avg_price = metrics["Средняя цена без СПП"]
                                            lost_revenue_per_sku = metrics["Упущенная выручка на 1 артикул"]
                                            
                                            # Находим мин/макс для нормализации
                                            all_revenues = [m["Выручка на 1 артикул"] for m in summary_data.values()]
                                            all_prices = [m["Средняя цена без СПП"] for m in summary_data.values()]
                                            all_lost = [m["Упущенная выручка на 1 артикул"] for m in summary_data.values()]
                                            
                                            # Нормализуем выручку (0-1, где 1 = максимум)
                                            if max(all_revenues) > min(all_revenues):
                                                norm_revenue = (revenue_per_sku - min(all_revenues)) / (max(all_revenues) - min(all_revenues))
                                            else:
                                                norm_revenue = 0.5
                                            
                                            # Нормализуем цену (0-1, где 1 = максимум)
                                            if max(all_prices) > min(all_prices):
                                                norm_price = (avg_price - min(all_prices)) / (max(all_prices) - min(all_prices))
                                            else:
                                                norm_price = 0.5
                                            
                                            # Нормализуем упущенную выручку (0-1, где 1 = минимум, т.е. лучше)
                                            if max(all_lost) > min(all_lost):
                                                norm_lost = 1 - (lost_revenue_per_sku - min(all_lost)) / (max(all_lost) - min(all_lost))
                                            else:
                                                norm_lost = 0.5
                                            
                                            # Взвешенная сумма с приоритетами
                                            score = (norm_revenue * 0.9) + (norm_price * 0.09) + (norm_lost * 0.01)
                                            return score
                                        
                                        # Сортируем по рассчитанному рейтингу
                                        sorted_by_score = sorted(
                                            summary_data.items(), 
                                            key=calculate_score,
                                            reverse=True
                                        )
                                        ratings = {param_val: i+1 for i, (param_val, _) in enumerate(sorted_by_score)}
                                        st.session_state["param_ratings"][param_rating_key] = ratings
                                        st.rerun()
                                
                                with col_export_analytics:
                                    if st.button("📊 Экспорт аналитики"):
                                        # Создаем полную таблицу для экспорта (с рейтингом)
                                        export_data = display_df.copy()
                                        rating_row = ["Рейтинг"] + [str(ratings.get(param_val, len(param_values_list)+1)) for param_val in param_values_list]
                                        rating_export_df = pd.DataFrame([rating_row], columns=export_data.columns)
                                        full_export_df = pd.concat([export_data, rating_export_df], ignore_index=True)
                                        
                                        csv_data = full_export_df.to_csv(encoding='utf-8-sig', index=False)
                                        # Получаем имя загруженного файла для названия экспорта
                                        base_filename = f"analytics_{selected_param}"
                                        if hasattr(uploaded, 'name') and uploaded.name:
                                            # Убираем расширение и добавляем суффикс
                                            name_without_ext = os.path.splitext(uploaded.name)[0]
                                            base_filename = f"{name_without_ext}_analytics_{selected_param}"
                                        
                                        st.download_button(
                                            label="💾 Скачать CSV",
                                            data=csv_data,
                                            file_name=f"{base_filename}.csv",
                                            mime="text/csv"
                                        )
                                
                                # Дополнительная статистика
                                st.divider()
                                st.write("**Дополнительная статистика:**")
                                
                                col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
                                
                                total_products = sum(item["Количество артикулов"] for item in summary_data.values())
                                total_revenue = sum(item["Общая выручка"] for item in summary_data.values())
                                total_lost_revenue = sum(item["Упущенная выручка"] for item in summary_data.values())
                                avg_position_all = sum(item["Позиция в выдаче (средняя)"] * item["Количество артикулов"] for item in summary_data.values()) / total_products if total_products > 0 else 0
                                
                                col_stat1.metric("Всего товаров", f"{total_products} шт.")
                                col_stat2.metric("Общая выручка", f"₽{total_revenue:,.0f}".replace(",", " "))
                                col_stat3.metric("Упущенная выручка", f"₽{total_lost_revenue:,.0f}".replace(",", " "))
                                col_stat4.metric("Средняя позиция", f"{avg_position_all:.1f}")
                                
                        else:
                            st.info("Нет данных для анализа по выбранному параметру")
                
                # Добавляем анализ лучших комбинаций параметров
                st.divider()
                st.subheader("🏆 Идеальная комбинация параметров")
                
                # Получаем рейтинги параметров из session_state
                param_ratings = {}
                for param in available_params:
                    rating_key = f"{param}_ratings"
                    if rating_key in st.session_state.get("param_ratings", {}):
                        param_ratings[param] = st.session_state["param_ratings"][rating_key]
                
                # Создаем идеальную комбинацию
                ideal_combination = {}
                for param_name, ratings in param_ratings.items():
                    # Ищем значение с рейтингом 1
                    for value, rating in ratings.items():
                        if rating == 1:
                            ideal_combination[param_name] = value
                            break
                
                if ideal_combination:
                    st.success(f"🏆 **Идеальная комбинация**: {' + '.join(ideal_combination.values())}")
                    
                    st.write("**Состав идеальной комбинации:**")
                    for param_name, value in ideal_combination.items():
                        st.write(f"• **{param_name}**: {value} (рейтинг 1)")
                    
                    st.info("💡 Создайте товары с такими параметрами для максимальной эффективности")
                    
                    # Создаем промпт для ChatGPT
                    prompt = f"""Проанализируй эту идеальную комбинацию параметров для товара и дай рекомендации по созданию продукта:

ИДЕАЛЬНАЯ КОМБИНАЦИЯ ПАРАМЕТРОВ:
{chr(10).join([f"• {param_name}: {value} (рейтинг 1)" for param_name, value in ideal_combination.items()])}

Пожалуйста, дай рекомендации по:
1. Как создать товар с такими параметрами
2. Какие дополнительные характеристики добавить
3. На что обратить внимание при производстве
4. Как позиционировать товар на рынке
5. Потенциальные проблемы и решения

Будь конкретным и практичным в рекомендациях."""
                    
                    # Показываем промпт и кнопку
                    st.subheader("🤖 Отправка в ChatGPT")
                    
                    col1, col2 = st.columns([2, 1])
                    
                    with col1:
                        st.text_area(
                            "Готовый промпт для ChatGPT:",
                            value=prompt,
                            height=200,
                            help="Скопируйте этот текст и отправьте в ChatGPT для получения рекомендаций"
                        )
                    
                    with col2:
                        st.write("**Действия:**")
                        
                        # Кнопка для копирования промпта
                        if st.button("📋 Скопировать промпт", type="primary"):
                            st.success("✅ Промпт скопирован в буфер обмена!")
                            # В реальном приложении здесь был бы JavaScript для копирования
                        
                        st.write("**Инструкция:**")
                        st.write("1. Нажмите кнопку 'Скопировать промпт'")
                        st.write("2. Откройте ChatGPT")
                        st.write("3. Вставьте промпт")
                        st.write("4. Получите рекомендации!")
                        
                        # Прямая ссылка на ChatGPT (откроется в новой вкладке)
                        st.link_button(
                            "🚀 Открыть ChatGPT",
                            "https://chat.openai.com/",
                            help="Откроет ChatGPT в новой вкладке"
                        )
                else:
                    st.warning("⚠️ Рейтинги параметров не найдены. Создайте рейтинги в аналитике отдельных параметров выше.")
        
        # Четвертая вкладка - Анализ сезонности
        if seasonality_available:
            with tab4:
                st.subheader("📅 Анализ сезонности")
                
                # Выбор источника данных
                data_source = st.radio(
                    "Выберите источник данных:",
                    ["📁 Файл sezon.csv", "📤 Загрузить свой файл", "✏️ Ручной ввод"],
                    help="Выберите способ загрузки данных для анализа сезонности"
                )
                
                seasonality_df = None
                
                if data_source == "📁 Файл sezon.csv":
                    # Проверяем наличие файла sezon.csv
                    if not os.path.exists('sezon.csv'):
                        st.error("❌ Файл sezon.csv не найден в текущей директории")
                        st.info("Пожалуйста, убедитесь, что файл sezon.csv находится в той же папке, что и приложение")
                    else:
                        # Загружаем данные сезонности
                        try:
                            seasonality_df = load_seasonality_data()
                            seasonality_df = clean_seasonality_data(seasonality_df)
                            st.success("✅ Данные из sezon.csv успешно загружены")
                        except Exception as e:
                            st.error(f"❌ Ошибка при загрузке данных: {e}")
                
                elif data_source == "📤 Загрузить свой файл":
                    st.info("📋 Поддерживаемые форматы: CSV, Excel")
                    st.info("📋 Обязательные столбцы: запрос, категория, январь, февраль, март, апрель, май, июнь, июль, август, сентябрь, октябрь, ноябрь, декабрь")
                    
                    uploaded_file = st.file_uploader(
                        "Выберите файл с данными сезонности:",
                        type=['csv', 'xlsx'],
                        help="Загрузите CSV или Excel файл с данными сезонности"
                    )
                    
                    if uploaded_file is not None:
                        custom_df, message = load_custom_data(uploaded_file)
                        if custom_df is not None:
                            seasonality_df = clean_seasonality_data(custom_df)
                            st.success(f"✅ {message}")
                        else:
                            st.error(f"❌ {message}")
                
                elif data_source == "✏️ Ручной ввод":
                    st.info("📝 Введите данные для одного запроса")
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        manual_query = st.text_input("Поисковый запрос:", placeholder="например: джинсы женские")
                        manual_category = st.text_input("Категория:", placeholder="например: джинсы")
                    
                    with col2:
                        st.write("**Частотность по месяцам:**")
                    
                    # Создаем поля для ввода частотности по месяцам
                    month_names = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
                                   'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
                    
                    col_months1, col_months2, col_months3, col_months4 = st.columns(4)
                    manual_frequencies = []
                    
                    for i, month in enumerate(month_names):
                        if i < 3:
                            with col_months1:
                                freq = st.number_input(f"{month}:", min_value=0, value=0, step=1, key=f"manual_{i}")
                                manual_frequencies.append(freq)
                        elif i < 6:
                            with col_months2:
                                freq = st.number_input(f"{month}:", min_value=0, value=0, step=1, key=f"manual_{i}")
                                manual_frequencies.append(freq)
                        elif i < 9:
                            with col_months3:
                                freq = st.number_input(f"{month}:", min_value=0, value=0, step=1, key=f"manual_{i}")
                                manual_frequencies.append(freq)
                        else:
                            with col_months4:
                                freq = st.number_input(f"{month}:", min_value=0, value=0, step=1, key=f"manual_{i}")
                                manual_frequencies.append(freq)
                    
                    if st.button("📊 Анализировать данные", type="primary"):
                        if manual_query and manual_category:
                            seasonality_df = create_manual_entry_data(manual_query, manual_category, manual_frequencies)
                            seasonality_df = clean_seasonality_data(seasonality_df)
                            st.success("✅ Данные успешно созданы и готовы для анализа")
                        else:
                            st.error("❌ Пожалуйста, заполните запрос и категорию")
                
                # Продолжаем только если данные загружены
                if seasonality_df is not None and not seasonality_df.empty:
                        
                        # Создаем вкладки для анализа сезонности
                        seasonality_tab1, seasonality_tab2 = st.tabs(["🔍 Анализ запроса", "📅 Анализ по месяцам"])
                        
                        with seasonality_tab1:
                            st.markdown("---")
                            st.subheader("🔍 Выбор товара для анализа")
                            
                            col1, col2 = st.columns([2, 3])
                            with col1:
                                if 'категория' in seasonality_df.columns:
                                    categories = sorted(seasonality_df['категория'].dropna().unique())
                                    selected_category = st.selectbox(
                                        "Выберите категорию:",
                                        categories,
                                        help="Сначала выберите категорию товаров"
                                    )
                                else:
                                    st.error("Столбец 'категория' не найден")
                                    selected_category = None
                            
                            with col2:
                                if selected_category:
                                    category_df = seasonality_df[seasonality_df['категория'] == selected_category]
                                    if 'запрос' in category_df.columns:
                                        queries_in_category = sorted(category_df['запрос'].dropna().unique())
                                        if queries_in_category:
                                            selected_item = st.selectbox(
                                                "Выберите запрос:",
                                                queries_in_category,
                                                help="Выберите поисковый запрос для анализа сезонности"
                                            )
                                            if selected_item:
                                                filtered_df = category_df[category_df['запрос'] == selected_item]
                                            else:
                                                filtered_df = pd.DataFrame()
                                        else:
                                            st.warning("В этой категории нет поисковых запросов")
                                            filtered_df = pd.DataFrame()
                                    else:
                                        st.error("Столбец 'запрос' не найден")
                                        filtered_df = pd.DataFrame()
                                else:
                                    st.info("Выберите категорию для продолжения")
                                    filtered_df = pd.DataFrame()
                            
                            st.markdown("---")
                            
                            if not filtered_df.empty:
                                # Основная информация
                                row = filtered_df.iloc[0]
                                col_info1, col_info2, col_info3 = st.columns(3)
                                
                                with col_info1:
                                    st.metric("Категория", row.get('категория', 'Н/Д'))
                                with col_info2:
                                    st.metric("Товар", row.get('наименование товара', 'Н/Д'))
                                with col_info3:
                                    st.metric("Запрос", row.get('запрос', 'Н/Д'))
                                
                                # График сезонности
                                fig = create_seasonality_graph(filtered_df, selected_item)
                                if fig:
                                    st.plotly_chart(fig, width='stretch')
                                
                                # Детальная таблица по месяцам
                                st.subheader("📊 Детальные данные по месяцам")
                                month_columns = ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь',
                                                'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь']
                                
                                month_data = []
                                for month in month_columns:
                                    if month in row.index:
                                        month_data.append({
                                            'Месяц': month.capitalize(),
                                            'Частота': f"{row[month]:,.0f}",
                                            'Процент от максимума': f"{(row[month] / max([row[m] for m in month_columns if m in row.index])) * 100:.1f}%"
                                        })
                                
                                if month_data:
                                    month_df = pd.DataFrame(month_data)
                                    st.dataframe(month_df, width='stretch')
                        
                        with seasonality_tab2:
                            st.subheader("📅 Анализ по месяцам")
                            
                            # Выбор месяца
                            month_columns = ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь',
                                            'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь']
                            
                            selected_month = st.selectbox(
                                "Выберите месяц для анализа:",
                                [month.capitalize() for month in month_columns],
                                help="Выберите месяц для анализа всех запросов"
                            )
                            
                            # Получаем данные для выбранного месяца
                            month_lower = selected_month.lower()
                            if month_lower in seasonality_df.columns:
                                columns_to_select = ['запрос', 'категория', 'наименование товара'] + month_columns
                                month_data = seasonality_df[columns_to_select].copy()
                                month_data = month_data[month_data[month_lower] > 0]  # Только с данными
                                
                                if not month_data.empty:
                                    st.write(f"**Найдено {len(month_data)} запросов с данными в {selected_month}**")
                                    
                                    # Получаем статистику по статусам
                                    stats, month_data_with_status = get_status_stats(month_data, month_lower)
                                    
                                    # Показываем KPI метрики
                                    col_kpi1, col_kpi2, col_kpi3, col_kpi4, col_kpi5 = st.columns(5)
                                    col_kpi1.metric("Всего", stats.get('Всего', 0))
                                    col_kpi2.metric("Пик max", stats.get('Пик max', 0))
                                    col_kpi3.metric("Пик min", stats.get('Пик min', 0))
                                    col_kpi4.metric("Рост", stats.get('Рост', 0))
                                    col_kpi5.metric("Падение", stats.get('Падение', 0) + stats.get('Большое падение', 0))
                                    
                                    # Фильтрация и сортировка
                                    col1, col2 = st.columns(2)
                                    
                                    with col1:
                                        status_options = ['Все', 'Пик max', 'Пик min', 'Рост', 'Падение', 'Большое падение']
                                        selected_status = st.selectbox(
                                            "Фильтр по статусу:",
                                            status_options,
                                            help="Выберите статус для фильтрации запросов"
                                        )
                                    
                                    with col2:
                                        sort_options = ['По цвету (зеленый → красный)', 'По частотности (высокая → низкая)', 'По алфавиту']
                                        selected_sort = st.selectbox(
                                            "Сортировка:",
                                            sort_options,
                                            help="Выберите способ сортировки запросов"
                                        )
                                    
                                    # Применяем фильтр по статусу
                                    if selected_status != 'Все':
                                        month_data_with_status = month_data_with_status[month_data_with_status['Статус'] == selected_status]
                                    
                                    # Применяем сортировку
                                    if selected_sort == 'По цвету (зеленый → красный)':
                                        def get_color_priority(row):
                                            current_month_value = row[month_lower]
                                            month_values = [row[month] for month in month_columns]
                                            max_val = max(month_values) if month_values else 1
                                            if max_val == 0:
                                                return 5
                                            
                                            intensity = current_month_value / max_val
                                            
                                            if intensity >= 0.9:
                                                return 1
                                            elif intensity >= 0.5:
                                                return 2
                                            elif intensity >= 0.3:
                                                return 3
                                            else:
                                                return 4
                                        
                                        month_data_with_status['sort_key'] = month_data_with_status.apply(get_color_priority, axis=1)
                                        month_data_with_status = month_data_with_status.sort_values('sort_key')
                                        month_data_with_status = month_data_with_status.drop('sort_key', axis=1)
                                        
                                    elif selected_sort == 'По частотности (высокая → низкая)':
                                        month_data_with_status = month_data_with_status.sort_values(month_lower, ascending=False)
                                        
                                    else:  # По алфавиту
                                        month_data_with_status = month_data_with_status.sort_values('запрос')
                                    
                                    # Показываем таблицу
                                    st.subheader("📋 Список запросов")
                                    
                                    # Стилизуем таблицу
                                    styled_df = style_dataframe(month_data_with_status, month_lower)
                                    st.dataframe(styled_df, width='stretch')
                                    
                                    # Легенда цветов
                                    st.markdown("---")
                                    st.caption("**Легенда цветов:**")
                                    col_legend1, col_legend2, col_legend3, col_legend4 = st.columns(4)
                                    col_legend1.markdown("🟢 **Зеленый** - 90%+ от максимума")
                                    col_legend2.markdown("🟡 **Желтый** - 50-90% от максимума")
                                    col_legend3.markdown("🟠 **Бледно-желтый** - 30-50% от максимума")
                                    col_legend4.markdown("🔴 **Красный** - <30% от максимума")
                                    
                                else:
                                    st.warning(f"Нет данных для месяца {selected_month}")
                            else:
                                st.error(f"Столбец '{selected_month}' не найден в данных")
                else:
                    st.info("📊 Выберите источник данных и загрузите информацию для анализа сезонности")
        else:
            st.warning("Модуль анализа сезонности недоступен")
        
        # Пятая вкладка - Прогнозирование с Prophet
        if PROPHET_AVAILABLE:
            if seasonality_available:
                with tab5:
                    st.subheader("🔮 Прогнозирование с Prophet")
                    
                    # Проверяем наличие данных
                    if df is not None and not df.empty:
                        # Настройки прогнозирования
                        col_settings1, col_settings2 = st.columns(2)
                        
                        with col_settings1:
                            # Выбор метрики для прогнозирования
                            numeric_columns = df.select_dtypes(include=[np.number]).columns.tolist()
                            if numeric_columns:
                                metric_choice = st.selectbox(
                                    "Выберите метрику для прогнозирования:",
                                    numeric_columns,
                                    key="prophet_metric_choice"
                                )
                            else:
                                st.warning("Нет числовых колонок для прогнозирования")
                                metric_choice = None
                        
                        with col_settings2:
                            # Выбор колонки с датами
                            date_columns = []
                            for col in df.columns:
                                if df[col].dtype == 'datetime64[ns]' or 'дата' in col.lower() or 'date' in col.lower():
                                    date_columns.append(col)
                            
                            if date_columns:
                                date_choice = st.selectbox(
                                    "Выберите колонку с датами (опционально):",
                                    ["Автоматически"] + date_columns,
                                    key="prophet_date_choice"
                                )
                                if date_choice == "Автоматически":
                                    date_choice = None
                            else:
                                date_choice = None
                                st.info("Колонка с датами не найдена, будет создана автоматически")
                        
                        # Дополнительные настройки
                        col_periods, col_seasonality = st.columns(2)
                        
                        with col_periods:
                            forecast_periods = st.number_input(
                                "Период прогнозирования (дни):",
                                min_value=1,
                                max_value=365,
                                value=30,
                                key="prophet_periods"
                            )
                        
                        with col_seasonality:
                            seasonality_mode = st.selectbox(
                                "Режим сезонности:",
                                ["additive", "multiplicative"],
                                key="prophet_seasonality"
                            )
                        
                        # Кнопка создания прогноза
                        if st.button("🔮 Создать прогноз", type="primary", key="create_forecast_btn"):
                            if metric_choice:
                                with st.spinner("Создание прогноза..."):
                                    # Подготавливаем данные
                                    df_prophet = prepare_data_for_prophet(df, metric_choice, date_choice)
                                    
                                    if df_prophet is not None and len(df_prophet) > 1:
                                        # Создаем прогноз
                                        model, forecast, future = create_prophet_forecast(
                                            df_prophet, 
                                            periods=forecast_periods,
                                            seasonality_mode=seasonality_mode
                                        )
                                        
                                        if model and forecast is not None:
                                            # Отображаем основной график прогноза
                                            st.subheader("📈 Прогноз")
                                            fig_forecast = plot_prophet_forecast(
                                                model, 
                                                forecast, 
                                                f"Прогноз {metric_choice}"
                                            )
                                            if fig_forecast:
                                                st.plotly_chart(fig_forecast, use_container_width=True)
                                            
                                            # Отображаем компоненты прогноза
                                            st.subheader("🔍 Компоненты прогноза")
                                            fig_components = plot_prophet_components(
                                                model, 
                                                forecast, 
                                                f"Компоненты прогноза {metric_choice}"
                                            )
                                            if fig_components:
                                                st.plotly_chart(fig_components, use_container_width=True)
                                            
                                            # Статистика прогноза
                                            st.subheader("📊 Статистика прогноза")
                                            
                                            # Получаем последние прогнозные значения
                                            forecast_future = forecast[forecast['ds'] > df_prophet['ds'].max()]
                                            
                                            if not forecast_future.empty:
                                                col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
                                                
                                                with col_stat1:
                                                    mean_forecast = forecast_future['yhat'].mean()
                                                    st.metric("Средний прогноз", f"{mean_forecast:,.0f}")
                                                
                                                with col_stat2:
                                                    max_forecast = forecast_future['yhat'].max()
                                                    st.metric("Максимальный прогноз", f"{max_forecast:,.0f}")
                                                
                                                with col_stat3:
                                                    min_forecast = forecast_future['yhat'].min()
                                                    st.metric("Минимальный прогноз", f"{min_forecast:,.0f}")
                                                
                                                with col_stat4:
                                                    trend = forecast_future['trend'].iloc[-1] - forecast_future['trend'].iloc[0]
                                                    st.metric("Изменение тренда", f"{trend:,.0f}")
                                                
                                                # Таблица с прогнозными значениями
                                                st.subheader("📋 Детальный прогноз")
                                                forecast_display = forecast_future[['ds', 'yhat', 'yhat_lower', 'yhat_upper']].copy()
                                                forecast_display.columns = ['Дата', 'Прогноз', 'Нижняя граница', 'Верхняя граница']
                                                forecast_display['Дата'] = forecast_display['Дата'].dt.strftime('%Y-%m-%d')
                                                
                                                st.dataframe(
                                                    forecast_display,
                                                    use_container_width=True,
                                                    hide_index=True
                                                )
                                                
                                                # Экспорт прогноза
                                                csv_data = forecast_display.to_csv(index=False)
                                                st.download_button(
                                                    label="💾 Скачать прогноз (CSV)",
                                                    data=csv_data,
                                                    file_name=f"prophet_forecast_{metric_choice}.csv",
                                                    mime="text/csv"
                                                )
                                            
                                        else:
                                            st.error("Не удалось создать прогноз. Проверьте данные.")
                                    else:
                                        st.error("Недостаточно данных для создания прогноза. Нужно минимум 2 точки данных.")
                            else:
                                st.warning("Выберите метрику для прогнозирования")
                    
                    else:
                        st.info("📊 Загрузите данные в первой вкладке для создания прогнозов")
            
            else:
                # Если нет вкладки сезонности, используем tab4
                with tab4:
                    st.subheader("🔮 Прогнозирование с Prophet")
                    
                    # Проверяем наличие данных
                    if df is not None and not df.empty:
                        # Настройки прогнозирования
                        col_settings1, col_settings2 = st.columns(2)
                        
                        with col_settings1:
                            # Выбор метрики для прогнозирования
                            numeric_columns = df.select_dtypes(include=[np.number]).columns.tolist()
                            if numeric_columns:
                                metric_choice = st.selectbox(
                                    "Выберите метрику для прогнозирования:",
                                    numeric_columns,
                                    key="prophet_metric_choice"
                                )
                            else:
                                st.warning("Нет числовых колонок для прогнозирования")
                                metric_choice = None
                        
                        with col_settings2:
                            # Выбор колонки с датами
                            date_columns = []
                            for col in df.columns:
                                if df[col].dtype == 'datetime64[ns]' or 'дата' in col.lower() or 'date' in col.lower():
                                    date_columns.append(col)
                            
                            if date_columns:
                                date_choice = st.selectbox(
                                    "Выберите колонку с датами (опционально):",
                                    ["Автоматически"] + date_columns,
                                    key="prophet_date_choice"
                                )
                                if date_choice == "Автоматически":
                                    date_choice = None
                            else:
                                date_choice = None
                                st.info("Колонка с датами не найдена, будет создана автоматически")
                        
                        # Дополнительные настройки
                        col_periods, col_seasonality = st.columns(2)
                        
                        with col_periods:
                            forecast_periods = st.number_input(
                                "Период прогнозирования (дни):",
                                min_value=1,
                                max_value=365,
                                value=30,
                                key="prophet_periods"
                            )
                        
                        with col_seasonality:
                            seasonality_mode = st.selectbox(
                                "Режим сезонности:",
                                ["additive", "multiplicative"],
                                key="prophet_seasonality"
                            )
                        
                        # Кнопка создания прогноза
                        if st.button("🔮 Создать прогноз", type="primary", key="create_forecast_btn"):
                            if metric_choice:
                                with st.spinner("Создание прогноза..."):
                                    # Подготавливаем данные
                                    df_prophet = prepare_data_for_prophet(df, metric_choice, date_choice)
                                    
                                    if df_prophet is not None and len(df_prophet) > 1:
                                        # Создаем прогноз
                                        model, forecast, future = create_prophet_forecast(
                                            df_prophet, 
                                            periods=forecast_periods,
                                            seasonality_mode=seasonality_mode
                                        )
                                        
                                        if model and forecast is not None:
                                            # Отображаем основной график прогноза
                                            st.subheader("📈 Прогноз")
                                            fig_forecast = plot_prophet_forecast(
                                                model, 
                                                forecast, 
                                                f"Прогноз {metric_choice}"
                                            )
                                            if fig_forecast:
                                                st.plotly_chart(fig_forecast, use_container_width=True)
                                            
                                            # Отображаем компоненты прогноза
                                            st.subheader("🔍 Компоненты прогноза")
                                            fig_components = plot_prophet_components(
                                                model, 
                                                forecast, 
                                                f"Компоненты прогноза {metric_choice}"
                                            )
                                            if fig_components:
                                                st.plotly_chart(fig_components, use_container_width=True)
                                            
                                            # Статистика прогноза
                                            st.subheader("📊 Статистика прогноза")
                                            
                                            # Получаем последние прогнозные значения
                                            forecast_future = forecast[forecast['ds'] > df_prophet['ds'].max()]
                                            
                                            if not forecast_future.empty:
                                                col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
                                                
                                                with col_stat1:
                                                    mean_forecast = forecast_future['yhat'].mean()
                                                    st.metric("Средний прогноз", f"{mean_forecast:,.0f}")
                                                
                                                with col_stat2:
                                                    max_forecast = forecast_future['yhat'].max()
                                                    st.metric("Максимальный прогноз", f"{max_forecast:,.0f}")
                                                
                                                with col_stat3:
                                                    min_forecast = forecast_future['yhat'].min()
                                                    st.metric("Минимальный прогноз", f"{min_forecast:,.0f}")
                                                
                                                with col_stat4:
                                                    trend = forecast_future['trend'].iloc[-1] - forecast_future['trend'].iloc[0]
                                                    st.metric("Изменение тренда", f"{trend:,.0f}")
                                                
                                                # Таблица с прогнозными значениями
                                                st.subheader("📋 Детальный прогноз")
                                                forecast_display = forecast_future[['ds', 'yhat', 'yhat_lower', 'yhat_upper']].copy()
                                                forecast_display.columns = ['Дата', 'Прогноз', 'Нижняя граница', 'Верхняя граница']
                                                forecast_display['Дата'] = forecast_display['Дата'].dt.strftime('%Y-%m-%d')
                                                
                                                st.dataframe(
                                                    forecast_display,
                                                    use_container_width=True,
                                                    hide_index=True
                                                )
                                                
                                                # Экспорт прогноза
                                                csv_data = forecast_display.to_csv(index=False)
                                                st.download_button(
                                                    label="💾 Скачать прогноз (CSV)",
                                                    data=csv_data,
                                                    file_name=f"prophet_forecast_{metric_choice}.csv",
                                                    mime="text/csv"
                                                )
                                            
                                        else:
                                            st.error("Не удалось создать прогноз. Проверьте данные.")
                                    else:
                                        st.error("Недостаточно данных для создания прогноза. Нужно минимум 2 точки данных.")
                            else:
                                st.warning("Выберите метрику для прогнозирования")
                    
                    else:
                        st.info("📊 Загрузите данные в первой вкладке для создания прогнозов")

# end of file
